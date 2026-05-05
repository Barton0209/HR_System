#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Converter - Конвертация Excel файлов между форматами XLSB и XLSX
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import threading
import shutil
import time
import gc
import traceback
from typing import Tuple, Optional, List

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# Попытка импорта pyxlsb
PYXLSB_AVAILABLE = False
try:
    from pyxlsb import open_workbook as open_xlsb
    PYXLSB_AVAILABLE = True
except ImportError:
    pass

# Попытка импорта win32com
WIN32COM_AVAILABLE = False
EXCEL_AVAILABLE = False
pythoncom = None
try:
    import win32com.client
    import pythoncom
    WIN32COM_AVAILABLE = True
    try:
        excel_test = win32com.client.Dispatch("Excel.Application")
        excel_test.Quit()
        del excel_test
        EXCEL_AVAILABLE = True
    except Exception:
        pass
except ImportError:
    pass


class ExcelConverterApp:
    """Главный класс приложения"""

    SUPPORTED_EXTENSIONS = ('.xlsx', '.xlsb')

    COM_FILE_FORMATS = {
        '.xlsx': 51,
        '.xlsb': 50,
    }

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Excel Converter - Конвертация XLSB <-> XLSX")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)

        self.files_list: List[str] = []
        self.stop_flag = False
        self.processing_thread: Optional[threading.Thread] = None

        # Переменные UI
        self.output_folder = tk.StringVar()
        self.conversion_mode = tk.StringVar(value="xlsb_to_xlsx")
        self.create_backup = tk.BooleanVar(value=False)
        self.use_excel_com = tk.BooleanVar(value=EXCEL_AVAILABLE)
        self.delay_between_files = tk.DoubleVar(value=0.5)

        self.setup_styles()
        self.setup_ui()

    def setup_styles(self) -> None:
        style = ttk.Style()
        style.configure('Header.TLabel', font=('Segoe UI', 12, 'bold'))
        style.configure('Subheader.TLabel', font=('Segoe UI', 9))
        style.configure('Error.TLabel', foreground='red')
        style.configure('Warning.TLabel', foreground='orange')
        style.configure('Success.TLabel', foreground='green')
        style.configure('Info.TLabel', foreground='gray')

    def setup_ui(self) -> None:
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1) # Лог теперь row 3

        # === ЗАГОЛОВОК ===
        header_frame = ttk.Frame(main_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        header_frame.columnconfigure(0, weight=1)
        
        ttk.Label(header_frame, text="Excel Converter", style='Header.TLabel').grid(row=0, column=0, sticky=tk.W)
        ttk.Label(header_frame, text="Массовая конвертация файлов между форматами XLSB и XLSX", style='Subheader.TLabel').grid(row=1, column=0, sticky=tk.W)

        # === БЛОК РЕЖИМА КОНВЕРТАЦИИ ===
        mode_frame = ttk.LabelFrame(main_frame, text=" Режим конвертации ", padding="10")
        mode_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        ttk.Radiobutton(mode_frame, text="XLSB  ➔  XLSX", variable=self.conversion_mode, value="xlsb_to_xlsx").pack(side=tk.LEFT, padx=(0, 30))
        ttk.Radiobutton(mode_frame, text="XLSX  ➔  XLSB", variable=self.conversion_mode, value="xlsx_to_xlsb").pack(side=tk.LEFT)

        # === ВКЛАДКИ ===
        notebook = ttk.Notebook(main_frame)
        notebook.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))

        files_tab = ttk.Frame(notebook, padding="10")
        notebook.add(files_tab, text=" Файлы ")
        files_tab.columnconfigure(0, weight=1)
        files_tab.rowconfigure(1, weight=1)

        extra_tab = ttk.Frame(notebook, padding="10")
        notebook.add(extra_tab, text=" Настройки ")
        extra_tab.columnconfigure(0, weight=1)

        self._setup_files_tab(files_tab)
        self._setup_settings_tab(extra_tab)

        # === КНОПКИ ДЕЙСТВИЙ И ПРОГРЕСС ===
        action_frame = ttk.Frame(main_frame)
        action_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10)) # Сдвинуто на row 3
        action_frame.columnconfigure(0, weight=1)

        self.progress = ttk.Progressbar(action_frame, mode='determinate')
        self.progress.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))

        self.status_label = ttk.Label(action_frame, text="Готово к работе")
        self.status_label.grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(5, 0))

        btn_frame = ttk.Frame(action_frame)
        btn_frame.grid(row=0, column=1, rowspan=2)
        
        self.start_btn = ttk.Button(btn_frame, text="▶ Запустить", command=self.start_processing)
        self.start_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.stop_btn = ttk.Button(btn_frame, text="⏹ Остановить", command=self.stop_processing, state='disabled')
        self.stop_btn.pack(side=tk.LEFT)

        # === ЛОГ ===
        log_frame = ttk.LabelFrame(main_frame, text=" Журнал операций ", padding="5")
        log_frame.grid(row=4, column=0, sticky=(tk.W, tk.E, tk.N, tk.S)) # Сдвинуто на row 4
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, state='disabled', wrap=tk.WORD, font=('Consolas', 9), bg='#1e1e1e', fg='#d4d4d4', insertbackground='white')
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.log_text.tag_configure('error', foreground='#f44747')
        self.log_text.tag_configure('success', foreground='#6a9955')
        self.log_text.tag_configure('warning', foreground='#cca700')
        self.log_text.tag_configure('info', foreground='#569cd6')

    def _setup_files_tab(self, parent: ttk.Frame) -> None:
        files_btn_frame = ttk.Frame(parent)
        files_btn_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        ttk.Button(files_btn_frame, text="Добавить файлы", command=self.add_files).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(files_btn_frame, text="Добавить папку", command=self.add_folder).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(files_btn_frame, text="Очистить", command=self.clear_files).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(files_btn_frame, text="Удалить выбранные", command=self.remove_selected).pack(side=tk.LEFT)

        list_frame = ttk.Frame(parent)
        list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)

        columns = ('filename', 'path', 'size', 'status')
        self.files_tree = ttk.Treeview(list_frame, columns=columns, show='headings', selectmode='extended', height=8)
        self.files_tree.heading('filename', text='Имя файла')
        self.files_tree.heading('path', text='Путь')
        self.files_tree.heading('size', text='Размер')
        self.files_tree.heading('status', text='Статус')
        self.files_tree.column('filename', width=180, minwidth=120)
        self.files_tree.column('path', width=350, minwidth=200)
        self.files_tree.column('size', width=80, minwidth=60)
        self.files_tree.column('status', width=120, minwidth=80)

        vsb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.files_tree.yview)
        hsb = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.files_tree.xview)
        self.files_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.files_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))

        output_frame = ttk.LabelFrame(parent, text="Папка для сохранения", padding="5")
        output_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        output_frame.columnconfigure(0, weight=1)
        
        entry_frame = ttk.Frame(output_frame)
        entry_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E))
        entry_frame.columnconfigure(0, weight=1)
        
        ttk.Entry(entry_frame, textvariable=self.output_folder).grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))
        ttk.Button(entry_frame, text="Обзор...", command=self.browse_output_folder).grid(row=0, column=1)
        
        ttk.Radiobutton(output_frame, text="Сохранять в исходную папку (с суффиксом _converted)", variable=self.output_folder, value="").grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))

    def _setup_settings_tab(self, parent: ttk.Frame) -> None:
        options_frame = ttk.LabelFrame(parent, text="Опции сохранения", padding="10")
        options_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Checkbutton(options_frame, text="Создать резервные копии (.backup)", variable=self.create_backup).pack(anchor=tk.W)

        delay_frame = ttk.LabelFrame(parent, text="Производительность", padding="10")
        delay_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(delay_frame, text="Задержка между файлами (сек):").pack(anchor=tk.W)
        ttk.Scale(delay_frame, from_=0.1, to=3.0, variable=self.delay_between_files, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(5, 0))
        ttk.Label(delay_frame, text="Увеличьте при ошибках доступа к файлам", style='Info.TLabel').pack(anchor=tk.W)

        xlsb_frame = ttk.LabelFrame(parent, text="Движок конвертации", padding="10")
        xlsb_frame.pack(fill=tk.X, pady=(0, 10))

        if EXCEL_AVAILABLE:
            ttk.Checkbutton(xlsb_frame, text="Использовать Excel COM (рекомендуется, сохраняет форматирование)", variable=self.use_excel_com).pack(anchor=tk.W)
            ttk.Label(xlsb_frame, text="✓ Excel найден - полное сохранение форматирования", style='Success.TLabel').pack(anchor=tk.W, pady=(5, 0))
        elif WIN32COM_AVAILABLE:
            ttk.Label(xlsb_frame, text="⚠ pywin32 установлен, но Microsoft Excel не найден", style='Warning.TLabel').pack(anchor=tk.W)
            self.use_excel_com.set(False)
        else:
            ttk.Label(xlsb_frame, text="✗ Компоненты Excel COM не установлены", style='Error.TLabel').pack(anchor=tk.W)
            self.use_excel_com.set(False)

        if PYXLSB_AVAILABLE:
            ttk.Label(xlsb_frame, text="ℹ pyxlsb доступен (для XLSB ➔ XLSX только данные без оформления)", style='Info.TLabel').pack(anchor=tk.W, pady=(5, 0))
        else:
            ttk.Label(xlsb_frame, text="⚠ pyxlsb не установлен (для XLSB ➔ XLSX без Excel он нужен)", style='Warning.TLabel').pack(anchor=tk.W)

    def add_files(self) -> None:
        files = filedialog.askopenfilenames(
            title="Выберите Excel файлы",
            filetypes=[("Excel Binary/XML", "*.xlsx *.xlsb"), ("Все файлы", "*.*")]
        )
        for file in files:
            self.add_file_to_list(file)

    def add_folder(self) -> None:
        folder = filedialog.askdirectory(title="Выберите папку с Excel файлами")
        if folder:
            count = 0
            for root_dir, _, files in os.walk(folder):
                for file in files:
                    if file.lower().endswith(self.SUPPORTED_EXTENSIONS):
                        full_path = os.path.join(root_dir, file)
                        self.add_file_to_list(full_path)
                        count += 1
            if count > 0:
                self.log(f"Добавлено {count} файлов из папки: {folder}", 'info')
            else:
                self.log(f"В папке не найдено XLSX/XLSB файлов: {folder}", 'warning')

    def add_file_to_list(self, file_path: str) -> None:
        normalized_path = os.path.normpath(file_path)
        for item in self.files_tree.get_children():
            existing_path = os.path.normpath(self.files_tree.item(item)['values'][1])
            if existing_path == normalized_path:
                return
        try:
            size = os.path.getsize(file_path)
            size_str = self.format_size(size)
            filename = os.path.basename(file_path)
            self.files_tree.insert('', tk.END, values=(filename, file_path, size_str, 'В очереди'))
            self.files_list.append(file_path)
        except OSError as e:
            self.log(f"Ошибка при добавлении файла {file_path}: {e}", 'error')

    def format_size(self, size_bytes: int) -> str:
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"

    def clear_files(self) -> None:
        for item in self.files_tree.get_children():
            self.files_tree.delete(item)
        self.files_list.clear()

    def remove_selected(self) -> None:
        selected = self.files_tree.selection()
        if not selected:
            return
        paths_to_remove = [self.files_tree.item(item)['values'][1] for item in selected]
        for item in selected:
            self.files_tree.delete(item)
        for path in paths_to_remove:
            if path in self.files_list:
                self.files_list.remove(path)

    def browse_output_folder(self) -> None:
        folder = filedialog.askdirectory(title="Выберите папку для сохранения")
        if folder:
            self.output_folder.set(folder)

    def log(self, message: str, tag: str = None) -> None:
        def _log():
            self.log_text.configure(state='normal')
            if tag:
                self.log_text.insert(tk.END, f"{message}\n", tag)
            else:
                self.log_text.insert(tk.END, f"{message}\n")
            self.log_text.see(tk.END)
            self.log_text.configure(state='disabled')
        
        if threading.current_thread() is threading.main_thread():
            _log()
        else:
            self.root.after(0, _log)

    def process_file_with_excel_com(self, file_path: str, output_path: str) -> Tuple[bool, str]:
        excel = None
        wb = None
        try:
            if pythoncom:
                pythoncom.CoInitialize()
            
            file_path = os.path.abspath(file_path)
            output_path = os.path.abspath(output_path)
            
            if os.path.exists(output_path):
                max_retries = 3
                for retry in range(max_retries):
                    try:
                        os.remove(output_path)
                        break
                    except PermissionError:
                        if retry < max_retries - 1:
                            time.sleep(1)
                        else:
                            return False, f"Не удалось удалить существующий файл: {output_path}"

            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.EnableEvents = False
            excel.AskToUpdateLinks = False
            
            wb = excel.Workbooks.Open(file_path, 0, True, None, None, None, True)

            ext = os.path.splitext(output_path)[1].lower()
            file_format = self.COM_FILE_FORMATS.get(ext, 51)

            wb.SaveAs(output_path, file_format, None, None, False, False, 0, 2)
            wb.Close(SaveChanges=False)
            wb = None

            time.sleep(self.delay_between_files.get())
            return True, f"[COM] {os.path.basename(file_path)} -> {os.path.basename(output_path)}"

        except Exception as e:
            error_msg = str(e)
            if 'xlmmain11.chm' in error_msg:
                error_msg = "Ошибка справки Excel (файл сохранен)"
            return False, f"Ошибка Excel COM: {error_msg}"
            
        finally:
            if wb:
                try: wb.Close(SaveChanges=False)
                except Exception: pass
                wb = None
            if excel:
                try: excel.Quit()
                except Exception: pass
                excel = None
            if pythoncom:
                try: pythoncom.CoUninitialize()
                except Exception: pass
            gc.collect()

    def convert_xlsb_to_xlsx_pyxlsb(self, xlsb_path: str, output_path: str) -> Tuple[bool, str]:
        if not PYXLSB_AVAILABLE:
            return False, "pyxlsb не установлен"
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            with open_xlsb(xlsb_path) as xlsb:
                for sheet_name in xlsb.sheets:
                    ws_xlsb = xlsb.get_sheet(sheet_name)
                    ws = wb.create_sheet(title=sheet_name)
                    for row_idx, row in enumerate(ws_xlsb.rows(), 1):
                        for col_idx, cell in enumerate(row, 1):
                            if cell.v is not None:
                                ws.cell(row=row_idx, column=col_idx, value=cell.v)
            wb.save(output_path)
            wb.close()
            return True, f"[pyxlsb] {os.path.basename(xlsb_path)} -> {os.path.basename(output_path)}"
        except Exception as e:
            return False, f"Ошибка pyxlsb: {e}"

    def process_file(self, file_path: str, output_folder: Optional[str]) -> Tuple[bool, str]:
        try:
            filename = os.path.basename(file_path)
            base_name, ext = os.path.splitext(filename)
            ext_lower = ext.lower()
            mode = self.conversion_mode.get()

            # Валидация входного файла
            if mode == "xlsb_to_xlsx" and ext_lower != '.xlsb':
                return False, "Пропущен (ожидается .xlsb)"
            if mode == "xlsx_to_xlsb" and ext_lower != '.xlsx':
                return False, "Пропущен (ожидается .xlsx)"

            # Определение выходного пути
            target_ext = ".xlsx" if mode == "xlsb_to_xlsx" else ".xlsb"
            output_filename = f"{base_name}_converted{target_ext}"

            if output_folder:
                os.makedirs(output_folder, exist_ok=True)
                output_path = os.path.join(output_folder, output_filename)
            else:
                src_dir = os.path.dirname(file_path)
                output_path = os.path.join(src_dir, output_filename)

            if self.create_backup.get():
                try:
                    shutil.copy2(file_path, f"{file_path}.backup")
                except Exception as e:
                    return False, f"Ошибка создания бэкапа: {e}"

            # Логика конвертации
            if mode == "xlsx_to_xlsb" and not self.use_excel_com.get():
                return False, "Ошибка: Для XLSX ➔ XLSB необходим Excel COM (openpyxl не умеет сохранять .xlsb)"

            if self.use_excel_com.get() and EXCEL_AVAILABLE:
                return self.process_file_with_excel_com(file_path, output_path)
            else:
                if mode == "xlsb_to_xlsx":
                    return self.convert_xlsb_to_xlsx_pyxlsb(file_path, output_path)
                else:
                    return False, "Ошибка: Движок не поддерживает XLSX ➔ XLSB"

        except InvalidFileException as e:
            return False, f"Ошибка формата: {e}"
        except PermissionError as e:
            return False, f"Нет доступа: {e}"
        except Exception as e:
            return False, f"Ошибка: {type(e).__name__}: {e}"

    def processing_worker(self) -> None:
        output_folder = self.output_folder.get() if self.output_folder.get() else None
        total = len(self.files_list)
        processed = 0
        success_count = 0
        error_count = 0
        skipped_count = 0

        self.root.after(0, lambda: self.start_btn.configure(state='disabled'))
        self.root.after(0, lambda: self.stop_btn.configure(state='normal'))

        mode_text = "XLSB ➔ XLSX" if self.conversion_mode.get() == "xlsb_to_xlsx" else "XLSX ➔ XLSB"
        engine_text = "Excel COM" if (EXCEL_AVAILABLE and self.use_excel_com.get()) else "openpyxl/pyxlsb"

        self.log("=" * 60, 'info')
        self.log(f"Начало обработки {total} файлов...", 'info')
        self.log(f"Режим: {mode_text} | Движок: {engine_text}", 'info')
        self.log("=" * 60, 'info')

        tree_items = self.files_tree.get_children()

        for i, file_path in enumerate(self.files_list):
            if self.stop_flag:
                self.log("Остановлено пользователем", 'warning')
                break

            if i >= len(tree_items):
                break

            item = tree_items[i]
            current_values = self.files_tree.item(item)['values']
            
            def update_status(vals, status_text):
                try:
                    self.files_tree.item(item, values=(vals[0], vals[1], vals[2], status_text))
                except tk.TclError:
                    pass

            self.root.after(0, update_status, current_values, '⏳ Обработка...')

            success, message = self.process_file(file_path, output_folder)
            
            if "Пропущен" in message:
                status = '⏭ Пропущен'
                skipped_count += 1
                self.root.after(0, update_status, current_values, status)
                self.log(f"[{i+1}/{total}] {os.path.basename(file_path)}: {message}", 'warning')
            else:
                status = '✓ OK' if success else '✗ Ошибка'
                if success:
                    success_count += 1
                else:
                    error_count += 1
                self.root.after(0, update_status, current_values, status)
                self.log(f"[{i+1}/{total}] {os.path.basename(file_path)}: {message}", 'success' if success else 'error')

            processed += 1
            progress_value = (processed / total) * 100
            
            def update_progress(val, txt):
                self.progress['value'] = val
                self.status_label['text'] = txt
                
            self.root.after(0, update_progress, progress_value, f"Обработано {processed} из {total}")

            if not self.stop_flag and i < total - 1:
                time.sleep(self.delay_between_files.get())

        self.log("=" * 60, 'info')
        self.log(f"Готово! Успешно: {success_count}, Пропущено: {skipped_count}, Ошибок: {error_count}", 'success' if error_count == 0 else 'warning')
        self.log("=" * 60, 'info')

        self.root.after(0, update_progress, 0, f"Готово. Успешно: {success_count}, Ошибок: {error_count}")

        self.root.after(0, lambda: self.start_btn.configure(state='normal'))
        self.root.after(0, lambda: self.stop_btn.configure(state='disabled'))

        if success_count > 0 and not self.stop_flag:
            def show_result():
                if error_count == 0:
                    messagebox.showinfo("Готово", f"Конвертация завершена успешно!\nОбработано: {success_count}")
                else:
                    messagebox.showwarning("Завершено с ошибками", f"Успешно: {success_count}\nОшибок: {error_count}\n\nПодробности в журнале.")
            self.root.after(100, show_result)

    def start_processing(self) -> None:
        if not self.files_list:
            messagebox.showwarning("Внимание", "Список файлов пуст!")
            return

        self.stop_flag = False
        self.processing_thread = threading.Thread(target=self.processing_worker, daemon=True)
        self.processing_thread.start()

    def stop_processing(self) -> None:
        if not self.stop_flag:
            self.stop_flag = True
            self.log("Остановка... (будет завершен текущий файл)", 'warning')


def main():
    try:
        root = tk.Tk()
        app = ExcelConverterApp(root)
        root.mainloop()
    except Exception as e:
        traceback.print_exc()
        try:
            root_err = tk.Tk()
            root_err.withdraw()
            messagebox.showerror("Критическая ошибка", f"Приложение завершилось с ошибкой:\n\n{str(e)}\n\nПодробности в консоли.")
            root_err.destroy()
        except:
            pass

if __name__ == "__main__":
    main()