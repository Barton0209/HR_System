"""
Auth Router - Аутентификация и авторизация пользователей
Данные загружаются из Excel_files/ПАРОЛЬ_ДОСТУП.xlsx
"""
import os
import hashlib
import logging
from pathlib import Path
from typing import Optional, Dict, List
from datetime import datetime, timedelta

from fastapi import APIRouter, HTTPException, Depends, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel

from hr_app.backend.database import get_conn, set_setting, get_setting

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/auth", tags=["auth"])

# Путь к файлу с паролями
PASSWORDS_FILE = Path(os.getenv("PASSWORDS_FILE", "Excel_files/ПАРОЛЬ_ДОСТУП.xlsx"))

# Хранилище пользователей в памяти (загружается из Excel)
_users_db: Dict[str, dict] = {}

# Секретный ключ для JWT (в production использовать env variable)
SECRET_KEY = os.getenv("SECRET_KEY", "hr_system_secret_key_change_in_production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 8  # 8 часов

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")


class UserLogin(BaseModel):
    username: str
    password: str


class UserInfo(BaseModel):
    login: str
    fio: str
    email: str
    position: str
    department: str
    access_level: str  # ДОСТУП - уровень доступа к информации по столбцу Базы AD - Площадка_ЕЖУ
    carnet_access: str  # Доступ к Карнет


def _hash_password(password: str) -> str:
    """Хеширование пароля."""
    return hashlib.sha256(password.encode()).hexdigest()


def load_users_from_excel():
    """Загрузка пользователей из Excel файла ПАРОЛЬ_ДОСТУП.xlsx"""
    global _users_db
    _users_db = {}
    
    if not PASSWORDS_FILE.exists():
        logger.warning(f"Файл с паролями не найден: {PASSWORDS_FILE}")
        return False
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(PASSWORDS_FILE), read_only=True, data_only=True)
        
        if 'ПАРОЛЬ_ДОСТУП' not in wb.sheetnames:
            logger.error(f"Лист 'ПАРОЛЬ_ДОСТУП' не найден в файле")
            wb.close()
            return False
        
        ws = wb['ПАРОЛЬ_ДОСТУП']
        
        # Читаем заголовки
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip() if h else '' for h in row]
            break
        
        # Находим индексы колонок
        col_map = {}
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if 'логин' in h_lower:
                col_map['login'] = i
            elif 'пароль' in h_lower:
                col_map['password'] = i
            elif 'доступ' in h_lower and 'карнет' not in h_lower:
                col_map['access'] = i
            elif h_lower == 'фио':
                col_map['fio'] = i
            elif 'email' in h_lower:
                col_map['email'] = i
            elif 'должность' in h_lower:
                col_map['position'] = i
            elif 'отдел' in h_lower:
                col_map['department'] = i
            elif 'карнет' in h_lower:
                col_map['carnet'] = i
        
        # Проверяем обязательные колонки
        required = ['login', 'password', 'fio']
        missing = [r for r in required if r not in col_map]
        if missing:
            logger.error(f"Отсутствуют обязательные колонки: {missing}")
            wb.close()
            return False
        
        # Читаем данные
        count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None for v in row):
                continue
            
            login = row[col_map['login']] if col_map['login'] < len(row) else None
            if not login:
                continue
            
            login = str(login).strip()
            password = row[col_map['password']] if col_map['password'] < len(row) else ''
            
            user_data = {
                'login': login,
                'password_hash': _hash_password(str(password)),
                'fio': str(row[col_map.get('fio', 3)]) if col_map.get('fio') and col_map['fio'] < len(row) else '',
                'email': str(row[col_map.get('email', 4)]) if col_map.get('email') and col_map['email'] < len(row) else '',
                'position': str(row[col_map.get('position', 5)]) if col_map.get('position') and col_map['position'] < len(row) else '',
                'department': str(row[col_map.get('department', 6)]) if col_map.get('department') and col_map['department'] < len(row) else '',
                'access_level': str(row[col_map.get('access', 2)]) if col_map.get('access') and col_map['access'] < len(row) else '',
                'carnet_access': str(row[col_map.get('carnet', 7)]) if col_map.get('carnet') and col_map['carnet'] < len(row) else 'Просмотр',
            }
            
            _users_db[login] = user_data
            count += 1
        
        wb.close()
        logger.info(f"Загружено {count} пользователей из {PASSWORDS_FILE}")
        return True
        
    except Exception as e:
        logger.exception(f"Ошибка загрузки пользователей из Excel: {e}")
        return False


def get_user_from_db(login: str) -> Optional[dict]:
    """Получение пользователя из базы."""
    return _users_db.get(login)


async def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    """Получение текущего пользователя из токена (упрощённая версия)."""
    # В реальной реализации здесь будет декодирование JWT
    # Для простоты используем логин как токен
    user = get_user_from_db(token)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Неверные учетные данные",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return user


@router.on_event("startup")
async def on_startup():
    """Загрузка пользователей при старте."""
    load_users_from_excel()


@router.post("/reload-users")
async def reload_users():
    """Перезагрузка пользователей из Excel файла."""
    success = load_users_from_excel()
    if success:
        return {"ok": True, "message": f"Загружено {len(_users_db)} пользователей"}
    else:
        raise HTTPException(500, "Ошибка загрузки пользователей")


@router.post("/login")
async def login(form_data: OAuth2PasswordRequestForm = Depends()):
    """Вход в систему."""
    user = get_user_from_db(form_data.username)
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Неверное имя пользователя или пароль",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Проверяем пароль
    password_hash = _hash_password(form_data.password)
    if user['password_hash'] != password_hash:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Неверное имя пользователя или пароль",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Возвращаем токен (в упрощённом виде - просто логин)
    # В production нужно использовать JWT
    access_token = user['login']  # Упрощённо - в реальности JWT
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "login": user['login'],
            "fio": user['fio'],
            "email": user['email'],
            "position": user['position'],
            "department": user['department'],
            "access_level": user['access_level'],
            "carnet_access": user['carnet_access'],
        }
    }


@router.get("/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    """Получение информации о текущем пользователе."""
    return {
        "login": current_user['login'],
        "fio": current_user['fio'],
        "email": current_user['email'],
        "position": current_user['position'],
        "department": current_user['department'],
        "access_level": current_user['access_level'],
        "carnet_access": current_user['carnet_access'],
    }


@router.get("/users")
async def list_users():
    """Список всех пользователей (для администрирования)."""
    return {
        "users": [
            {
                "login": u['login'],
                "fio": u['fio'],
                "email": u['email'],
                "position": u['position'],
                "department": u['department'],
                "access_level": u['access_level'],
                "carnet_access": u['carnet_access'],
            }
            for u in _users_db.values()
        ]
    }
