"""Карнет — объединение Excel-карнетов (веб-версия MergeCarnetTab)"""
import os
import json
import tempfile
import logging
from pathlib import Path
from typing import List, Dict, Any

from fastapi import APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import StreamingResponse

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/carnet", tags=["carnet"])

REQUIRED_FIELDS = [
    "Табельный номер", "Фактическая должность", "Участок по факту", "Прораб",
    "Позиция", "ДЕНЬ/НОЧЬ", "да/нет", "ОП/Проект", "Оценка", "ФИО",
    "Должность (по утвержденном списку в параметрах)",
    "Фактическое структурное подразделение (по утвержденном списку в параметрах)",
    "Гражданство", "Должность", "Разряд", "Подразделение",
    "Сотрудник официально трудоустроен на проекте (в 1с Территория): (данные на последний день в месяце)",
    "Удостоверение Серия", "Удостоверение Номер", "Компания", "ИТР",
    "Сектор", "Вид работ", "Виза/ гражданство",
] + [str(i) for i in range(1, 32)] + [
    "ПРИМЕЧАНИЕ", "Итого произв. Часов", "Итого актираных часов"
]


def _read_headers(tmp_path: str, header_row: int = 1) -> List[Dict]:
    """Читает листы и заголовки из Excel файла."""
    from openpyxl import load_workbook

    ext = Path(tmp_path).suffix.lower()
    sheets = []

    if ext in ('.xlsx', '.xlsm', '.xlsb'):
        wb = load_workbook(tmp_path, data_only=True, read_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            max_col = ws.max_column or 50
            headers = []
            for col in range(1, min(max_col, 200) + 1):
                cell = ws.cell(row=header_row, column=col)
                val = cell.value
                headers.append(str(val).strip() if val is not None else f"Столбец {col}")
            state = getattr(ws, 'sheet_state', 'visible')
            sheets.append({
                "name": name,
                "headers": headers,
                "hidden": state != 'visible',
                "header_row": header_row,
                "max_col": len(headers),
                "selected": not (state != 'visible'),
            })
        wb.close()

    elif ext == '.xls':
        try:
            import xlrd
            wb = xlrd.open_workbook(tmp_path)
            for name in wb.sheet_names():
                ws = wb.sheet_by_name(name)
                headers = []
                if header_row <= ws.nrows:
                    for i, v in enumerate(ws.row_values(header_row - 1)):
                        headers.append(str(v).strip() if v else f"Столбец {i + 1}")
                sheets.append({
                    "name": name,
                    "headers": headers,
                    "hidden": False,
                    "header_row": header_row,
                    "max_col": len(headers),
                    "selected": True,
                })
        except ImportError:
            raise HTTPException(500, "xlrd не установлен — .xls не поддерживается")

    return sheets


@router.post("/scan")
async def scan_files(files: List[UploadFile] = File(...)):
    """Загружает файлы и возвращает их листы + заголовки."""
    results = []
    for upload in files:
        suffix = Path(upload.filename or "file.xlsx").suffix.lower()
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(await upload.read())
            tmp_path = tmp.name
        try:
            sheets = _read_headers(tmp_path)
            results.append({
                "name": upload.filename or "file",
                "type": suffix.lstrip("."),
                "sheets": sheets,
            })
        except Exception as e:
            logger.error("Scan error %s: %s", upload.filename, e)
            results.append({
                "name": upload.filename or "file",
                "type": suffix.lstrip("."),
                "sheets": [],
                "error": str(e),
            })
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
    return {"files": results, "required_fields": REQUIRED_FIELDS}


@router.post("/reload-headers")
async def reload_headers(
    file: UploadFile = File(...),
    header_row: int = Form(1),
    sheet_name: str = Form(""),
):
    """Перечитывает заголовки одного листа при смене строки заголовка."""
    suffix = Path(file.filename or "file.xlsx").suffix.lower()
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        sheets = _read_headers(tmp_path, header_row)
        target = next((s for s in sheets if s["name"] == sheet_name), sheets[0] if sheets else None)
        return {"headers": target["headers"] if target else [], "max_col": target["max_col"] if target else 0}
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


@router.post("/merge")
async def merge_carnets(
    files: List[UploadFile] = File(...),
    mapping_json: str = Form(...),
):
    """
    Объединяет карнеты.
    mapping_json: [{filename, sheet, header_row, columns:[{src_header, target_field}]}]
    """
    from openpyxl import Workbook, load_workbook
    from urllib.parse import quote

    try:
        mapping: List[Dict] = json.loads(mapping_json)
    except Exception:
        raise HTTPException(400, "Неверный формат mapping_json")

    file_map: Dict[str, str] = {}
    tmp_files: List[str] = []
    for upload in files:
        suffix = Path(upload.filename or "file.xlsx").suffix.lower()
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(await upload.read())
            tmp_files.append(tmp.name)
            file_map[upload.filename or ""] = tmp.name

    try:
        seen = dict()
        target_cols: List[str] = []
        for entry in mapping:
            for col in entry.get("columns", []):
                tf = col.get("target_field", "").strip()
                if tf and tf not in seen:
                    seen[tf] = True
                    target_cols.append(tf)
        if not target_cols:
            target_cols = REQUIRED_FIELDS[:]

        all_rows: List[Dict] = []
        log_lines: List[str] = []

        for entry in mapping:
            fname = entry.get("filename", "")
            sheet_name = entry.get("sheet", "")
            header_row = int(entry.get("header_row", 1))
            columns = entry.get("columns", [])

            tmp_path = file_map.get(fname)
            if not tmp_path or not Path(tmp_path).exists():
                log_lines.append(f"Файл не найден: {fname}")
                continue

            col_map = {c["src_header"]: c["target_field"] for c in columns if c.get("target_field")}

            try:
                wb = load_workbook(tmp_path, data_only=True, read_only=True)
                if sheet_name not in wb.sheetnames:
                    log_lines.append(f"Лист не найден: {fname}/{sheet_name}")
                    wb.close()
                    continue

                ws = wb[sheet_name]
                hdr_vals = []
                for col in range(1, (ws.max_column or 200) + 1):
                    cell = ws.cell(row=header_row, column=col)
                    hdr_vals.append(str(cell.value).strip() if cell.value is not None else "")

                idx_to_target = {
                    ci: col_map[hdr]
                    for ci, hdr in enumerate(hdr_vals)
                    if hdr in col_map
                }

                row_count = 0
                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    row_dict: Dict[str, str] = {}
                    for ci, val in enumerate(row):
                        if ci in idx_to_target:
                            row_dict[idx_to_target[ci]] = str(val) if val is not None else ""
                    if any(v for v in row_dict.values()):
                        all_rows.append(row_dict)
                        row_count += 1

                wb.close()
                log_lines.append(f"✅ {fname} / {sheet_name}: {row_count} строк")
            except Exception as e:
                logger.error("Merge error %s/%s: %s", fname, sheet_name, e)
                log_lines.append(f"❌ {fname}/{sheet_name}: {e}")

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Итоговый карнет"
        out_ws.append(target_cols)
        for row_dict in all_rows:
            out_ws.append([row_dict.get(f, "") for f in target_cols])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as out_tmp:
            out_path = out_tmp.name
        out_wb.save(out_path)

        def file_iter():
            with open(out_path, "rb") as f:
                yield from f
            try:
                os.unlink(out_path)
            except Exception:
                pass

        fname_enc = quote("Итоговый_карнет.xlsx", safe="")
        return StreamingResponse(
            file_iter(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"carnet_merged.xlsx\"; filename*=UTF-8''{fname_enc}",
                "X-Merge-Log": json.dumps(log_lines, ensure_ascii=False)[:1000],
                "X-Row-Count": str(len(all_rows)),
            },
        )
    finally:
        for tp in tmp_files:
            try:
                os.unlink(tp)
            except Exception:
                pass
