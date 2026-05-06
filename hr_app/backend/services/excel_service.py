"""
Excel Service — загрузка и обработка Excel файлов
"""
import re
import logging
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime

import pandas as pd
import numpy as np

from hr_app.backend.database import (
    upsert_employees, log_action, save_ticket_orders, get_setting, get_conn
)

logger = logging.getLogger(__name__)


def safe_str(val, default="") -> str:
    if val is None:
        return default
    try:
        if pd.isna(val):
            return default
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    return default if s in ("nan", "None", "NaT", "") else s


def safe_date(val) -> str:
    """Приводит дату к формату ДД.ММ.ГГГГ."""
    s = safe_str(val)
    if not s:
        return ""
    # Already formatted
    if re.match(r'^\d{2}\.\d{2}\.\d{4}$', s):
        return s
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d",
                "%d-%m-%Y", "%Y-%m-%dT%H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:10], fmt[:len(fmt)]).strftime("%d.%m.%Y")
        except ValueError:
            pass
    m = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{4})', s)
    if m:
        d, mo, y = m.groups()
        return f"{d.zfill(2)}.{mo.zfill(2)}.{y}"
    return s


# Column index → field name (0-based, columns A=0, B=1, ...)
# Based on the БАЗА.xlsx description
_COL_BY_POS = {
    0:  "unique1",        # A
    1:  "unique2",        # B
    2:  "tab_num",        # C
    3:  "org",            # D
    4:  "territory",      # E
    5:  "fio",            # F
    6:  "citizenship",    # G
    7:  "birth_date",     # H
    8:  "doc_series",     # I
    9:  "doc_num",        # J
    10: "position",       # K
    11: "grade",          # L
    12: "department",     # M
    13: "section",        # N
    14: "section2",       # O
    15: "work_schedule",  # P
    16: "hire_date",      # Q
    17: "fire_date",      # R
    18: "status",         # S
    19: "work_start_date",# T
    20: "birth_place",    # U
    21: "doc_issuer",     # V
    22: "doc_issue_date", # W
    23: "address",        # X
    24: "phone_home",     # Y
    25: "phone_mobile",   # Z
    26: "phone_work",     # AA
    27: "total",          # AB
    28: "region_eju",     # AC
    29: "platform_eju",   # AD
    30: "position_eju",   # AE
    31: "section_eju",    # AF
    32: "section2_eju",   # AG
    33: "visa_eju",       # AH
    34: "visa_type_eju",  # AI
    35: "visa_region_eju",# AJ
    36: "visa_expire_eju",# AK
    37: "shift_start_eju",# AL
    38: "status_op",      # AM
    39: "subdivision_blt",# AN  (placeholder, filled from settings)
    40: "classification",  # AO
    41: "dept_category",  # AP
    42: "doc_type",       # AQ
}

_NAME_MAP = {
    "фио+датарождения_unique1": "unique1",
    "объединенный_номер_unique2": "unique2",
    "табельный_номер_unique3": "tab_num",
    "организация_1с": "org",
    "территория_1с": "territory",
    "фио_1с": "fio",
    "гражданство_1с": "citizenship",
    "дата_рождения_1с": "birth_date",
    "удостоверениесерия_1с": "doc_series",
    "удостоверениесерия": "doc_series",
    "удостоверениеномер_1с": "doc_num",
    "удостоверениеномер": "doc_num",
    "должность_1с": "position",
    "разряд_1с": "grade",
    "подразделение_1с": "department",
    "участок_отдел_1с": "section",
    "участок_отдел2_1с": "section2",
    "график_работы_1с": "work_schedule",
    "дата_приема_1с": "hire_date",
    "дата_увольнения_1с": "fire_date",
    "статус_сотрудника_1с": "status",
    "сотрудникдата_выхода_на_работу_1с": "work_start_date",
    "место_рождения_1с": "birth_place",
    "удостоверениекем_выдан_1с": "doc_issuer",
    "удостоверениедата выдачи_1с": "doc_issue_date",
    "удостоверениедата_выдачи_1с": "doc_issue_date",
    "физическое_лицоадрес_по_прописке_1с": "address",
    "физическое_лицодомашний_телефон_1с": "phone_home",
    "физическое_лицоличный_мобильный_телефон_1с": "phone_mobile",
    "физическое_лицорабочий_телефон_1с": "phone_work",
    "итого_1с": "total",
    "регион_ежу": "region_eju",
    "площадка_ежу": "platform_eju",
    "фактическая_должность_ежу": "position_eju",
    "фактический_участок_отдел_ежу": "section_eju",
    "фактический_участок_отдел2_ежу": "section2_eju",
    "виза_ежу": "visa_eju",
    "вид_визы_ежу": "visa_type_eju",
    "регион_визы_ежу": "visa_region_eju",
    "срок_до_ежу": "visa_expire_eju",
    "дата_начала_вахты_прием_на_работу_ежу": "shift_start_eju",
    "статус_оп": "status_op",
    "статус_ оп": "status_op",
    "подразделение_блт": "subdivision_blt",
    "классификация_сотрудников_(итр_или_рабочие)_блт": "classification",
    "классификация_сотрудников_итр_или_рабочие_блт": "classification",
    "отдел_(сму,_умит,_отиз)_и_т.д_блт": "dept_category",
    "вид_документа_блт": "doc_type",
}


def _normalize_col(col: str) -> str:
    """Нормализует имя столбца для сопоставления."""
    col = str(col).strip().lower()
    col = col.replace(" ", "_").replace(".", "").replace(",", "").replace("(", "").replace(")", "")
    return col


def _map_df_columns(df: pd.DataFrame) -> Dict[str, str]:
    """Returns mapping: df_column_name -> field_name"""
    mapping = {}
    for col in df.columns:
        normalized = _normalize_col(col)
        if normalized in _NAME_MAP:
            mapping[col] = _NAME_MAP[normalized]
    return mapping


def load_main_base(file_path: str) -> Tuple[bool, str, int]:
    """Загрузка основной базы БАЗА.xlsx лист База_1С"""
    try:
        logger.info("Loading main base from %s", file_path)
        xl = pd.ExcelFile(file_path)

        # Find the correct sheet
        sheet = None
        for candidate in ["База_1С", "База_1С ", "ВСЕ ОП", "Sheet1", "Лист1"]:
            if candidate in xl.sheet_names:
                sheet = candidate
                break
        if sheet is None:
            sheet = xl.sheet_names[0]

        df = pd.read_excel(file_path, sheet_name=sheet, header=0, dtype=str)
        df.columns = df.columns.astype(str)

        # Map columns by name
        col_map = _map_df_columns(df)

        # Fallback: map by position
        for pos_idx, field in _COL_BY_POS.items():
            if pos_idx < len(df.columns):
                col = df.columns[pos_idx]
                if col not in col_map and field not in col_map.values():
                    col_map[col] = field

        # Rename
        df = df.rename(columns=col_map)

        # Keep only known fields
        known_fields = set(_COL_BY_POS.values())
        df = df[[c for c in df.columns if c in known_fields]]

        # Ensure fio exists
        if "fio" not in df.columns:
            return False, "Не найден столбец ФИО (F / ФИО_1С)", 0

        # Clean up
        for col in df.columns:
            df[col] = df[col].apply(safe_str)

        # Date columns
        for date_col in ["birth_date", "hire_date", "fire_date", "doc_issue_date",
                         "work_start_date", "shift_start_eju", "visa_expire_eju"]:
            if date_col in df.columns:
                df[date_col] = df[date_col].apply(safe_date)

        # Filter empty FIO
        df = df[df["fio"].notna() & (df["fio"] != "")]

        records = df.to_dict("records")
        count = upsert_employees(records)
        log_action("load_main_base", os.path.basename(file_path), count, "ok")
        return True, f"Загружено {count} сотрудников (лист: {sheet})", count

    except Exception as e:
        logger.exception("Error loading main base")
        log_action("load_main_base", os.path.basename(file_path), 0, "error", str(e))
        return False, f"Ошибка загрузки: {e}", 0


def load_total_experience(file_path: str) -> Tuple[bool, str, int]:
    """Загрузка общего стажа из файла ОБЩИЙ_СТАЖ.xlsx
    
    Ожидает колонки:
    - Табельный_номер_UNIQUE3 (или类似)
    - Общий_стаж
    - Стаж_за_период
    """
    try:
        logger.info("Loading total experience from %s", file_path)
        
        # Try to read with openpyxl engine for better compatibility
        df = pd.read_excel(file_path, engine='openpyxl', dtype=str)
        df.columns = df.columns.astype(str).str.strip()
        
        # Normalize column names
        col_map = {}
        for col in df.columns:
            col_lower = col.lower().replace(' ', '_').replace('-', '_')
            if 'табельный' in col_lower or 'tab_num' in col_lower or 'unique3' in col_lower:
                col_map[col] = 'tab_num'
            elif 'общий_стаж' in col_lower or 'total_experience' in col_lower or col_lower == 'общий_стаж':
                col_map[col] = 'total_experience'
            elif 'стаж_за_период' in col_lower or 'period_experience' in col_lower:
                col_map[col] = 'experience_period'
        
        # If no mapping found, try positional mapping
        if not col_map:
            # Assume first column is tab_num, second is total_experience
            if len(df.columns) >= 2:
                col_map[df.columns[0]] = 'tab_num'
                col_map[df.columns[1]] = 'total_experience'
        
        df = df.rename(columns=col_map)
        
        # Ensure we have tab_num
        if 'tab_num' not in df.columns:
            return False, "Не найдена колонка с табельным номером", 0
        
        # Clean data
        for col in df.columns:
            df[col] = df[col].apply(lambda x: safe_str(x) if pd.notna(x) else "")
        
        # Filter rows with tab_num
        df = df[df['tab_num'].notna() & (df['tab_num'] != "")]
        
        # Update employees table with experience data
        updated_count = 0
        with get_conn() as conn:
            for _, row in df.iterrows():
                tab_num = row.get('tab_num', '')
                if not tab_num:
                    continue
                    
                total_exp = row.get('total_experience', '')
                exp_period = row.get('experience_period', '')
                
                # Check if employee exists
                existing = conn.execute(
                    "SELECT id FROM employees WHERE tab_num=?", (tab_num,)
                ).fetchone()
                
                if existing:
                    # Update existing employee
                    if total_exp or exp_period:
                        updates = []
                        params = []
                        if total_exp:
                            updates.append("total=?")
                            params.append(total_exp)
                        if exp_period:
                            updates.append("extra_json=json_insert(coalesce(extra_json, '{}'), '$.experience_period', ?)")
                            params.append(exp_period)
                        params.append(tab_num)
                        
                        if updates:
                            conn.execute(
                                f"UPDATE employees SET {','.join(updates)} WHERE tab_num=?",
                                params
                            )
                            updated_count += 1
                else:
                    # Create new employee record with minimal data
                    conn.execute(
                        """INSERT INTO employees (tab_num, total, extra_json)
                           VALUES (?, ?, json_object('experience_period', ?))""",
                        (tab_num, total_exp, exp_period if exp_period else "")
                    )
                    updated_count += 1
        
        log_action("load_total_experience", os.path.basename(file_path), updated_count, "ok")
        return True, f"Обновлен стаж для {updated_count} сотрудников", updated_count

    except Exception as e:
        logger.exception("Error loading total experience")
        log_action("load_total_experience", os.path.basename(file_path), 0, "error", str(e))
        return False, f"Ошибка загрузки стажа: {e}", 0


def load_daily_tracking_files(folder_path: str, track_date: str) -> Tuple[bool, str, int]:
    """
    Загружает файлы Ежедневного учёта из папки.
    Ищет все *.xls* файлы, читает лист 'ЕЖЕДНЕВНЫЙ УЧЕТ',
    фильтрует красные строки, объединяет данные.
    """
    from hr_app.backend.database import get_conn
    import openpyxl

    folder = Path(folder_path)
    if not folder.exists():
        return False, f"Папка не найдена: {folder_path}", 0

    files = list(folder.glob("*.xls*")) + list(folder.glob("*.xlsb"))
    if not files:
        return False, "Файлы Excel не найдены в папке", 0

    RED_COLORS = {
        (255, 0, 0), (255, 199, 206)
    }

    all_rows = []
    processed = 0
    skipped = 0

    for fp in files:
        try:
            wb = openpyxl.load_workbook(str(fp), read_only=False, data_only=True)
            sheet_found = None
            for sname in wb.sheetnames:
                if sname.strip().upper() == "ЕЖЕДНЕВНЫЙ УЧЕТ":
                    sheet_found = sname
                    break

            if not sheet_found:
                skipped += 1
                continue

            ws = wb[sheet_found]
            rows_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=False), start=6):
                # Check cell E (index 4) not empty
                cell_e = row[4] if len(row) > 4 else None
                if cell_e is None or cell_e.value is None or str(cell_e.value).strip() == "":
                    break

                # Check fill color (red = skip)
                skip_row = False
                for cell in row[:16]:
                    fill = cell.fill
                    if fill and fill.fgColor and fill.fgColor.type == "rgb":
                        rgb_hex = fill.fgColor.rgb
                        if len(rgb_hex) == 8:
                            r = int(rgb_hex[2:4], 16)
                            g = int(rgb_hex[4:6], 16)
                            b = int(rgb_hex[6:8], 16)
                            if (r, g, b) in RED_COLORS or (r == 255 and g == 0 and b == 0):
                                skip_row = True
                                break
                if skip_row:
                    continue

                vals = [cell.value for cell in row[:16]]
                rows_data.append(vals)

            wb.close()

            # Map columns: A=0 region, B=1 platform, C=2 fio(?), D=3 tab_num, J=9 position...
            for vals in rows_data:
                while len(vals) < 16:
                    vals.append(None)
                all_rows.append({
                    "track_date": track_date,
                    "source_file": fp.name,
                    "region": safe_str(vals[1]),      # B
                    "platform": safe_str(vals[2]),    # C
                    "tab_num": safe_str(vals[3]),     # D
                    "fio": safe_str(vals[4]),         # E
                    "position": safe_str(vals[9]),    # J
                    "section": safe_str(vals[10]),    # K
                    "visa": safe_str(vals[11]),       # L
                    "visa_type": safe_str(vals[12]),  # M
                    "visa_region": safe_str(vals[13]),# N
                    "visa_expire": safe_str(vals[14]),# O
                    "shift_start": safe_str(vals[15]),# P
                    "status": "",
                    "extra_json": "",
                })
            processed += 1

        except Exception as e:
            logger.warning("Error processing %s: %s", fp.name, e)
            skipped += 1

    if all_rows:
        with get_conn() as conn:
            conn.execute(
                "DELETE FROM daily_tracking WHERE track_date=?", (track_date,)
            )
            conn.executemany(
                """INSERT INTO daily_tracking
                   (track_date,source_file,region,platform,tab_num,fio,position,
                    section,visa,visa_type,visa_region,visa_expire,shift_start,status)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [(r["track_date"], r["source_file"], r["region"], r["platform"],
                  r["tab_num"], r["fio"], r["position"], r["section"],
                  r["visa"], r["visa_type"], r["visa_region"], r["visa_expire"],
                  r["shift_start"], r["status"]) for r in all_rows]
            )

    msg = f"Обработано файлов: {processed}, пропущено: {skipped}, строк загружено: {len(all_rows)}"
    log_action("load_daily_tracking", folder_path, len(all_rows), "ok", msg)
    return True, msg, len(all_rows)


def load_ticket_costs(file_paths: List[str]) -> Tuple[bool, str, int]:
    """Загружает реестры по затратам на билеты."""
    from hr_app.backend.database import get_conn
    total_rows = 0
    for fp in file_paths:
        try:
            df = pd.read_excel(fp, header=0, dtype=str)
            df.columns = df.columns.astype(str).str.strip()
            # Generic load - keep all rows
            rows = []
            for _, row in df.iterrows():
                vals = [safe_str(v) for v in row.values]
                padded = (vals + [""] * 10)[:10]
                rows.append(padded)
            with get_conn() as conn:
                conn.executemany(
                    """INSERT INTO ticket_costs
                       (source_file,tab_num,fio,route,flight_date,ticket_num,amount,payment,org,department,note)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    [(os.path.basename(fp), *r) for r in rows]
                )
            total_rows += len(rows)
        except Exception as e:
            logger.warning("Error loading ticket costs from %s: %s", fp, e)
    return True, f"Загружено {total_rows} записей о затратах", total_rows


def export_ticket_orders_excel(orders: List[Dict], output_path: str) -> bool:
    """Экспорт заявок на билеты в Excel (формат как ОП __ Заявка)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Заявка"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        columns = [
            "№", "Подразделение", "Отдел", "Операция", "Классификация", "Дата заказа",
            "Организация", "Ф.И.О.", "Ф.И.О лат", "Табельный номер", "Гражданство",
            "Дата рождения", "Вид документа", "Серия", "Номер", "Дата выдачи",
            "Дата окончания", "Кем выдан", "Адрес", "Маршрут", "Обоснование",
            "ПС", "АВИА/ЖД", "Дата вылета", "Примечание", "Ответственный",
            "Дата выписки", "Билет", "Сумма", "Оплата", "Причина возврата",
            "Последний перелет", "Телефон", "Трансфер",
        ]
        db_cols = [
            "num", "department", "section_dept", "operation", "classification", "order_date",
            "org", "fio", "fio_lat", "tab_num", "citizenship", "birth_date", "doc_type",
            "doc_series", "doc_num", "doc_issue_date", "doc_expire_date", "doc_issuer",
            "address", "route", "reason", "ps", "transport_type", "flight_date", "note",
            "responsible", "issue_date", "ticket", "amount", "payment", "return_reason",
            "last_flight", "phone", "transfer",
        ]

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row_idx, order in enumerate(orders, 2):
            for col_idx, db_col in enumerate(db_cols, 1):
                val = order.get(db_col, "")
                if val is None:
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center")

        # Auto width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)
        ws.row_dimensions[1].height = 30

        wb.save(output_path)
        return True
    except Exception as e:
        logger.exception("Error exporting to Excel")
        return False


# ============================================================================
# ЗАГРУЗКА ПАРОЛЕЙ ДОСТУПА И СПРАВОЧНИКОВ
# ============================================================================

def load_password_access(file_path: str) -> Tuple[bool, str, int]:
    """Загрузка ПАРОЛЬ_ДОСТУП.xlsx - лист ПАРОЛЬ_ДОСТУП
    Колонки: A-Логин, B-Пароль, C-ДОСТУП (Площадка_ЕЖУ), D-ФИО, E-Email, F-Должность, G-Отдел, H-Доступ к Карнет
    """
    try:
        logger.info("Loading password access from %s", file_path)
        
        xl = pd.ExcelFile(file_path)
        sheet = "ПАРОЛЬ_ДОСТУП" if "ПАРОЛЬ_ДОСТУП" in xl.sheet_names else xl.sheet_names[0]
        
        df = pd.read_excel(file_path, sheet_name=sheet, header=0, dtype=str)
        df.columns = df.columns.astype(str).str.strip()
        
        # Map columns
        col_map = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            if 'логин' in col_lower:
                col_map[col] = 'login'
            elif 'пароль' in col_lower:
                col_map[col] = 'password'
            elif 'доступ' in col_lower and ('площадк' in col_lower or 'ежу' in col_lower or col_lower == 'доступ'):
                col_map[col] = 'access_platform'
            elif 'фио' in col_lower:
                col_map[col] = 'fio'
            elif 'email' in col_lower:
                col_map[col] = 'email'
            elif 'должност' in col_lower:
                col_map[col] = 'position'
            elif 'отдел' in col_lower:
                col_map[col] = 'department'
            elif 'карнет' in col_lower:
                col_map[col] = 'carnet_access'
        
        # Fallback: positional mapping
        if not col_map and len(df.columns) >= 8:
            cols = list(df.columns)[:8]
            col_map = {
                cols[0]: 'login',
                cols[1]: 'password',
                cols[2]: 'access_platform',
                cols[3]: 'fio',
                cols[4]: 'email',
                cols[5]: 'position',
                cols[6]: 'department',
                cols[7]: 'carnet_access',
            }
        
        df = df.rename(columns=col_map)
        
        # Clean data
        for col in df.columns:
            df[col] = df[col].apply(lambda x: safe_str(x) if pd.notna(x) else "")
        
        # Filter rows with login
        df = df[df['login'].notna() & (df['login'] != "")]
        
        # Store in database table for user access control
        count = 0
        with get_conn() as conn:
            # Create table if not exists
            conn.execute("""
                CREATE TABLE IF NOT EXISTS user_access (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    login TEXT UNIQUE,
                    password TEXT,
                    access_platform TEXT,
                    fio TEXT,
                    email TEXT,
                    position TEXT,
                    department TEXT,
                    carnet_access TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            for _, row in df.iterrows():
                try:
                    conn.execute("""
                        INSERT OR REPLACE INTO user_access 
                        (login, password, access_platform, fio, email, position, department, carnet_access)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        row.get('login', ''),
                        row.get('password', ''),
                        row.get('access_platform', ''),
                        row.get('fio', ''),
                        row.get('email', ''),
                        row.get('position', ''),
                        row.get('department', ''),
                        row.get('carnet_access', ''),
                    ))
                    count += 1
                except Exception as e:
                    logger.warning("Error inserting user %s: %s", row.get('login'), e)
        
        log_action("load_password_access", os.path.basename(file_path), count, "ok")
        return True, f"Загружено {count} пользователей", count
        
    except Exception as e:
        logger.exception("Error loading password access")
        log_action("load_password_access", os.path.basename(file_path), 0, "error", str(e))
        return False, f"Ошибка загрузки: {e}", 0


def load_departments(file_path: str) -> Tuple[bool, str, int]:
    """Загрузка Подразделение_Отдел_Участок.xlsx"""
    try:
        logger.info("Loading departments from %s", file_path)
        
        df = pd.read_excel(file_path, dtype=str)
        df.columns = df.columns.astype(str).str.strip()
        
        # Clean data
        for col in df.columns:
            df[col] = df[col].apply(lambda x: safe_str(x) if pd.notna(x) else "")
        
        count = len(df)
        
        # Store as JSON setting
        data = df.to_dict('records')
        set_setting("departments_data", json.dumps(data, ensure_ascii=False))
        
        log_action("load_departments", os.path.basename(file_path), count, "ok")
        return True, f"Загружено {count} записей подразделений", count
        
    except Exception as e:
        logger.exception("Error loading departments")
        log_action("load_departments", os.path.basename(file_path), 0, "error", str(e))
        return False, f"Ошибка загрузки: {e}", 0


def load_areas(file_path: str) -> Tuple[bool, str, int]:
    """Загрузка Терр_ПЛОЩ_ПОДР_СтатусОП_Регион.xlsx"""
    try:
        logger.info("Loading areas from %s", file_path)
        
        df = pd.read_excel(file_path, dtype=str)
        df.columns = df.columns.astype(str).str.strip()
        
        # Clean data
        for col in df.columns:
            df[col] = df[col].apply(lambda x: safe_str(x) if pd.notna(x) else "")
        
        count = len(df)
        
        # Store as JSON setting
        data = df.to_dict('records')
        set_setting("areas_data", json.dumps(data, ensure_ascii=False))
        
        log_action("load_areas", os.path.basename(file_path), count, "ok")
        return True, f"Загружено {count} записей территорий", count
        
    except Exception as e:
        logger.exception("Error loading areas")
        log_action("load_areas", os.path.basename(file_path), 0, "error", str(e))
        return False, f"Ошибка загрузки: {e}", 0


def load_positions(file_path: str) -> Tuple[bool, str, int]:
    """Загрузка Должность, Классификация.xlsx"""
    try:
        logger.info("Loading positions from %s", file_path)
        
        df = pd.read_excel(file_path, dtype=str)
        df.columns = df.columns.astype(str).str.strip()
        
        # Clean data
        for col in df.columns:
            df[col] = df[col].apply(lambda x: safe_str(x) if pd.notna(x) else "")
        
        count = len(df)
        
        # Store as JSON setting
        data = df.to_dict('records')
        set_setting("positions_data", json.dumps(data, ensure_ascii=False))
        
        log_action("load_positions", os.path.basename(file_path), count, "ok")
        return True, f"Загружено {count} записей должностей", count
        
    except Exception as e:
        logger.exception("Error loading positions")
        log_action("load_positions", os.path.basename(file_path), 0, "error", str(e))
        return False, f"Ошибка загрузки: {e}", 0


def generate_total_experience_report(output_path: str) -> Tuple[bool, str, int]:
    """Генерация отчета ОБЩИЙ_СТАЖ.xlsx из основной базы сотрудников"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        with get_conn() as conn:
            employees = conn.execute("""
                SELECT tab_num, fio, position, department, hire_date, fire_date, total
                FROM employees
                WHERE tab_num IS NOT NULL AND tab_num != ''
                ORDER BY tab_num
            """).fetchall()
        
        if not employees:
            return False, "Нет данных для генерации отчета", 0
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ОБЩИЙ_СТАЖ"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        headers = ["Табельный номер", "ФИО", "Должность", "Подразделение", 
                   "Дата приема", "Дата увольнения", "Общий стаж"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for row_idx, emp in enumerate(employees, 2):
            for col_idx, val in enumerate(emp, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val if val else "")
                cell.border = border
                cell.font = Font(name="Calibri", size=10)
        
        # Auto width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)
        
        wb.save(output_path)
        
        log_action("generate_total_experience_report", output_path, len(employees), "ok")
        return True, f"Сгенерирован отчет для {len(employees)} сотрудников", len(employees)
        
    except Exception as e:
        logger.exception("Error generating total experience report")
        log_action("generate_total_experience_report", output_path, 0, "error", str(e))
        return False, f"Ошибка генерации: {e}", 0


def generate_ticket_costs_report(output_path: str) -> Tuple[bool, str, int]:
    """Генерация Реестр_по_затратам_на_билеты.xlsx из загруженных реестров"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        with get_conn() as conn:
            costs = conn.execute("""
                SELECT source_file, tab_num, fio, route, flight_date, 
                       ticket_num, amount, payment, org, department, note
                FROM ticket_costs
                ORDER BY flight_date DESC, tab_num
            """).fetchall()
        
        if not costs:
            return False, "Нет данных о затратах на билеты", 0
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Затраты на билеты"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        headers = ["Источник", "Таб.номер", "ФИО", "Маршрут", "Дата вылета",
                   "Номер билета", "Сумма", "Оплата", "Организация", "Подразделение", "Примечание"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for row_idx, cost in enumerate(costs, 2):
            for col_idx, val in enumerate(cost, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val if val else "")
                cell.border = border
                cell.font = Font(name="Calibri", size=10)
        
        # Auto width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)
        
        wb.save(output_path)
        
        log_action("generate_ticket_costs_report", output_path, len(costs), "ok")
        return True, f"Сгенерирован отчет по {len(costs)} записям", len(costs)
        
    except Exception as e:
        logger.exception("Error generating ticket costs report")
        log_action("generate_ticket_costs_report", output_path, 0, "error", str(e))
        return False, f"Ошибка генерации: {e}", 0
