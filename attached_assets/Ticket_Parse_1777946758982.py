#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ticket Parser Application - Только для авиабилетов
Парсинг PDF билетов с извлечением данных о рейсах
"""

import sys
import os
import re
import json
import logging
import io
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict

# Excel handling
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PyQt5
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QComboBox, QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox, QProgressBar, QGroupBox, QGridLayout,
    QHeaderView, QAbstractItemView, QCheckBox, QTextEdit, QSplitter, 
    QTabWidget, QStatusBar, QToolBar, QAction, QMenu
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QRunnable, QThreadPool, QObject, QSettings
from PyQt5.QtGui import QPixmap, QIcon, QFont, QKeySequence

# PDF support
try:
    import fitz  # PyMuPDF
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# Image support (для сканов билетов)
try:
    from PIL import Image
    import pytesseract
    IMAGE_SUPPORT = True
except ImportError:
    IMAGE_SUPPORT = False

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ticket_parser.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class TicketData:
    """Структура данных билета"""
    passenger: str = 'Не указано'
    ticket_number: str = 'Не указано'
    order_number: str = 'Не указано'
    issue_date: str = 'Не указано'
    carrier: str = 'Не указано'
    flight_number: str = 'Не указано'
    departure_time: str = 'Не указано'
    departure_date: str = 'Не указано'
    route: str = 'Не указано'
    arrival_airport: str = 'Не указано'
    arrival_time: str = 'Не указано'
    arrival_date: str = 'Не указано'
    total_price: str = 'Не указано'
    currency: str = 'RUB'
    source_file: str = ''
    
    def to_dict(self) -> Dict[str, str]:
        return {
            'Пассажир': self.passenger,
            'Номер билета': self.ticket_number,
            'Номер заказа': self.order_number,
            'Дата выдачи': self.issue_date,
            'Перевозчик': self.carrier,
            'Номер рейса': self.flight_number,
            'Время отправления': self.departure_time,
            'Дата отправления': self.departure_date,
            'Маршрут': self.route,
            'Аэропорт прибытия': self.arrival_airport,
            'Время прибытия': self.arrival_time,
            'Дата прибытия': self.arrival_date,
            'Стоимость': self.total_price,
            'Валюта': self.currency,
            'Источник': self.source_file
        }


class TicketParser:
    """Парсер авиабилетов из PDF и изображений - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
    
    # IATA коды аэропортов (как в оригинале)
    IATA_CODES = {
        'SVO': 'МОСКВА', 'DME': 'МОСКВА', 'VKO': 'МОСКВА', 'ZIA': 'МОСКВА',
        'LED': 'САНКТ-ПЕТЕРБУРГ', 'AER': 'СОЧИ', 'KRR': 'КРАСНОДАР',
        'ROV': 'РОСТОВ-НА-ДОНУ', 'KZN': 'КАЗАНЬ', 'UFA': 'УФА',
        'KUF': 'САМАРА', 'GOJ': 'НИЖНИЙ НОВГОРОД', 'SVX': 'ЕКАТЕРИНБУРГ',
        'CEK': 'ЧЕЛЯБИНСК', 'TJM': 'ТЮМЕНЬ', 'OVB': 'НОВОСИБИРСК',
        'OMS': 'ОМСК', 'KJA': 'КРАСНОЯРСК', 'IKT': 'ИРКУТСК',
        'TAS': 'ТАШКЕНТ', 'TMJ': 'ТЕРМЕЗ', 'BHK': 'БУХАРА',
        'SKD': 'САМАРКАНД', 'NVI': 'НАВОИ', 'NCU': 'НУКУС',
        'UGC': 'УРГЕНЧ', 'FEG': 'ФЕРГАНА', 'KSQ': 'КАРШИ',
        'AZN': 'АНДИЖАН', 'NMA': 'НАМАНГАН', 'ALA': 'АЛМАТЫ',
        'NQZ': 'АСТАНА', 'CIT': 'ШЫМКЕНТ', 'MSQ': 'МИНСК',
        'EVN': 'ЕРЕВАН', 'GYD': 'БАКУ', 'TBS': 'ТБИЛИСИ',
        'FRU': 'БИШКЕК', 'DYU': 'ДУШАНБЕ', 'LBD': 'ХУДЖАНД',
        'IST': 'СТАМБУЛ', 'SAW': 'СТАМБУЛ', 'AYT': 'АНТАЛЬЯ',
        'DXB': 'ДУБАЙ', 'AUH': 'АБУ-ДАБИ'
    }
    
    # Месяцы для парсинга дат (как в оригинале)
    MONTH_MAP = {
        'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04', 'MAY': '05', 'JUN': '06',
        'JUL': '07', 'AUG': '08', 'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12',
        'ЯНВ': '01', 'ФЕВ': '02', 'МАР': '03', 'АПР': '04', 'МАЙ': '05', 'ИЮН': '06',
        'ИЮЛ': '07', 'АВГ': '08', 'СЕН': '09', 'ОКТ': '10', 'НОЯ': '11', 'ДЕК': '12',
        'MAP': '03'  # Опечатка в оригинале
    }
    
    def __init__(self):
        self.settings = QSettings('TicketParser', 'Settings')
        
    def parse_file(self, file_path: str) -> TicketData:
        """Основной метод парсинга файла"""
        path = Path(file_path)
        filename = path.stem
        
        # Определяем тип файла
        if path.suffix.lower() == '.pdf':
            text = self._extract_text_from_pdf(file_path)
        elif path.suffix.lower() in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
            text = self._extract_text_from_image(file_path)
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {path.suffix}")
        
        return self._parse_text(text, filename)
    
    def _extract_text_from_pdf(self, pdf_path: str) -> str:
        """Извлекает текст из PDF"""
        if not PDF_SUPPORT:
            raise ImportError("PyMuPDF не установлен")
        
        try:
            doc = fitz.open(pdf_path)
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            return text
        except Exception as e:
            logger.error(f"Ошибка чтения PDF {pdf_path}: {e}")
            return ""
    
    def _extract_text_from_image(self, image_path: str) -> str:
        """Извлекает текст из изображения через OCR"""
        if not IMAGE_SUPPORT:
            raise ImportError("PIL или pytesseract не установлены")
        
        try:
            image = Image.open(image_path)
            image = image.convert('L')
            text = pytesseract.image_to_string(image, lang='rus+eng')
            return text
        except Exception as e:
            logger.error(f"Ошибка OCR {image_path}: {e}")
            return ""
    
    def _parse_text(self, text: str, filename: str) -> TicketData:
        """Парсит текст билета - ИСПРАВЛЕННЫЙ МЕТОД"""
        data = TicketData()
        data.source_file = filename
        
        # === 1. ПАССАЖИР (как в оригинале) ===
        data.passenger = self._extract_passenger(filename, text)
        
        # === 2. НОМЕР БИЛЕТА (как в оригинале) ===
        data.ticket_number = self._extract_ticket_number(text)
        
        # === 3. НОМЕР ЗАКАЗА (как в оригинале) ===
        data.order_number = self._extract_order_number(text, filename)
        
        # === 4. ДАТА ВЫДАЧИ (как в оригинале) ===
        data.issue_date = self._extract_issue_date(text)
        
        # === 5. ПЕРЕВОЗЧИК (УНИВЕРСАЛЬНЫЙ МЕТОД) ===
        data.carrier = self._extract_carrier(text)
        
        # === 6. НОМЕР РЕЙСА (как в оригинале) ===
        data.flight_number = self._extract_flight_number(text)
        
        # === 7. ДАТЫ И ВРЕМЯ (ИСПРАВЛЕНО - как в оригинале) ===
        flight_info = self._extract_flight_times(text, data.issue_date)
        data.departure_time = flight_info.get('dep_time', 'Не указано')
        data.departure_date = flight_info.get('dep_date', 'Не указано')
        data.arrival_time = flight_info.get('arr_time', 'Не указано')
        data.arrival_date = flight_info.get('arr_date', 'Не указано')
        
        # === 8. МАРШРУТ (как в оригинале) ===
        route_info = self._extract_route(text)
        data.route = route_info.get('route', 'Не указано')
        data.arrival_airport = route_info.get('arrival', 'Не указано')
        
        # === 9. СТОИМОСТЬ (как в оригинале) ===
        price_info = self._extract_price(text)
        data.total_price = price_info.get('price', 'Не указано')
        data.currency = price_info.get('currency', 'RUB')
        
        return data
    
    def _extract_passenger(self, filename: str, text: str) -> str:
        """Извлекает имя пассажира - как в оригинале"""
        # Из имени файла
        name_match = re.match(r'^([A-Za-zА-Яа-я]+ [A-Za-zА-Яа-я]+)', filename)
        if name_match:
            return name_match.group(1).strip()
        
        # Из текста формат ФАМИЛИЯ/ИМЯ
        name_match = re.search(r'ФАМИЛИЯ:\s*([A-Z]+/[A-Z]+)', text)
        if name_match:
            return name_match.group(1).replace('/', ' ')
        
        return 'Не указано'
    
    def _extract_ticket_number(self, text: str) -> str:
        """Извлекает номер билета - как в оригинале"""
        ticket_match = re.search(r'НОМЕР БИЛЕТА\s*:\s*(\d+\s+\d+)', text)
        if ticket_match:
            return ticket_match.group(1).replace(' ', '')
        return 'Не указано'
    
    def _extract_order_number(self, text: str, filename: str) -> str:
        """Извлекает номер заказа - как в оригинале"""
        # Из имени файла
        order_match = re.search(r'-\s*([A-Z0-9]+)', filename)
        if order_match:
            return order_match.group(1)
        return 'Не указано'
    
    def _extract_issue_date(self, text: str) -> str:
        """Извлекает дату выдачи - как в оригинале"""
        date_issue = re.search(r'ДАТА:\s*(\d{2})([А-ЯA-Z]{3})(\d{2})', text)
        if date_issue:
            day, month_str, year = date_issue.groups()
            month_num = self.MONTH_MAP.get(month_str[:3].upper(), '01')
            return f"{day}.{month_num}.20{year}"
        return 'Не указано'
    
    def _extract_carrier(self, text: str) -> str:
        """Определяет авиакомпанию - УНИВЕРСАЛЬНЫЙ МЕТОД"""
        text_upper = text.upper()
        lines = text.split('\n')
        
        # 1. Ищем по ключевым маркерам в тексте
        carrier_patterns = [
            r'ВЫДАН\s+ОТ\s*[:;]?\s*([A-ZА-Я][A-ZА-Я\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'ISSUED\s+BY\s*[:;]?\s*([A-Z][A-Z\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'CARRIER\s*[:;]?\s*([A-Z][A-Z\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'АВИАКОМПАНИЯ\s*[:;]?\s*([A-ZА-Я][A-ZА-Я\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'OPERATED\s+BY\s*[:;]?\s*([A-Z][A-Z\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'VALIDATING\s+CARRIER\s*[:;]?\s*([A-Z]{2})',
            r'MARKETING\s+CARRIER\s*[:;]?\s*([A-Z]{2})',
        ]
        
        for pattern in carrier_patterns:
            match = re.search(pattern, text_upper)
            if match:
                carrier = match.group(1).strip()
                carrier = re.sub(r'\s+', ' ', carrier)
                if len(carrier) > 2 and len(carrier) < 50:
                    return self._clean_carrier_name(carrier)
        
        # 2. Ищем по коду рейса (IATA код авиакомпании)
        flight_patterns = [
            r'\b([A-Z]{2})\s*[-]?\s*\d{3,4}\b',
            r'FLIGHT\s*[:;]?\s*([A-Z]{2})\s*\d{3,4}',
            r'РЕЙС\s*[:;]?\s*([A-Z]{2})\s*\d{3,4}',
        ]
        
        for pattern in flight_patterns:
            match = re.search(pattern, text_upper)
            if match:
                code = match.group(1)
                if code not in ['TK', 'BG', 'NP', 'PC', 'NDC', 'NUC', 'ROE', 'RUB', 'USD', 'EUR']:
                    full_name = self._find_carrier_name_by_code(code, text)
                    return full_name if full_name else code
        
        # 3. Ищем известные авиакомпании по названию
        known_carriers = [
            'LUFTHANSA', 'AIR FRANCE', 'KLM', 'BRITISH AIRWAYS', 'VIRGIN ATLANTIC',
            'TURKISH AIRLINES', 'PEGASUS AIRLINES', 'SUNEXPRESS', 'CORENDON AIRLINES',
            'AEROFLOT', 'S7 AIRLINES', 'ROSSIYA', 'URAL AIRLINES', 'POBEDA', 'AZIMUT',
            'NORDWIND', 'SMARTAVIA', 'AZUR AIR', 'ROYAL FLIGHT', 'IKAR', 'RED WINGS',
            'BELAVIA', 'AIR MOLDOVA', 'LOT POLISH AIRLINES', 'CZECH AIRLINES',
            'AUSTRIAN AIRLINES', 'SWISS', 'SAS', 'NORWEGIAN', 'FINNAIR',
            'IBERIA', 'TAP PORTUGAL', 'ALITALIA', 'AEGEAN AIRLINES', 'EL AL',
            'EMIRATES', 'ETIHAD AIRWAYS', 'QATAR AIRWAYS', 'FLY DUBAI', 'AIR ARABIA',
            'SAUDI ARABIAN AIRLINES', 'KUWAIT AIRWAYS', 'GULF AIR', 'OMAN AIR',
            'SINGAPORE AIRLINES', 'CATHAY PACIFIC', 'THAI AIRWAYS', 'MALAYSIA AIRLINES',
            'GARUDA INDONESIA', 'PHILIPPINE AIRLINES', 'VIETNAM AIRLINES',
            'JAPAN AIRLINES', 'ALL NIPPON AIRWAYS', 'KOREAN AIR', 'ASIANA AIRLINES',
            'CHINA SOUTHERN', 'CHINA EASTERN', 'AIR CHINA', 'HAINAN AIRLINES',
            'INDIGO', 'SPICEJET', 'AIR INDIA', 'GO FIRST',
            'UZBEKISTAN AIRWAYS', 'AIR ASTANA', 'SCAT AIRLINES', 'QANOT SHARQ',
            'KYRGYZ AIRLINES', 'TAJIK AIR', 'TURKMENISTAN AIRLINES',
            'AMERICAN AIRLINES', 'DELTA AIR LINES', 'UNITED AIRLINES', 'SOUTHWEST',
            'AIR CANADA', 'WESTJET', 'JETBLUE', 'ALASKA AIRLINES', 'SPIRIT AIRLINES',
            'LATAM', 'GOL', 'AZUL', 'AEROMEXICO', 'COPA AIRLINES', 'AVIANCA',
            'EGYPTAIR', 'ROYAL AIR MAROC', 'ETHIOPIAN AIRLINES', 'SOUTH AFRICAN AIRWAYS',
            'KENYA AIRWAYS', 'QANTAS', 'VIRGIN AUSTRALIA', 'AIR NEW ZEALAND',
            'RYANAIR', 'EASYJET', 'WIZZ AIR', 'VUELING', 'TRANSAVIA', 'EUROWINGS',
            'JET2', 'TUI AIRWAYS', 'NORWEGIAN AIR SHUTTLE',
            'AIR BALTIC', 'UKRAINE INTERNATIONAL AIRLINES', 'GEORGIAN AIRWAYS',
            'ARMENIA AIRWAYS', 'BUTA AIRWAYS', 'IRAN AIR', 'MAHAN AIR',
        ]
        
        known_carriers.sort(key=len, reverse=True)
        
        for carrier in known_carriers:
            if carrier.upper() in text_upper:
                return carrier
        
        # 4. Ищем по шаблону "Xxx Airways/Airlines/Air"
        generic_match = re.search(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\s+(?:AIRWAYS|AIRLINES|AIR|AVIA))\b', text)
        if generic_match:
            return generic_match.group(1)
        
        # 5. Последняя попытка - ищем строку с "air" или "авиа"
        for line in lines:
            line_upper = line.upper().strip()
            if any(word in line_upper for word in ['AIR', 'AVIA', 'АВИА']) and len(line) > 3 and len(line) < 40:
                clean = re.sub(r'[^\w\s&\-]', '', line).strip()
                if len(clean) > 3:
                    return clean
        
        return 'Не указано'

    def _clean_carrier_name(self, name: str) -> str:
        """Очищает название авиакомпании от мусора"""
        remove_words = ['TICKET', 'E-TICKET', 'ETICKET', 'BOOKING', 'PNR', 'FARE', 'TAX', 'TOTAL']
        name_upper = name.upper()
        for word in remove_words:
            name_upper = name_upper.replace(word, '')
        
        name = name_upper.strip()
        name = re.sub(r'\s+', ' ', name)
        name = re.sub(r'\s+[A-Z0-9]$', '', name)
        
        return name if len(name) > 1 else 'Не указано'

    def _find_carrier_name_by_code(self, code: str, text: str) -> Optional[str]:
        """Ищет полное название авиакомпании по IATA коду в тексте"""
        iata_to_name = {
            'HY': 'UZBEKISTAN AIRWAYS', 'KC': 'AIR ASTANA', 'DV': 'SCAT AIRLINES',
            'SU': 'AEROFLOT', 'S7': 'S7 AIRLINES', 'U6': 'URAL AIRLINES',
            'DP': 'POBEDA', 'A4': 'AZIMUT', 'B2': 'BELAVIA', 'TK': 'TURKISH AIRLINES',
            'PC': 'PEGASUS AIRLINES', 'XQ': 'SUNEXPRESS', 'LH': 'LUFTHANSA',
            'AF': 'AIR FRANCE', 'BA': 'BRITISH AIRWAYS', 'EK': 'EMIRATES',
            'EY': 'ETIHAD AIRWAYS', 'QR': 'QATAR AIRWAYS', 'FZ': 'FLY DUBAI',
            'AA': 'AMERICAN AIRLINES', 'DL': 'DELTA AIR LINES', 'UA': 'UNITED AIRLINES',
            'FR': 'RYANAIR', 'U2': 'EASYJET', 'W6': 'WIZZ AIR', 'LX': 'SWISS',
            'OS': 'AUSTRIAN AIRLINES', 'KL': 'KLM', 'VS': 'VIRGIN ATLANTIC',
            'SQ': 'SINGAPORE AIRLINES', 'CX': 'CATHAY PACIFIC', 'JL': 'JAPAN AIRLINES',
            'NH': 'ALL NIPPON AIRWAYS', 'QF': 'QANTAS', 'NZ': 'AIR NEW ZEALAND',
            'SA': 'SOUTH AFRICAN AIRWAYS', 'MS': 'EGYPTAIR', 'ET': 'ETHIOPIAN AIRLINES',
            'AI': 'AIR INDIA', '6E': 'INDIGO', 'SG': 'SPICEJET', 'CZ': 'CHINA SOUTHERN',
            'MU': 'CHINA EASTERN', 'CA': 'AIR CHINA', 'HU': 'HAINAN AIRLINES',
            'TG': 'THAI AIRWAYS', 'MH': 'MALAYSIA AIRLINES', 'GA': 'GARUDA INDONESIA',
            'PR': 'PHILIPPINE AIRLINES', 'VN': 'VIETNAM AIRLINES', 'KE': 'KOREAN AIR',
            'OZ': 'ASIANA AIRLINES', 'AC': 'AIR CANADA', 'WS': 'WESTJET',
            'B6': 'JETBLUE', 'AS': 'ALASKA AIRLINES', 'LA': 'LATAM', 'G3': 'GOL',
            'AD': 'AZUL', 'AM': 'AEROMEXICO', 'CM': 'COPA AIRLINES', 'AV': 'AVIANCA',
            'ME': 'MIDDLE EAST AIRLINES', 'RJ': 'ROYAL JORDANIAN', 'KU': 'KUWAIT AIRWAYS',
            'GF': 'GULF AIR', 'WY': 'OMAN AIR', 'SV': 'SAUDI ARABIAN AIRLINES',
            'AT': 'ROYAL AIR MAROC', 'KM': 'AIR MALTA', 'CY': 'CYPRUS AIRWAYS',
            'A3': 'AEGEAN AIRLINES', 'RO': 'TAROM', 'LO': 'LOT POLISH AIRLINES',
            'OK': 'CZECH AIRLINES', 'JU': 'AIR SERBIA', 'FB': 'BULGARIA AIR',
            'OU': 'CROATIA AIRLINES', 'BT': 'AIR BALTIC', 'PS': 'UKRAINE INTERNATIONAL AIRLINES',
            'A9': 'GEORGIAN AIRWAYS', 'SZ': 'SOMON AIR', '5W': 'WIZZ AIR ABU DHABI',
            '9U': 'AIR MOLDOVA', '7B': 'BEES AIRLINE', 'IU': 'SUPER AIR JET',
        }
        
        if code in iata_to_name:
            return iata_to_name[code]
        
        pattern = rf'\b{code}\b\s*[-:]\s*["\']?([A-Z][A-Za-z\s&]+?)["\']?(?:\n|$)'
        match = re.search(pattern, text)
        if match:
            name = match.group(1).strip()
            if len(name) > 2 and len(name) < 40:
                return name
        
        return None
    
    def _extract_flight_number(self, text: str) -> str:
        """Извлекает номер рейса - как в оригинале"""
        flight_match = re.search(r'\b([A-Z]{2})\s*(\d{3,4})\b', text)
        if flight_match:
            return f"{flight_match.group(1)}{flight_match.group(2)}"
        return 'Не указано'
    
    def _extract_flight_times(self, text: str, issue_date: str) -> Dict[str, str]:
        """Извлекает время и даты вылета/прилета - как в оригинале"""
        result = {
            'dep_time': 'Не указано',
            'dep_date': 'Не указано',
            'arr_time': 'Не указано',
            'arr_date': 'Не указано'
        }
        
        time_match = re.search(r'(\d{2})([А-ЯA-Z]{3})\s+(\d{2})(\d{2})\s+(\d{2})(\d{2})', text)
        
        if time_match:
            day, month_str, dep_h, dep_m, arr_h, arr_m = time_match.groups()
            month_num = self.MONTH_MAP.get(month_str[:3].upper(), '01')
            
            year = '2026'
            if issue_date != 'Не указано':
                year = issue_date.split('.')[-1]
            
            result['dep_time'] = f"{dep_h}:{dep_m}"
            result['arr_time'] = f"{arr_h}:{arr_m}"
            result['dep_date'] = f"{day}.{month_num}.{year}"
            
            try:
                dep_dt = datetime(int(year), int(month_num), int(day), int(dep_h), int(dep_m))
                arr_dt = datetime(int(year), int(month_num), int(day), int(arr_h), int(arr_m))
                
                if arr_dt < dep_dt:
                    arr_dt += timedelta(days=1)
                
                result['arr_date'] = arr_dt.strftime('%d.%m.%Y')
            except Exception as e:
                logger.warning(f"Ошибка расчета даты прилета: {e}")
                result['arr_date'] = result['dep_date']
        
        return result
    
    def _extract_route(self, text: str) -> Dict[str, str]:
        """Определяет маршрут - как в оригинале"""
        result = {'route': 'Не указано', 'arrival': 'Не указано'}
        
        codes = re.findall(r'\b([A-Z]{3})\b', text)
        exclude = {'NUC', 'ROE', 'RUB', 'EUR', 'PC', 'NDC', 'USD', 'END', 'MOW'}
        valid = [c for c in codes if c not in exclude]
        
        dep_code = arr_code = None
        for c in valid:
            if c == 'TMJ':
                dep_code = c
            elif c in ['VKO', 'DME', 'SVO', 'ZIA']:
                arr_code = c
        
        if dep_code and arr_code:
            dep_city = self.IATA_CODES.get(dep_code, dep_code)
            arr_city = self.IATA_CODES.get(arr_code, arr_code)
            result['route'] = f"{dep_city} - {arr_city}"
            result['arrival'] = f"{arr_city} ({arr_code})"
        
        return result
    
    def _extract_price(self, text: str) -> Dict[str, str]:
        """Извлекает стоимость - как в оригинале"""
        result = {'price': 'Не указано', 'currency': 'RUB'}
        
        total_match = re.search(r'ИТОГО\s*:\s*(\d+)\s*РУБ', text, re.I)
        if total_match:
            result['price'] = total_match.group(1)
        else:
            total_match = re.search(r'ИТОГО К ОПЛАТЕ\s*:\s*RUB(\d+)', text)
            if total_match:
                result['price'] = total_match.group(1)
        
        return result


class ProcessingThread(QThread):
    """Поток для обработки файлов без блокировки UI"""
    progress = pyqtSignal(int)
    finished_signal = pyqtSignal(list)
    error_signal = pyqtSignal(str)
    
    def __init__(self, files: List[str], parser: TicketParser):
        super().__init__()
        self.files = files
        self.parser = parser
        self._is_running = True
        
    def run(self):
        results = []
        for i, file_path in enumerate(self.files):
            if not self._is_running:
                break
            
            try:
                data = self.parser.parse_file(file_path)
                results.append(data)
                self.progress.emit(i + 1)
            except Exception as e:
                logger.error(f"Ошибка обработки {file_path}: {e}")
                self.error_signal.emit(f"Ошибка в файле {Path(file_path).name}: {str(e)}")
        
        self.finished_signal.emit(results)
    
    def stop(self):
        self._is_running = False


class TicketParserApp(QMainWindow):
    """Главное окно приложения для парсинга билетов"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Ticket Parser - Парсер авиабилетов')
        self.setGeometry(100, 100, 1600, 900)
        
        self.parser = TicketParser()
        self.tickets: List[TicketData] = []
        self.processing_thread: Optional[ProcessingThread] = None
        
        self.init_ui()
        self.load_settings()
        self.check_dependencies()
    
    def init_ui(self):
        """Инициализация интерфейса"""
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # === ПАНЕЛЬ УПРАВЛЕНИЯ ===
        control_group = QGroupBox('Управление')
        control_layout = QHBoxLayout(control_group)
        
        control_layout.addWidget(QLabel('Папка с билетами:'))
        self.folder_input = QLineEdit()
        self.folder_input.setPlaceholderText('Выберите папку с PDF или изображениями билетов...')
        control_layout.addWidget(self.folder_input, stretch=1)
        
        btn_browse = QPushButton('📁 Обзор')
        btn_browse.clicked.connect(self.select_folder)
        control_layout.addWidget(btn_browse)
        
        btn_scan = QPushButton('🔍 Сканировать')
        btn_scan.clicked.connect(self.scan_folder)
        btn_scan.setStyleSheet('background-color: #4CAF50; color: white; font-weight: bold;')
        control_layout.addWidget(btn_scan)
        
        control_layout.addSpacing(20)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        control_layout.addWidget(self.progress_bar, stretch=1)
        
        layout.addWidget(control_group)
        
        # === ТАБЛИЦА ДАННЫХ ===
        self.table = QTableWidget()
        self.table.setColumnCount(13)
        headers = [
            'Пассажир', 'Номер билета', 'Номер заказа', 'Дата выдачи',
            'Перевозчик', 'Номер рейса', 'Время вылета', 'Дата вылета',
            'Маршрут', 'Аэропорт прибытия', 'Время прилета', 'Дата прилета', 'Стоимость'
        ]
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_table_context_menu)
        layout.addWidget(self.table, stretch=1)
        
        # === НИЖНЯЯ ПАНЕЛЬ ===
        bottom_layout = QHBoxLayout()
        
        stats_group = QGroupBox('Статистика')
        stats_layout = QVBoxLayout(stats_group)
        self.stats_label = QLabel('Билетов: 0')
        self.stats_label.setStyleSheet('font-size: 12px;')
        stats_layout.addWidget(self.stats_label)
        bottom_layout.addWidget(stats_group)
        
        actions_group = QGroupBox('Действия')
        actions_layout = QHBoxLayout(actions_group)
        
        btn_export = QPushButton('📊 Экспорт в Excel')
        btn_export.clicked.connect(self.export_to_excel)
        btn_export.setEnabled(False)
        self.btn_export = btn_export
        actions_layout.addWidget(btn_export)
        
        btn_export_json = QPushButton('📋 Экспорт в JSON')
        btn_export_json.clicked.connect(self.export_to_json)
        btn_export_json.setEnabled(False)
        self.btn_export_json = btn_export_json
        actions_layout.addWidget(btn_export_json)
        
        btn_clear = QPushButton('🗑 Очистить')
        btn_clear.clicked.connect(self.clear_data)
        actions_layout.addWidget(btn_clear)
        
        bottom_layout.addWidget(actions_group, stretch=1)
        
        layout.addLayout(bottom_layout)
        
        self.statusBar().showMessage('Готово к работе')
        self.create_menu()
    
    def create_menu(self):
        """Создает меню приложения"""
        menubar = self.menuBar()
        
        file_menu = menubar.addMenu('Файл')
        
        action_open = QAction('Открыть папку...', self)
        action_open.setShortcut(QKeySequence.Open)
        action_open.triggered.connect(self.select_folder)
        file_menu.addAction(action_open)
        
        action_export = QAction('Экспорт в Excel', self)
        action_export.setShortcut('Ctrl+S')
        action_export.triggered.connect(self.export_to_excel)
        file_menu.addAction(action_export)
        
        file_menu.addSeparator()
        
        action_exit = QAction('Выход', self)
        action_exit.setShortcut('Ctrl+Q')
        action_exit.triggered.connect(self.close)
        file_menu.addAction(action_exit)
        
        tools_menu = menubar.addMenu('Инструменты')
        
        action_settings = QAction('Настройки', self)
        action_settings.triggered.connect(self.show_settings)
        tools_menu.addAction(action_settings)
        
        action_logs = QAction('Открыть логи', self)
        action_logs.triggered.connect(self.open_logs)
        tools_menu.addAction(action_logs)
        
        help_menu = menubar.addMenu('Справка')
        
        action_about = QAction('О программе', self)
        action_about.triggered.connect(self.show_about)
        help_menu.addAction(action_about)
    
    def check_dependencies(self):
        """Проверяет наличие необходимых библиотек"""
        missing = []
        if not PDF_SUPPORT:
            missing.append('PyMuPDF (pip install PyMuPDF)')
        
        if missing:
            QMessageBox.warning(
                self, 'Отсутствуют зависимости',
                f'Следующие библиотеки не установлены:\n\n' + 
                '\n'.join(missing) +
                '\n\nPDF файлы не будут обрабатываться.'
            )
    
    def select_folder(self):
        """Выбор папки с билетами"""
        folder = QFileDialog.getExistingDirectory(
            self, 
            'Выберите папку с билетами',
            self.parser.settings.value('last_folder', '')
        )
        
        if folder:
            self.folder_input.setText(folder)
            self.parser.settings.setValue('last_folder', folder)
            self.scan_folder()
    
    def scan_folder(self):
        """Сканирование папки и обработка файлов"""
        folder = self.folder_input.text()
        if not folder or not Path(folder).exists():
            QMessageBox.warning(self, 'Ошибка', 'Выберите корректную папку!')
            return
        
        path = Path(folder)
        files = []
        
        if PDF_SUPPORT:
            files.extend(path.glob('*.pdf'))
            files.extend(path.glob('*.PDF'))
        
        for ext in ['*.jpg', '*.jpeg', '*.png', '*.bmp', '*.tiff']:
            files.extend(path.glob(ext))
            files.extend(path.glob(ext.upper()))
        
        files = sorted(list(set(files)))
        
        if not files:
            QMessageBox.information(
                self, 'Нет файлов', 
                'В папке не найдено PDF файлов или изображений.'
            )
            return
        
        reply = QMessageBox.question(
            self, 'Найдены файлы',
            f'Найдено файлов: {len(files)}\n\nНачать обработку?',
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.process_files([str(f) for f in files])
    
    def process_files(self, files: List[str]):
        """Обработка файлов в отдельном потоке"""
        self.tickets = []
        self.update_table()
        
        self.progress_bar.setMaximum(len(files))
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        self.progress_bar.setFormat('Обработано: %v/%m (%p%)')
        
        self.processing_thread = ProcessingThread(files, self.parser)
        self.processing_thread.progress.connect(self.on_progress)
        self.processing_thread.finished_signal.connect(self.on_processing_finished)
        self.processing_thread.error_signal.connect(self.on_processing_error)
        self.processing_thread.start()
        
        self.statusBar().showMessage(f'Обработка {len(files)} файлов...')
    
    def on_progress(self, value: int):
        """Обновление прогресса"""
        self.progress_bar.setValue(value)
    
    def on_processing_finished(self, results: List[TicketData]):
        """Завершение обработки"""
        self.tickets = results
        self.progress_bar.setVisible(False)
        self.update_table()
        self.update_stats()
        
        success_count = len([t for t in results if t.ticket_number != 'Не указано'])
        self.statusBar().showMessage(
            f'Обработано: {len(results)} файлов, '
            f'распознано билетов: {success_count}'
        )
        
        if success_count > 0:
            self.btn_export.setEnabled(True)
            self.btn_export_json.setEnabled(True)
        
        QMessageBox.information(
            self, 'Готово',
            f'Обработано файлов: {len(results)}\n'
            f'Успешно распознано: {success_count}'
        )
    
    def on_processing_error(self, error_msg: str):
        """Обработка ошибки"""
        logger.error(error_msg)
        self.statusBar().showMessage(f'Ошибка: {error_msg}', 5000)
    
    def update_table(self):
        """Обновление таблицы данными"""
        self.table.setRowCount(len(self.tickets))
        
        for i, ticket in enumerate(self.tickets):
            data = ticket.to_dict()
            values = [
                data['Пассажир'],
                data['Номер билета'],
                data['Номер заказа'],
                data['Дата выдачи'],
                data['Перевозчик'],
                data['Номер рейса'],
                data['Время отправления'],
                data['Дата отправления'],
                data['Маршрут'],
                data['Аэропорт прибытия'],
                data['Время прибытия'],
                data['Дата прибытия'],
                f"{data['Стоимость']} {data['Валюта']}" if data['Стоимость'] != 'Не указано' else 'Не указано'
            ]
            
            for j, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if value == 'Не указано' or value == '':
                    item.setBackground(Qt.yellow)
                self.table.setItem(i, j, item)
        
        self.table.resizeColumnsToContents()
    
    def update_stats(self):
        """Обновление статистики"""
        total = len(self.tickets)
        with_tickets = len([t for t in self.tickets if t.ticket_number != 'Не указано'])
        with_flights = len([t for t in self.tickets if t.flight_number != 'Не указано'])
        with_times = len([t for t in self.tickets if t.departure_time != 'Не указано'])
        
        stats_text = (
            f'<b>Всего билетов:</b> {total}<br>'
            f'<b>С номером билета:</b> {with_tickets}<br>'
            f'<b>С номером рейса:</b> {with_flights}<br>'
            f'<b>С временем вылета:</b> {with_times}'
        ) if total > 0 else 'Билетов: 0'
        
        self.stats_label.setText(stats_text)
    
    def export_to_excel(self):
        """Экспорт в Excel"""
        if not self.tickets:
            QMessageBox.warning(self, 'Нет данных', 'Нет данных для экспорта!')
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            'Сохранить как',
            f'tickets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            'Excel Files (*.xlsx)'
        )
        
        if not file_path:
            return
        
        try:
            data = [t.to_dict() for t in self.tickets]
            df = pd.DataFrame(data)
            
            column_order = [
                'Пассажир', 'Номер билета', 'Номер заказа', 'Дата выдачи',
                'Перевозчик', 'Номер рейса', 'Маршрут', 'Аэропорт прибытия',
                'Дата отправления', 'Время отправления',
                'Дата прибытия', 'Время прибытия',
                'Стоимость', 'Валюта', 'Источник'
            ]
            df = df[column_order]
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Билеты', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Билеты']
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            self.statusBar().showMessage(f'Сохранено: {file_path}', 5000)
            QMessageBox.information(self, 'Успех', f'Данные сохранены в:\n{file_path}')
            
        except Exception as e:
            logger.error(f"Ошибка экспорта в Excel: {e}")
            QMessageBox.critical(self, 'Ошибка', f'Не удалось сохранить файл:\n{str(e)}')
    
    def export_to_json(self):
        """Экспорт в JSON"""
        if not self.tickets:
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            'Сохранить как',
            f'tickets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json',
            'JSON Files (*.json)'
        )
        
        if not file_path:
            return
        
        try:
            data = [t.to_dict() for t in self.tickets]
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            self.statusBar().showMessage(f'Сохранено: {file_path}', 5000)
            
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось сохранить файл:\n{str(e)}')
    
    def clear_data(self):
        """Очистка всех данных"""
        if self.tickets and QMessageBox.question(
            self, 'Подтверждение',
            'Очистить все данные?',
            QMessageBox.Yes | QMessageBox.No
        ) == QMessageBox.No:
            return
        
        self.tickets = []
        self.table.setRowCount(0)
        self.stats_label.setText('Билетов: 0')
        self.btn_export.setEnabled(False)
        self.btn_export_json.setEnabled(False)
        self.statusBar().showMessage('Данные очищены')
    
    def show_table_context_menu(self, position):
        """Контекстное меню для таблицы"""
        menu = QMenu()
        
        action_copy = QAction('Копировать', self)
        action_copy.triggered.connect(self.copy_selected_cells)
        menu.addAction(action_copy)
        
        action_copy_row = QAction('Копировать строку', self)
        action_copy_row.triggered.connect(self.copy_selected_row)
        menu.addAction(action_copy_row)
        
        menu.exec_(self.table.viewport().mapToGlobal(position))
    
    def copy_selected_cells(self):
        """Копирование выбранных ячеек"""
        selected = self.table.selectedItems()
        if selected:
            text = '\t'.join([item.text() for item in selected])
            QApplication.clipboard().setText(text)
    
    def copy_selected_row(self):
        """Копирование выбранной строки"""
        row = self.table.currentRow()
        if row >= 0:
            values = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                values.append(item.text() if item else '')
            text = '\t'.join(values)
            QApplication.clipboard().setText(text)
    
    def show_settings(self):
        """Диалог настроек"""
        QMessageBox.information(
            self, 'Настройки',
            'Настройки будут доступны в следующей версии.'
        )
    
    def open_logs(self):
        """Открытие файла логов"""
        log_file = Path('ticket_parser.log')
        if log_file.exists():
            import subprocess
            subprocess.Popen(['notepad.exe', str(log_file)])
        else:
            QMessageBox.information(self, 'Логи', 'Файл логов еще не создан')
    
    def show_about(self):
        """О программе"""
        QMessageBox.about(
            self,
            'О программе',
            '<h2>Ticket Parser v1.0</h2>'
            '<p>Парсер авиабилетов из PDF и изображений</p>'
        )
    
    def load_settings(self):
        """Загрузка настроек"""
        last_folder = self.parser.settings.value('last_folder', '')
        if last_folder:
            self.folder_input.setText(last_folder)
    
    def closeEvent(self, event):
        """Обработка закрытия окна"""
        if self.processing_thread and self.processing_thread.isRunning():
            self.processing_thread.stop()
            self.processing_thread.wait(2000)
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    font = QFont('Segoe UI', 10)
    app.setFont(font)
    
    app.setStyleSheet('''
        QGroupBox {
            font-weight: bold;
            border: 1px solid #cccccc;
            border-radius: 5px;
            margin-top: 10px;
            padding-top: 10px;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px;
        }
        QPushButton {
            padding: 5px 15px;
            border-radius: 3px;
        }
        QPushButton:hover {
            background-color: #e0e0e0;
        }
        QTableWidget {
            gridline-color: #cccccc;
        }
        QHeaderView::section {
            background-color: #f0f0f0;
            padding: 5px;
            border: 1px solid #d0d0d0;
            font-weight: bold;
        }
    ''')
    
    window = TicketParserApp()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()