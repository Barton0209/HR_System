#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File Renamer Pro - Отдельное приложение для переименования файлов
PySide6 Edition with Excel Template Support
"""
import sys
import os
from pathlib import Path
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QTableWidget, QTableWidgetItem,
    QFileDialog, QMessageBox, QGroupBox, QCheckBox, QHeaderView,
    QAbstractItemView
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont

# Проверка наличия openpyxl
try:
    import openpyxl
except ImportError:
    print("Ошибка: Библиотека openpyxl не найдена.")
    print("Установите её командой: pip install openpyxl")
    sys.exit(1)

class RenameWorker:
    """Логика переименования"""
    @staticmethod
    def get_files(folder_path: str, recursive: bool = False) -> list:
        """Получить список файлов в папке"""
        files = []
        path = Path(folder_path)
        if not path.exists():
            return files
        
        pattern = "**/*" if recursive else "*"
        for item in path.glob(pattern):
            if item.is_file():
                files.append(item)
        return sorted(files)

    @staticmethod
    def apply_rename(file_path: Path, new_name: str) -> tuple[bool, str]:
        """Переименовать файл"""
        try:
            new_path = file_path.parent / new_name
            if new_path.exists() and new_path != file_path:
                return False, f"Файл уже существует: {new_name}"
            file_path.rename(new_path)
            return True, "Успешно"
        except Exception as e:
            return False, str(e)

class FileRenamerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("File Renamer Pro")
        self.setMinimumSize(1000, 700)
        
        # Список словарей: {'path': Path, 'old_name': str, 'new_name': str, 'status': str}
        self.files_data = [] 
        self.current_folder = ""
        
        self.init_ui()
        self.apply_style()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # --- Верхняя панель: Выбор папки ---
        top_group = QGroupBox("📁 Источник")
        top_layout = QHBoxLayout(top_group)
        
        self.folder_input = QLineEdit()
        self.folder_input.setPlaceholderText("Выберите папку с файлами...")
        self.folder_input.setReadOnly(True)
        top_layout.addWidget(self.folder_input)
        
        btn_browse = QPushButton("📂 Обзор")
        btn_browse.clicked.connect(self.browse_folder)
        top_layout.addWidget(btn_browse)
        
        self.recursive_check = QCheckBox("Включая подпапки")
        top_layout.addWidget(self.recursive_check)
        
        btn_load = QPushButton("🔄 Загрузить файлы")
        btn_load.clicked.connect(self.load_files)
        top_layout.addWidget(btn_load)
        
        main_layout.addWidget(top_group)

        # --- Средняя панель: Инструменты переименования ---
        tools_group = QGroupBox("🛠 Инструменты переименования")
        tools_layout = QVBoxLayout(tools_group)
        
        # Строка 1: Префикс и Суффикс
        row1 = QHBoxLayout()
        self.prefix_input = QLineEdit()
        self.prefix_input.setPlaceholderText("Префикс (в начало)")
        row1.addWidget(QLabel("Префикс:"))
        row1.addWidget(self.prefix_input)
        
        self.suffix_input = QLineEdit()
        self.suffix_input.setPlaceholderText("Суффикс (перед расширением)")
        row1.addWidget(QLabel("Суффикс:"))
        row1.addWidget(self.suffix_input)
        
        btn_add_prefix_suffix = QPushButton("➕ Добавить")
        btn_add_prefix_suffix.clicked.connect(self.add_prefix_suffix)
        row1.addWidget(btn_add_prefix_suffix)
        tools_layout.addLayout(row1)

        # Строка 2: Замена текста
        row2 = QHBoxLayout()
        self.find_text = QLineEdit()
        self.find_text.setPlaceholderText("Найти...")
        row2.addWidget(QLabel("Найти:"))
        row2.addWidget(self.find_text)
        
        self.replace_text = QLineEdit()
        self.replace_text.setPlaceholderText("Заменить на...")
        row2.addWidget(QLabel("Заменить на:"))
        row2.addWidget(self.replace_text)
        
        btn_replace = QPushButton("🔄 Заменить")
        btn_replace.clicked.connect(self.replace_text_in_names)
        row2.addWidget(btn_replace)
        tools_layout.addLayout(row2)

        # Строка 3: Дополнительные опции
        row3 = QHBoxLayout()
        self.lowercase_check = QCheckBox("В нижний регистр")
        row3.addWidget(self.lowercase_check)
        
        self.uppercase_check = QCheckBox("В верхний регистр")
        row3.addWidget(self.uppercase_check)
        
        self.remove_spaces_check = QCheckBox("Удалить пробелы")
        row3.addWidget(self.remove_spaces_check)
        
        btn_apply_format = QPushButton("🎨 Применить формат")
        btn_apply_format.clicked.connect(self.apply_formatting)
        row3.addWidget(btn_apply_format)
        row3.addStretch()
        tools_layout.addLayout(row3)
        
        main_layout.addWidget(tools_group)

        # --- Таблица файлов ---
        table_group = QGroupBox("📄 Список файлов")
        table_layout = QVBoxLayout(table_group)
        
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Старое имя", "Новое имя", "Статус"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        table_layout.addWidget(self.table)
        
        main_layout.addWidget(table_group)

        # --- Нижняя панель: Действия ---
        bottom_layout = QHBoxLayout()
        self.status_label = QLabel("Готов к работе")
        bottom_layout.addWidget(self.status_label)
        bottom_layout.addStretch()
        
        # Кнопки шаблонов
        btn_create_template = QPushButton("📥 Создать шаблон")
        btn_create_template.clicked.connect(self.create_template)
        bottom_layout.addWidget(btn_create_template)
        
        btn_load_template = QPushButton("📤 Загрузить шаблон")
        btn_load_template.clicked.connect(self.load_template)
        bottom_layout.addWidget(btn_load_template)
        
        btn_reset = QPushButton("❌ Сбросить изменения")
        btn_reset.clicked.connect(self.reset_names)
        bottom_layout.addWidget(btn_reset)
        
        btn_rename = QPushButton("✅ ПРИМЕНИТЬ ПЕРЕИМЕНОВАНИЕ")
        btn_rename.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px;")
        btn_rename.clicked.connect(self.execute_rename)
        bottom_layout.addWidget(btn_rename)
        
        main_layout.addLayout(bottom_layout)

    def apply_style(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #f0f0f0; }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ccc;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }
            QPushButton {
                background-color: #0078d7;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton:hover { background-color: #005a9e; }
            QPushButton:pressed { background-color: #004578; }
            QLineEdit {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
            QTableWidget {
                border: 1px solid #ccc;
                gridline-color: #ddd;
            }
            QHeaderView::section {
                background-color: #e0e0e0;
                padding: 5px;
                border: 1px solid #ccc;
            }
        """)

    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку")
        if folder:
            self.folder_input.setText(folder)
            self.current_folder = folder

    def load_files(self):
        folder = self.folder_input.text()
        if not folder or not os.path.exists(folder):
            QMessageBox.warning(self, "Ошибка", "Выберите корректную папку!")
            return
        
        recursive = self.recursive_check.isChecked()
        files = RenameWorker.get_files(folder, recursive)
        
        if not files:
            QMessageBox.information(self, "Информация", "В папке нет файлов.")
            return
        
        self.files_data = []
        for file_path in files:
            self.files_data.append({
                'path': file_path,
                'old_name': file_path.name,
                'new_name': "",  # Изначально пустое новое имя
                'status': ''
            })
        
        self.update_table()
        self.status_label.setText(f"Загружено файлов: {len(self.files_data)}")

    def update_table(self):
        self.table.setRowCount(len(self.files_data))
        for i, data in enumerate(self.files_data):
            # Старое имя (не редактируется)
            old_item = QTableWidgetItem(data['old_name'])
            old_item.setFlags(old_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(i, 0, old_item)
            
            # Новое имя (можно редактировать вручную при необходимости)
            new_item = QTableWidgetItem(data['new_name'])
            # Разрешаем редактирование нового имени пользователем
            new_item.setFlags(new_item.flags() | Qt.ItemIsEditable)
            self.table.setItem(i, 1, new_item)
            
            # Статус (не редактируется)
            status_item = QTableWidgetItem(data['status'])
            status_item.setFlags(status_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(i, 2, status_item)

    def add_prefix_suffix(self):
        prefix = self.prefix_input.text()
        suffix = self.suffix_input.text()
        
        if not prefix and not suffix:
            return
        
        for data in self.files_data:
            # Если new_name пуст, берем old_name для основы
            current_name = data['new_name'] if data['new_name'] else data['old_name']
            
            name = Path(current_name).stem
            ext = Path(current_name).suffix
            
            if prefix:
                name = prefix + name
            if suffix:
                name = name + suffix
                
            data['new_name'] = name + ext
        
        self.update_table()
        self.status_label.setText("Префикс/Суффикс добавлены")

    def replace_text_in_names(self):
        find_txt = self.find_text.text()
        replace_txt = self.replace_text.text()
        
        if not find_txt:
            return
        
        for data in self.files_data:
            # Если new_name пуст, инициализируем его старым именем перед заменой
            if not data['new_name']:
                data['new_name'] = data['old_name']
                
            data['new_name'] = data['new_name'].replace(find_txt, replace_txt)
        
        self.update_table()
        self.status_label.setText(f"Замена '{find_txt}' на '{replace_txt}' выполнена")

    def apply_formatting(self):
        for data in self.files_data:
            # Если new_name пуст, инициализируем его старым именем
            if not data['new_name']:
                data['new_name'] = data['old_name']
                
            name = Path(data['new_name']).stem
            ext = Path(data['new_name']).suffix
            
            if self.lowercase_check.isChecked():
                name = name.lower()
            elif self.uppercase_check.isChecked():
                name = name.upper()
                
            if self.remove_spaces_check.isChecked():
                name = name.replace(" ", "")
                
            data['new_name'] = name + ext
        
        self.update_table()
        self.status_label.setText("Форматирование применено")

    def reset_names(self):
        for data in self.files_data:
            data['new_name'] = ""  # Сбрасываем в пустоту, как при загрузке
            data['status'] = ''
        self.update_table()
        self.status_label.setText("Изменения сброшены")

    def create_template(self):
        """Создает Excel файл с текущим списком файлов"""
        if not self.files_data:
            QMessageBox.warning(self, "Ошибка", "Сначала загрузите файлы!")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Сохранить шаблон", 
            "rename_template.xlsx", 
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "RenameTemplate"
            
            # Заголовки
            ws.append(["Старое имя", "Новое имя"])
            
            # Данные
            for data in self.files_data:
                ws.append([data['old_name'], data['new_name']])
            
            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Шаблон сохранен:\n{file_path}")
            self.status_label.setText("Шаблон создан")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{str(e)}")

    def load_template(self):
        """Загружает новые имена из Excel файла"""
        if not self.files_data:
            QMessageBox.warning(self, "Ошибка", "Сначала загрузите файлы!")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Загрузить шаблон", 
            "", 
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Пропускаем заголовок, начинаем со второй строки
            updates_count = 0
            
            # Создаем словарь для быстрого поиска: {OldName: NewName}
            template_map = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 2:
                    old_name_val = str(row[0]).strip() if row[0] is not None else ""
                    new_name_val = str(row[1]).strip() if row[1] is not None else ""
                    if old_name_val:
                        template_map[old_name_val] = new_name_val
            
            # Применяем данные к текущему списку
            for data in self.files_data:
                old_name = data['old_name']
                if old_name in template_map:
                    data['new_name'] = template_map[old_name]
                    updates_count += 1
            
            self.update_table()
            QMessageBox.information(self, "Успех", f"Загружено имен: {updates_count}")
            self.status_label.setText(f"Шаблон загружен. Обновлено: {updates_count}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл:\n{str(e)}")

    def execute_rename(self):
        if not self.files_data:
            return
        
        # Проверка, все ли новые имена заполнены (опционально, можно убрать если разрешены пустые)
        empty_names = [d['old_name'] for d in self.files_data if not d['new_name']]
        if empty_names:
            reply = QMessageBox.question(
                self,
                "Предупреждение",
                f"У {len(empty_names)} файлов не указано новое имя. Они будут пропущены или переименованы в исходные. Продолжить?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                return

        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите переименовать файлы?\nЭто действие нельзя отменить!",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.No:
            return
        
        success_count = 0
        error_count = 0
        skip_count = 0
        
        for i, data in enumerate(self.files_data):
            # Если новое имя пустое, пропускаем или используем старое
            if not data['new_name']:
                data['status'] = "⚠️ Пропущено (нет имени)"
                skip_count += 1
                continue

            if data['old_name'] == data['new_name']:
                data['status'] = "Без изменений"
                continue
            
            ok, msg = RenameWorker.apply_rename(data['path'], data['new_name'])
            if ok:
                data['status'] = "✅ Успешно"
                # Обновляем путь и старое имя для корректности дальнейших операций
                data['path'] = data['path'].parent / data['new_name']
                data['old_name'] = data['new_name']
                success_count += 1
            else:
                data['status'] = f"❌ Ошибка: {msg}"
                error_count += 1
        
        self.update_table()
        self.status_label.setText(f"Готово! Успешно: {success_count}, Ошибок: {error_count}, Пропущено: {skip_count}")
        
        if error_count > 0:
            QMessageBox.warning(self, "Внимание", f"При переименовании возникло {error_count} ошибок.")
        else:
            QMessageBox.information(self, "Успех", f"Операция завершена! Успешно: {success_count}")

def main():
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    window = FileRenamerApp()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()