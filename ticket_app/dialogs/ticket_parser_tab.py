# ticket_app/dialogs/ticket_parser_tab.py
"""
Вкладка парсера авиабилетов — встраивает логику из АвиаЖД_билеты_only_RU.py
в tkinter-фрейм без изменения оригинального файла.
"""

import sys
import os
import re
import json
import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd

logger = logging.getLogger(__name__)

# --- Попытка импорта PyMuPDF ---
try:
    import fitz
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# --- Попытка импорта PIL + pytesseract ---
try:
    from PIL import Image
    import pytesseract
    IMAGE_SUPPORT = True
except ImportError:
    IMAGE_SUPPORT = False


# -----------------------------------------------------------------------
# Парсер (логика из оригинального файла, без изменений)
# -----------------------------------------------------------------------

IATA_CODES = {
    'SVO': 'МОСКВА', 'DME': 'МОСКВА', 'VKO': 'МОСКВА', 'ZIA': 'МОСКВА',
    'LED': 'САНКТ-ПЕТЕРБУРГ', 'AER': 'СОЧИ', 'KRR': 'КРАСНОДАР',
    'ROV': 'РОСТОВ-НА-ДОНУ', 'KZN': 'КАЗАНЬ', 'UFA': 'УФА',
    'TAS': 'ТАШКЕНТ', 'TMJ': 'ТЕРМЕЗ', 'BHK': 'БУХАРА',
    'SKD': 'САМАРКАНД', 'FEG': 'ФЕРГАНА', 'AZN': 'АНДИЖАН',
    'NMA': 'НАМАНГАН', 'ALA': 'АЛМАТЫ', 'NQZ': 'АСТАНА',
    'MSQ': 'МИНСК', 'EVN': 'ЕРЕВАН', 'GYD': 'БАКУ',
    'FRU': 'БИШКЕК', 'DYU': 'ДУШАНБЕ', 'LBD': 'ХУДЖАНД',
    'IST': 'СТАМБУЛ', 'SAW': 'СТАМБУЛ', 'DXB': 'ДУБАЙ',
}

MONTH_MAP = {
    'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04', 'MAY': '05', 'JUN': '06',
    'JUL': '07', 'AUG': '08', 'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12',
    'ЯНВ': '01', 'ФЕВ': '02', 'МАР': '03', 'АПР': '04', 'МАЙ': '05', 'ИЮН': '06',
    'ИЮЛ': '07', 'АВГ': '08', 'СЕН': '09', 'ОКТ': '10', 'НОЯ': '11', 'ДЕК': '12',
    'MAP': '03',
}

IATA_TO_NAME = {
    'HY': 'UZBEKISTAN AIRWAYS', 'KC': 'AIR ASTANA', 'SU': 'AEROFLOT',
    'S7': 'S7 AIRLINES', 'U6': 'URAL AIRLINES', 'DP': 'POBEDA',
    'B2': 'BELAVIA', 'TK': 'TURKISH AIRLINES', 'PC': 'PEGASUS AIRLINES',
    'LH': 'LUFTHANSA', 'AF': 'AIR FRANCE', 'EK': 'EMIRATES',
    'FZ': 'FLY DUBAI', 'W6': 'WIZZ AIR', 'FR': 'RYANAIR',
}


def _extract_text_from_pdf(pdf_path: str) -> str:
    if not PDF_SUPPORT:
        return ""
    try:
        doc = fitz.open(pdf_path)
        text = "".join(page.get_text() for page in doc)
        doc.close()
        return text
    except Exception as e:
        logger.error("Ошибка чтения PDF %s: %s", pdf_path, e)
        return ""


def _extract_text_from_image(image_path: str) -> str:
    if not IMAGE_SUPPORT:
        return ""
    try:
        img = Image.open(image_path).convert('L')
        return pytesseract.image_to_string(img, lang='rus+eng')
    except Exception as e:
        logger.error("Ошибка OCR %s: %s", image_path, e)
        return ""


def _parse_ticket(text: str, filename: str) -> Dict:
    data = {
        'Пассажир': 'Не указано', 'Номер билета': 'Не указано',
        'Номер заказа': 'Не указано', 'Дата выдачи': 'Не указано',
        'Перевозчик': 'Не указано', 'Номер рейса': 'Не указано',
        'Время отправления': 'Не указано', 'Дата отправления': 'Не указано',
        'Маршрут': 'Не указано', 'Аэропорт прибытия': 'Не указано',
        'Время прибытия': 'Не указано', 'Дата прибытия': 'Не указано',
        'Стоимость': 'Не указано', 'Валюта': 'RUB', 'Источник': Path(filename).stem,
    }

    # Пассажир из имени файла
    m = re.match(r'^([A-Za-zА-Яа-я]+ [A-Za-zА-Яа-я]+)', Path(filename).stem)
    if m:
        data['Пассажир'] = m.group(1).strip()

    # Номер билета
    m = re.search(r'НОМЕР БИЛЕТА\s*:\s*(\d+\s+\d+)', text)
    if m:
        data['Номер билета'] = m.group(1).replace(' ', '')

    # Номер заказа из имени файла
    m = re.search(r'-\s*([A-Z0-9]+)', Path(filename).stem)
    if m:
        data['Номер заказа'] = m.group(1)

    # Дата выдачи
    m = re.search(r'ДАТА:\s*(\d{2})([А-ЯA-Z]{3})(\d{2})', text)
    if m:
        d, mo, y = m.groups()
        mn = MONTH_MAP.get(mo[:3].upper(), '01')
        data['Дата выдачи'] = f"{d}.{mn}.20{y}"

    # Номер рейса
    m = re.search(r'\b([A-Z]{2})\s*(\d{3,4})\b', text)
    if m:
        data['Номер рейса'] = f"{m.group(1)}{m.group(2)}"
        code = m.group(1)
        data['Перевозчик'] = IATA_TO_NAME.get(code, code)

    # Время и дата вылета
    m = re.search(r'(\d{2})([А-ЯA-Z]{3})\s+(\d{2})(\d{2})\s+(\d{2})(\d{2})', text)
    if m:
        day, mo_str, dh, dm, ah, am = m.groups()
        mn = MONTH_MAP.get(mo_str[:3].upper(), '01')
        year = data['Дата выдачи'].split('.')[-1] if data['Дата выдачи'] != 'Не указано' else '2026'
        data['Время отправления'] = f"{dh}:{dm}"
        data['Время прибытия'] = f"{ah}:{am}"
        data['Дата отправления'] = f"{day}.{mn}.{year}"
        try:
            dep = datetime(int(year), int(mn), int(day), int(dh), int(dm))
            arr = dep if int(ah) >= int(dh) else dep + timedelta(days=1)
            data['Дата прибытия'] = arr.strftime('%d.%m.%Y')
        except Exception:
            data['Дата прибытия'] = data['Дата отправления']

    # Маршрут по IATA кодам
    codes = re.findall(r'\b([A-Z]{3})\b', text)
    exclude = {'NUC', 'ROE', 'RUB', 'EUR', 'USD', 'END', 'MOW', 'NDC', 'PC'}
    valid = [c for c in codes if c not in exclude and c in IATA_CODES]
    if len(valid) >= 2:
        dep_city = IATA_CODES.get(valid[0], valid[0])
        arr_city = IATA_CODES.get(valid[1], valid[1])
        data['Маршрут'] = f"{dep_city} - {arr_city}"
        data['Аэропорт прибытия'] = f"{arr_city} ({valid[1]})"

    # Стоимость
    m = re.search(r'ИТОГО\s*:\s*(\d+)\s*РУБ', text, re.I)
    if m:
        data['Стоимость'] = m.group(1)
    else:
        m = re.search(r'ИТОГО К ОПЛАТЕ\s*:\s*RUB(\d+)', text)
        if m:
            data['Стоимость'] = m.group(1)

    return data


def parse_file(file_path: str) -> Dict:
    path = Path(file_path)
    if path.suffix.lower() == '.pdf':
        text = _extract_text_from_pdf(file_path)
    else:
        text = _extract_text_from_image(file_path)
    return _parse_ticket(text, path.name)


# -----------------------------------------------------------------------
# Виджет вкладки
# -----------------------------------------------------------------------

COLUMNS = [
    'Пассажир', 'Номер билета', 'Номер заказа', 'Дата выдачи',
    'Перевозчик', 'Номер рейса', 'Время отправления', 'Дата отправления',
    'Маршрут', 'Аэропорт прибытия', 'Время прибытия', 'Дата прибытия',
    'Стоимость', 'Валюта', 'Источник',
]


class TicketParserTab(tk.Frame):
    """Вкладка парсера авиабилетов для встраивания в ttk.Notebook."""

    def __init__(self, parent):
        super().__init__(parent)
        self.tickets: List[Dict] = []
        self._build_ui()

    def _build_ui(self):
        # Панель управления
        ctrl = tk.Frame(self, bg='#16213e', height=45)
        ctrl.pack(fill='x')
        ctrl.pack_propagate(False)

        btn_cfg = dict(fg='white', font=('Arial', 10), padx=12, pady=6)

        tk.Button(ctrl, text="📁 Выбрать папку", command=self._select_folder,
                  bg='#0f3460', **btn_cfg).pack(side='left', padx=5, pady=5)
        tk.Button(ctrl, text="📄 Выбрать файл(ы)", command=self._select_files,
                  bg='#0f3460', **btn_cfg).pack(side='left', padx=5, pady=5)
        tk.Button(ctrl, text="📊 Экспорт в Excel", command=self._export_excel,
                  bg='#28a745', **btn_cfg).pack(side='left', padx=5, pady=5)
        tk.Button(ctrl, text="📋 Экспорт в JSON", command=self._export_json,
                  bg='#17a2b8', **btn_cfg).pack(side='left', padx=5, pady=5)
        tk.Button(ctrl, text="🗑 Очистить", command=self._clear,
                  bg='#dc3545', **btn_cfg).pack(side='left', padx=5, pady=5)

        self.progress = ttk.Progressbar(ctrl, mode='determinate', length=200)
        self.progress.pack(side='left', padx=10, pady=8)
        self.progress_lbl = tk.Label(ctrl, text="", font=('Arial', 9),
                                     bg='#16213e', fg='white')
        self.progress_lbl.pack(side='left')

        # Таблица
        tbl_frame = tk.Frame(self)
        tbl_frame.pack(fill='both', expand=True, padx=5, pady=5)

        vsb = ttk.Scrollbar(tbl_frame, orient='vertical')
        vsb.pack(side='right', fill='y')
        hsb = ttk.Scrollbar(tbl_frame, orient='horizontal')
        hsb.pack(side='bottom', fill='x')

        self.table = ttk.Treeview(tbl_frame, columns=COLUMNS, show='headings',
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=self.table.yview)
        hsb.config(command=self.table.xview)

        col_w = {'Пассажир': 140, 'Маршрут': 160, 'Перевозчик': 130,
                 'Номер билета': 110, 'Источник': 120}
        for col in COLUMNS:
            self.table.heading(col, text=col)
            self.table.column(col, width=col_w.get(col, 90), anchor='w')
        self.table.pack(fill='both', expand=True)

        # Статус
        self.status = tk.Label(self, text="Готов", bd=1, relief='sunken',
                               anchor='w', font=('Arial', 9), bg='#f0f0f0')
        self.status.pack(side='bottom', fill='x')

    # ------------------------------------------------------------------

    def _select_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с билетами")
        if not folder:
            return
        files = []
        for ext in ('*.pdf', '*.PDF', '*.jpg', '*.jpeg', '*.png', '*.tiff'):
            files.extend(Path(folder).glob(ext))
        if not files:
            messagebox.showinfo("Нет файлов", "PDF и изображения не найдены")
            return
        self._process([str(f) for f in sorted(set(files))])

    def _select_files(self):
        files = filedialog.askopenfilenames(
            title="Выберите файлы",
            filetypes=[("PDF и изображения", "*.pdf *.jpg *.jpeg *.png *.tiff *.PDF")])
        if files:
            self._process(list(files))

    def _process(self, files: List[str]):
        self.progress['maximum'] = len(files)
        self.progress['value'] = 0
        errors = 0
        for i, fp in enumerate(files):
            self.progress_lbl.config(text=f"{i+1}/{len(files)}: {Path(fp).name}")
            self.update_idletasks()
            try:
                data = parse_file(fp)
                self.tickets.append(data)
                vals = [data.get(c, '') for c in COLUMNS]
                tag = 'unknown' if 'Не указано' in (data.get('Номер билета',''), data.get('Маршрут','')) else ''
                iid = self.table.insert('', 'end', values=vals, tags=(tag,))
                if tag:
                    self.table.tag_configure('unknown', background='#fff3cd')
            except Exception as e:
                logger.error("Ошибка парсинга %s: %s", fp, e)
                errors += 1
            self.progress['value'] = i + 1
        self.progress_lbl.config(text="")
        self.status.config(
            text=f"Обработано: {len(files)} файлов | Билетов: {len(self.tickets)} | Ошибок: {errors}")

    def _export_excel(self):
        if not self.tickets:
            messagebox.showwarning("Нет данных", "Нет данных для экспорта")
            return
        fp, _ = filedialog.asksaveasfilename(
            title="Сохранить Excel",
            defaultextension=".xlsx",
            initialfile=f"билеты_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel", "*.xlsx")]), None
        if not fp:
            return
        try:
            df = pd.DataFrame(self.tickets)[COLUMNS]
            with pd.ExcelWriter(fp, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Билеты', index=False)
                ws = writer.sheets['Билеты']
                from openpyxl.styles import Font, PatternFill, Alignment
                hf = Font(bold=True, color='FFFFFF')
                hfill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                for cell in ws[1]:
                    cell.font = hf
                    cell.fill = hfill
                    cell.alignment = Alignment(horizontal='center')
                for col in ws.columns:
                    w = max((len(str(c.value or '')) for c in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(w + 2, 40)
            messagebox.showinfo("Успех", f"Сохранено:\n{fp}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _export_json(self):
        if not self.tickets:
            return
        fp, _ = filedialog.asksaveasfilename(
            title="Сохранить JSON",
            defaultextension=".json",
            initialfile=f"билеты_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            filetypes=[("JSON", "*.json")]), None
        if not fp:
            return
        try:
            with open(fp, 'w', encoding='utf-8') as f:
                json.dump(self.tickets, f, ensure_ascii=False, indent=2)
            self.status.config(text=f"JSON сохранён: {fp}")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _clear(self):
        self.tickets.clear()
        self.table.delete(*self.table.get_children())
        self.status.config(text="Очищено")
