#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FolderForge Pro - Ultimate Toolset
Multi-tool application with responsive UI design
PyQt6 Edition
"""

import sys
import os
import re
import json
import logging
import subprocess
import io
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional, Dict, Tuple, Any, Set
from dataclasses import dataclass, asdict
from enum import Enum

# Excel & Data handling
import pandas as pd
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import pyxlsb
    from pyxlsb import open_workbook as open_xlsb
except ImportError:
    pyxlsb = None

# PDF support
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

# PyQt5 Imports
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# --- CONSTANTS & DATA STRUCTURES ---

class StructureType(Enum):
    FLAT = "flat"
    ALPHABETICAL = "alphabetical"
    BY_DATE = "by_date"
    BY_FIRST_WORD = "by_first_word"


@dataclass
class FolderConfig:
    structure_type: StructureType = StructureType.FLAT
    remove_duplicates: bool = True
    validate_names: bool = True
    create_log: bool = True
    open_after_creation: bool = True
    date_format: str = "%Y-%m-%d"
    prefix: str = ""
    suffix: str = ""


# --- THEME ---

class ModernStyle:
    """Modern Catppuccin Mocha Theme with Responsive Typography"""

    # Colors
    BASE = "#1e1e2e"
    MANTLE = "#181825"
    CRUST = "#11111b"
    SURFACE0 = "#313244"
    SURFACE1 = "#45475a"
    SURFACE2 = "#585b70"
    OVERLAY0 = "#6c7086"
    OVERLAY1 = "#7f849c"
    OVERLAY2 = "#9399b2"
    TEXT = "#cdd6f4"
    SUBTEXT0 = "#a6adc8"
    SUBTEXT1 = "#bac2de"
    BLUE = "#89b4fa"
    LAVENDER = "#b4befe"
    SAPPHIRE = "#74c7ec"
    SKY = "#89dceb"
    TEAL = "#94e2d5"
    GREEN = "#a6e3a1"
    YELLOW = "#f9e2af"
    PEACH = "#fab387"
    MAROON = "#eba0ac"
    RED = "#f38ba8"
    MAUVE = "#cba6f7"
    PINK = "#f5c2e7"
    FLAMINGO = "#f2cdcd"
    ROSEWATER = "#f5e0dc"

    # Typography Scale (responsive)
    FONT_FAMILY = "Segoe UI", "SF Pro Display", "Helvetica Neue", "Arial", "sans-serif"

    @classmethod
    def get_font_sizes(cls, base_size: int = 10) -> Dict[str, int]:
        """Calculate font sizes based on base size"""
        return {
            'tiny': int(base_size * 0.75),
            'small': int(base_size * 0.875),
            'normal': base_size,
            'medium': int(base_size * 1.125),
            'large': int(base_size * 1.25),
            'xlarge': int(base_size * 1.5),
            'xxlarge': int(base_size * 1.75),
            'huge': int(base_size * 2),
            'title': int(base_size * 2.5),
        }

    @classmethod
    def get_stylesheet(cls, base_font_size: int = 10) -> str:
        fonts = cls.get_font_sizes(base_font_size)

        return f"""
        QMainWindow, QDialog {{
            background-color: {cls.BASE};
            color: {cls.TEXT};
            font-family: "Segoe UI", "SF Pro Display", "Helvetica Neue", Arial, sans-serif;
            font-size: {fonts['normal']}pt;
        }}

        QWidget {{
            background-color: {cls.BASE};
            color: {cls.TEXT};
            font-size: {fonts['normal']}pt;
        }}

        /* === TYPOGRAPHY === */
        QLabel {{
            color: {cls.TEXT};
            font-size: {fonts['normal']}pt;
            padding:4px 0px;
        }}

        QLabel[class="title"] {{
            font-size: {fonts['title']}pt;
            font-weight: 700;
            color: {cls.LAVENDER};
            padding: 16px 0px 8px 0px;
        }}

        QLabel[class="subtitle"] {{
            font-size: {fonts['xlarge']}pt;
            font-weight: 600;
            color: {cls.BLUE};
            padding: 12px 0px 6px 0px;
        }}

        QLabel[class="heading"] {{
            font-size: {fonts['large']}pt;
            font-weight: 600;
            color: {cls.TEXT};
            padding: 8px 0px 4px 0px;
        }}

        QLabel[class="caption"] {{
            font-size: {fonts['small']}pt;
            color: {cls.SUBTEXT0};
            padding: 2px 0px;
        }}

        /* === BUTTONS === */
        QPushButton {{
            background-color: {cls.SURFACE0};
            color: {cls.TEXT};
            border:1px solid {cls.SURFACE1};
            border-radius: 8px;
            padding: 10px 20px;
            font-size: {fonts['normal']}pt;
            font-weight: 500;
            min-width: 100px;
            min-height: 20px;
        }}

        QPushButton:hover {{
            background-color: {cls.SURFACE1};
            border-color: {cls.SURFACE2};
        }}

        QPushButton:pressed {{
            background-color: {cls.SURFACE2};
        }}

        QPushButton:disabled {{
            background-color: {cls.CRUST};
            color: {cls.OVERLAY0};
            border-color: {cls.SURFACE0};
        }}

        QPushButton[class="primary"] {{
            background-color: {cls.BLUE};
            color: {cls.CRUST};
            border: none;
            font-weight: 600;
            padding: 12px 24px;
        }}

        QPushButton[class="primary"]:hover {{
            background-color: {cls.SAPPHIRE};
        }}

        QPushButton[class="success"] {{
            background-color: {cls.GREEN};
            color: {cls.CRUST};
            border: none;
            font-weight: 600;
        }}

        QPushButton[class="danger"] {{
            background-color: {cls.RED};
            color: {cls.CRUST};
            border: none;
            font-weight: 600;
        }}

        QPushButton[class="icon-button"] {{
            min-width: 36px;
            min-height: 36px;
            max-width: 36px;
            max-height: 36px;
            padding: 0px;
            border-radius: 6px;
            font-size: {fonts['large']}pt;
        }}

        /* === INPUT FIELDS === */
        QLineEdit, QTextEdit, QComboBox {{
            background-color: {cls.SURFACE0};
            color: {cls.TEXT};
            border: 2px solid {cls.SURFACE1};
            border-radius: 6px;
            padding: 8px 12px;
            font-size: {fonts['normal']}pt;
            min-height: 20px;
        }}

        QLineEdit:focus, QTextEdit:focus, QComboBox:focus {{
            border-color: {cls.BLUE};
        }}

        QLineEdit:disabled, QTextEdit:disabled, QComboBox:disabled {{
            background-color: {cls.CRUST};
            color: {cls.OVERLAY0};
        }}

        QComboBox::drop-down {{
            border: none;
            width: 30px;
        }}

        QComboBox QAbstractItemView {{
            background-color: {cls.SURFACE0};
            color: {cls.TEXT};
            selection-background-color: {cls.BLUE};
            selection-color: {cls.CRUST};
            border:1px solid {cls.SURFACE1};
        }}

        /* === TABLES === */
        QTableWidget, QTreeWidget {{
            background-color: {cls.MANTLE};
            color: {cls.TEXT};
            border: 1px solid {cls.SURFACE0};
            border-radius: 8px;
            gridline-color: {cls.SURFACE0};
            font-size: {fonts['normal']}pt;
        }}

        QTableWidget::item, QTreeWidget::item {{
            padding: 8px 12px;
            border-bottom: 1px solid {cls.SURFACE0};
        }}

        QTableWidget::item:selected, QTreeWidget::item:selected {{
            background-color: {cls.BLUE};
            color: {cls.CRUST};
        }}

        QHeaderView::section {{
            background-color: {cls.SURFACE0};
            color: {cls.TEXT};
            padding: 10px 12px;
            border: none;
            border-bottom: 2px solid {cls.SURFACE1};
            font-weight: 600;
            font-size: {fonts['small']}pt;
        }}

        QHeaderView::section:hover {{
            background-color: {cls.SURFACE1};
        }}

        /* === GROUP BOXES === */
        QGroupBox {{
            background-color: {cls.MANTLE};
            border: 1px solid {cls.SURFACE0};
            border-radius: 12px;
            margin-top: 16px;
            padding-top: 16px;
            font-size: {fonts['medium']}pt;
            font-weight: 600;
        }}

        QGroupBox::title {{
            subcontrol-origin: margin;
            left: 16px;
            padding: 0px 12px;
            color: {cls.LAVENDER};
        }}

        /* === PROGRESS BAR === */
        QProgressBar {{
            background-color: {cls.SURFACE0};
            border: none;
            border-radius: 6px;
            height: 8px;
            text-align: center;
            font-size: {fonts['small']}pt;
        }}

        QProgressBar::chunk {{
            background-color: {cls.BLUE};
            border-radius: 6px;
        }}

        /* === SCROLL BARS === */
        QScrollBar:vertical {{
            background-color: {cls.CRUST};
            width: 12px;
            border-radius: 6px;
        }}

        QScrollBar::handle:vertical {{
            background-color: {cls.SURFACE0};
            border-radius: 6px;
            min-height: 30px;
        }}

        QScrollBar::handle:vertical:hover {{
            background-color: {cls.SURFACE1};
        }}

        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
            height: 0px;
        }}

        /* === TABS === */
        QTabWidget::pane {{
            border: 1px solid {cls.SURFACE0};
            border-radius: 8px;
            background-color: {cls.MANTLE};
        }}

        QTabBar::tab {{
            background-color: {cls.SURFACE0};
            color: {cls.SUBTEXT0};
            padding: 10px 20px;
            margin-right: 4px;
            border-top-left-radius: 6px;
            border-top-right-radius: 6px;
            font-size: {fonts['normal']}pt;
        }}

        QTabBar::tab:selected {{
            background-color: {cls.BLUE};
            color: {cls.CRUST};
        }}

        QTabBar::tab:hover:!selected {{
            background-color: {cls.SURFACE1};
            color: {cls.TEXT};
        }}

        /* === MENU === */
        QMenuBar {{
            background-color: {cls.MANTLE};
            color: {cls.TEXT};
            padding: 4px;
            font-size: {fonts['normal']}pt;
        }}

        QMenuBar::item:selected {{
            background-color: {cls.SURFACE0};
            border-radius: 4px;
        }}

        QMenu {{
            background-color: {cls.SURFACE0};
            color: {cls.TEXT};
            border: 1px solid {cls.SURFACE1};
            padding: 8px;
            font-size: {fonts['normal']}pt;
        }}

        QMenu::item {{
            padding: 8px 24px;
            border-radius: 4px;
        }}

        QMenu::item:selected {{
            background-color: {cls.BLUE};
            color: {cls.CRUST};
        }}

        /* === STATUS BAR === */
        QStatusBar {{
            background-color: {cls.MANTLE};
            color: {cls.SUBTEXT0};
            border-top: 1px solid {cls.SURFACE0};
            font-size: {fonts['small']}pt;
            padding: 4px 16px;
        }}

        /* === TOOL BAR === */
        QToolBar {{
            background-color: {cls.MANTLE};
            border: none;
            spacing: 8px;
            padding: 8px;
        }}

        QToolButton {{
            background-color: transparent;
            border: none;
            border-radius: 6px;
            padding: 8px;
            font-size: {fonts['normal']}pt;
        }}

        QToolButton:hover {{
            background-color: {cls.SURFACE0};
        }}

        /* === LIST WIDGET === */
        QListWidget {{
            background-color: {cls.MANTLE};
            border: 1px solid {cls.SURFACE0};
            border-radius: 8px;
            padding: 8px;
            font-size: {fonts['normal']}pt;
        }}

        QListWidget::item {{
            padding: 10px 12px;
            border-radius: 6px;
            margin: 2px 0px;
        }}

        QListWidget::item:selected {{
            background-color: {cls.BLUE};
            color: {cls.CRUST};
        }}

        QListWidget::item:hover:!selected {{
            background-color: {cls.SURFACE0};
        }}

        /* === CHECKBOX === */
        QCheckBox {{
            font-size: {fonts['normal']}pt;
            spacing: 8px;
        }}

        QCheckBox::indicator {{
            width: 20px;
            height: 20px;
            border-radius: 4px;
            border: 2px solid {cls.SURFACE1};
            background-color: {cls.SURFACE0};
        }}

        QCheckBox::indicator:checked {{
            background-color: {cls.BLUE};
            border-color: {cls.BLUE};
        }}

        /* === FRAME === */
        QFrame[class="card"] {{
            background-color: {cls.MANTLE};
            border: 1px solid {cls.SURFACE0};
            border-radius: 12px;
            padding: 16px;
        }}

        QFrame[class="divider"] {{
            background-color: {cls.SURFACE0};
            max-height: 1px;
            margin: 16px 0px;
        }}

        /* === SPLITTER === */
        QSplitter::handle {{
            background-color: {cls.SURFACE0};
        }}

        QSplitter::handle:horizontal {{
            width: 2px;
        }}

        QSplitter::handle:vertical {{
            height: 2px;
        }}
        """


# --- TAB 1: FOLDER CREATION (Full Logic Restored) ---

class FolderWorker(QThread):
    """Worker thread for folder creation"""
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    finished = pyqtSignal(bool, list)

    def __init__(self, folder_names: List[str], base_path: str, config: FolderConfig):
        super().__init__()
        self.folder_names = folder_names
        self.base_path = base_path
        self.config = config
        self.created_folders = []
        self._is_running = True

    def run(self):
        try:
            total = len(self.folder_names)

            for i, name in enumerate(self.folder_names):
                if not self._is_running:
                    break

                # Create folder based on structure type
                folder_path = self._create_folder_structure(name)
                if folder_path:
                    self.created_folders.append(folder_path)

                progress = int((i + 1) / total * 100)
                self.progress.emit(progress)
                self.status.emit(f"Создание: {name}")

            self.finished.emit(True, self.created_folders)
        except Exception as e:
            logger.error(f"Worker error: {e}")
            self.finished.emit(False, [str(e)])

    def _create_folder_structure(self, name: str) -> Optional[str]:
        """Create folder with selected structure"""
        try:
            base = Path(self.base_path)

            if self.config.structure_type == StructureType.FLAT:
                folder_path = base / f"{self.config.prefix}{name}{self.config.suffix}"
            elif self.config.structure_type == StructureType.ALPHABETICAL:
                first_char = name[0].upper() if name else "_"
                folder_path = base / first_char / f"{self.config.prefix}{name}{self.config.suffix}"
            elif self.config.structure_type == StructureType.BY_DATE:
                today = datetime.now().strftime(self.config.date_format)
                folder_path = base / today / f"{self.config.prefix}{name}{self.config.suffix}"
            elif self.config.structure_type == StructureType.BY_FIRST_WORD:
                first_word = name.split()[0] if name.split() else "_"
                folder_path = base / first_word / f"{self.config.prefix}{name}{self.config.suffix}"
            else:
                folder_path = base / name

            folder_path.mkdir(parents=True, exist_ok=True)
            return str(folder_path)
        except Exception as e:
            logger.error(f"Error creating folder {name}: {e}")
            return None

    def stop(self):
        self._is_running = False


class BaseTabWidget(QWidget):
    """Базовый класс для всех вкладок"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.scale_factor = 1.0
        self.setup_ui()

    def setup_ui(self):
        """Метод для переопределения в дочерних классах"""
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Содержимое вкладки"))


class FolderCreationTab(BaseTabWidget):
    """Вкладка 1: Создание папок"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.folder_names: List[str] = []
        self.config = FolderConfig()
        self.current_file: Optional[str] = None

    def setup_ui(self):
        main_layout = QHBoxLayout(self)
        sp = 8
        main_layout.setContentsMargins(sp, sp, sp, sp)
        main_layout.setSpacing(sp)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(self.create_left_panel())
        splitter.addWidget(self.create_right_panel())
        splitter.setSizes([400, 600])
        
        main_layout.addWidget(splitter)

    def create_left_panel(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        header = QLabel("📂 Создание папок")
        header.setProperty("class", "title")
        layout.addWidget(header)

        subtitle = QLabel("Массовое создание папок из Excel/CSV/TXT")
        subtitle.setProperty("class", "caption")
        layout.addWidget(subtitle)

        divider = QFrame()
        divider.setProperty("class", "divider")
        divider.setFrameShape(QFrame.Shape.HLine)
        layout.addWidget(divider)

        file_group = QGroupBox("📁 Источник данных")
        file_layout = QVBoxLayout(file_group)

        file_input_layout = QHBoxLayout()
        self.file_path_input = QLineEdit()
        self.file_path_input.setPlaceholderText("Перетащите файл или нажмите 'Открыть'...")
        self.file_path_input.setMinimumHeight(36)
        file_input_layout.addWidget(self.file_path_input)

        browse_btn = QPushButton("📂 Открыть")
        browse_btn.setProperty("class", "primary")
        browse_btn.clicked.connect(self.browse_file)
        file_input_layout.addWidget(browse_btn)
        file_layout.addLayout(file_input_layout)

        paste_label = QLabel("⚡ Быстрый ввод:")
        paste_label.setProperty("class", "heading")
        file_layout.addWidget(paste_label)

        self.quick_input = QTextEdit()
        self.quick_input.setPlaceholderText("Например: Проект A, Проект B...")
        self.quick_input.setMaximumHeight(80)
        self.quick_input.textChanged.connect(self.on_quick_input_changed)
        file_layout.addWidget(self.quick_input)
        layout.addWidget(file_group)

        preview_group = QGroupBox("👁 Предпросмотр")
        preview_layout = QVBoxLayout(preview_group)
        self.stats_label = QLabel("Загружено: 0 папок")
        self.stats_label.setProperty("class", "caption")
        preview_layout.addWidget(self.stats_label)

        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(3)
        self.preview_table.setHorizontalHeaderLabels(["№", "Имя папки", "Статус"])
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.preview_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setMinimumHeight(300)
        preview_layout.addWidget(self.preview_table)

        btn_layout = QHBoxLayout()
        self.validate_btn = QPushButton("✓ Проверить имена")
        self.validate_btn.clicked.connect(self.validate_names)
        self.validate_btn.setEnabled(False)
        btn_layout.addWidget(self.validate_btn)

        self.remove_duplicates_btn = QPushButton("🗑 Удалить дубликаты")
        self.remove_duplicates_btn.clicked.connect(self.remove_duplicates)
        self.remove_duplicates_btn.setEnabled(False)
        btn_layout.addWidget(self.remove_duplicates_btn)
        
        self.sort_btn = QPushButton("⇅ Сортировать")
        self.sort_btn.clicked.connect(self.sort_names)
        self.sort_btn.setEnabled(False)
        btn_layout.addWidget(self.sort_btn)

        preview_layout.addLayout(btn_layout)
        layout.addWidget(preview_group, 1)
        return panel

    def create_right_panel(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        settings_group = QGroupBox("⚙ Настройки структуры")
        settings_layout = QVBoxLayout(settings_group)
        
        structure_label = QLabel("Тип организации:")
        structure_label.setProperty("class", "heading")
        settings_layout.addWidget(structure_label)

        self.structure_combo = QComboBox()
        self.structure_combo.addItems([
            "📂 Плоская", "🔤 По алфавиту", "📅 По дате", "📝 По первому слову"
        ])
        self.structure_combo.currentIndexChanged.connect(self.on_structure_changed)
        settings_layout.addWidget(self.structure_combo)

        prefix_layout = QHBoxLayout()
        prefix_layout.addWidget(QLabel("Префикс:"))
        self.prefix_input = QLineEdit()
        self.prefix_input.setPlaceholderText("[2024] ")
        prefix_layout.addWidget(self.prefix_input)
        prefix_layout.addWidget(QLabel("Суффикс:"))
        self.suffix_input = QLineEdit()
        self.suffix_input.setPlaceholderText("_backup")
        prefix_layout.addWidget(self.suffix_input)
        settings_layout.addLayout(prefix_layout)

        self.validate_check = QCheckBox("Проверять недопустимые символы")
        self.validate_check.setChecked(True)
        settings_layout.addWidget(self.validate_check)

        self.remove_dup_check = QCheckBox("Удалять дубликаты автоматически")
        self.remove_dup_check.setChecked(True)
        settings_layout.addWidget(self.remove_dup_check)

        self.open_after_check = QCheckBox("Открыть папку после создания")
        self.open_after_check.setChecked(True)
        settings_layout.addWidget(self.open_after_check)
        layout.addWidget(settings_group)

        export_group = QGroupBox("📤 Экспорт скриптов")
        export_layout = QHBoxLayout(export_group)
        
        self.export_bat_btn = QPushButton("🪟 BAT")
        self.export_bat_btn.clicked.connect(lambda: self.export_script("bat"))
        self.export_bat_btn.setEnabled(False)
        export_layout.addWidget(self.export_bat_btn)

        self.export_ps_btn = QPushButton("💻 PS1")
        self.export_ps_btn.clicked.connect(lambda: self.export_script("ps1"))
        self.export_ps_btn.setEnabled(False)
        export_layout.addWidget(self.export_ps_btn)

        self.export_py_btn = QPushButton("🐍 Python")
        self.export_py_btn.clicked.connect(lambda: self.export_script("py"))
        self.export_py_btn.setEnabled(False)
        export_layout.addWidget(self.export_py_btn)
        layout.addWidget(export_group)

        output_group = QGroupBox("📂 Выходная директория")
        output_layout = QHBoxLayout(output_group)
        self.output_path_input = QLineEdit()
        self.output_path_input.setText(str(Path.home() / "Documents" / "CreatedFolders"))
        output_layout.addWidget(self.output_path_input)
        
        output_browse_btn = QPushButton("📁")
        output_browse_btn.setProperty("class", "icon-button")
        output_browse_btn.clicked.connect(self.browse_output)
        output_layout.addWidget(output_browse_btn)
        layout.addWidget(output_group)

        progress_label = QLabel("Прогресс:")
        progress_label.setProperty("class", "heading")
        layout.addWidget(progress_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        layout.addWidget(self.progress_bar)

        self.status_label = QLabel("Готов к работе")
        self.status_label.setProperty("class", "caption")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.status_label)

        layout.addStretch()

        self.create_btn = QPushButton("🚀 СОЗДАТЬ ПАПКИ")
        self.create_btn.setProperty("class", "success")
        self.create_btn.setMinimumHeight(50)
        font = self.create_btn.font()
        font.setBold(True)
        self.create_btn.setFont(font)
        self.create_btn.clicked.connect(self.create_folders)
        self.create_btn.setEnabled(False)
        layout.addWidget(self.create_btn)

        return panel

    # Logic
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите файл", "", "Excel files (*.xlsx *.xls);;CSV files (*.csv);;Text files (*.txt)")
        if file_path:
            self.load_file(file_path)

    def load_file(self, file_path: str):
        try:
            self.current_file = file_path
            self.file_path_input.setText(file_path)
            ext = Path(file_path).suffix.lower()
            if ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path, header=None)
                self.folder_names = df.iloc[:, 0].dropna().astype(str).tolist()
            elif ext == '.csv':
                df = pd.read_csv(file_path, header=None)
                self.folder_names = df.iloc[:, 0].dropna().astype(str).tolist()
            elif ext == '.txt':
                with open(file_path, 'r', encoding='utf-8') as f:
                    self.folder_names = [line.strip() for line in f if line.strip()]
            else:
                raise ValueError("Неподдерживаемый формат")
            
            if self.remove_dup_check.isChecked():
                self.remove_duplicates()
            self.update_preview()
            
            main_win = self.window()
            if isinstance(main_win, MainWindow):
                main_win.statusbar.showMessage(f"Загружено {len(self.folder_names)} папок")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")

    def on_quick_input_changed(self):
        text = self.quick_input.toPlainText()
        if text:
            self.folder_names = [name.strip() for name in re.split(r'[,\n]', text) if name.strip()]
            self.update_preview()

    def update_preview(self):
        self.preview_table.setRowCount(len(self.folder_names))
        for i, name in enumerate(self.folder_names):
            self.preview_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.preview_table.setItem(i, 1, QTableWidgetItem(name))
            status = "✓" if self.is_valid_name(name) else "⚠"
            self.preview_table.setItem(i, 2, QTableWidgetItem(status))
        
        self.stats_label.setText(f"Загружено: {len(self.folder_names)} папок")
        has_data = len(self.folder_names) > 0
        self.create_btn.setEnabled(has_data)
        self.validate_btn.setEnabled(has_data)
        self.remove_duplicates_btn.setEnabled(has_data)
        self.sort_btn.setEnabled(has_data)
        self.export_bat_btn.setEnabled(has_data)
        self.export_ps_btn.setEnabled(has_data)
        self.export_py_btn.setEnabled(has_data)

    def is_valid_name(self, name: str) -> bool:
        invalid_chars = r'[<>:"/\\|?*]'
        return not re.search(invalid_chars, name) and name not in ['CON', 'PRN', 'AUX', 'NUL']

    def validate_names(self):
        cleaned = []
        for name in self.folder_names:
            clean = re.sub(r'[<>:"/\\|?*]', '_', name)
            cleaned.append(clean)
        self.folder_names = cleaned
        self.update_preview()

    def remove_duplicates(self):
        self.folder_names = list(dict.fromkeys(self.folder_names))
        self.update_preview()

    def sort_names(self):
        self.folder_names.sort(key=str.lower)
        self.update_preview()

    def export_script(self, format_type: str):
        if not self.folder_names: return
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить скрипт", f"create_folders.{format_type}")
        if not file_path: return
        
        try:
            content = ""
            if format_type == "bat":
                content = f"@echo off\ncd /d \"{self.output_path_input.text()}\"\n" + "\n".join([f'mkdir "{n}"' for n in self.folder_names])
            elif format_type == "ps1":
                content = f"$basePath = '{self.output_path_input.text()}'\nNew-Item -ItemType Directory -Force -Path $basePath | Out-Null\n" + "\n".join([f'New-Item -ItemType Directory -Force -Path "$basePath\\{n}"' for n in self.folder_names])
            elif format_type == "py":
                content = f"import os\nbase_path = r'{self.output_path_input.text()}'\n" + "\n".join([f'os.makedirs(os.path.join(base_path, "{n}"), exist_ok=True)' for n in self.folder_names])
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            QMessageBox.information(self, "Успех", "Скрипт сохранен")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def browse_output(self):
        dir_path = QFileDialog.getExistingDirectory(self, "Выберите директорию", self.output_path_input.text())
        if dir_path:
            self.output_path_input.setText(dir_path)

    def on_structure_changed(self, index: int):
        structures = [StructureType.FLAT, StructureType.ALPHABETICAL, StructureType.BY_DATE, StructureType.BY_FIRST_WORD]
        if 0 <= index < len(structures):
            self.config.structure_type = structures[index]

    def create_folders(self):
        if not self.folder_names: return
        output_path = self.output_path_input.text()
        if not output_path: return

        self.config.prefix = self.prefix_input.text()
        self.config.suffix = self.suffix_input.text()
        self.config.validate_names = self.validate_check.isChecked()
        self.config.remove_duplicates = self.remove_dup_check.isChecked()
        self.config.open_after_creation = self.open_after_check.isChecked()

        if self.config.validate_names:
            self.validate_names()

        Path(output_path).mkdir(parents=True, exist_ok=True)
        self.worker = FolderWorker(self.folder_names, output_path, self.config)
        self.worker.progress.connect(self.on_progress)
        self.worker.status.connect(self.on_status)
        self.worker.finished.connect(self.on_finished)
        self.create_btn.setEnabled(False)
        self.worker.start()

    def on_progress(self, value: int):
        self.progress_bar.setValue(value)

    def on_status(self, message: str):
        self.status_label.setText(message)

    def on_finished(self, success: bool, results: list):
        self.create_btn.setEnabled(True)
        if success:
            self.progress_bar.setValue(100)
            self.status_label.setText(f"✓ Создано {len(results)} папок")
            if self.config.open_after_creation and results:
                subprocess.Popen(f'explorer "{self.output_path_input.text()}"')
            QMessageBox.information(self, "Успех", f"Создано {len(results)} папок")
        else:
            self.status_label.setText("✗ Ошибка")
            QMessageBox.critical(self, "Ошибка", str(results[0]) if results else "Unknown error")

    # --- SCANNING LOGIC (RESTORED) ---
    def scan_folder_to_excel(self):
        """Scan folder and create Excel with file names"""
        folder_path = QFileDialog.getExistingDirectory(self, "Выберите папку для сканирования", str(Path.home()))
        if not folder_path: return

        dialog = QDialog(self)
        dialog.setWindowTitle("Настройки сканирования")
        layout = QVBoxLayout(dialog)
        
        subfolders_check = QCheckBox("Включить подпапки (рекурсивно)")
        subfolders_check.setChecked(False)
        layout.addWidget(subfolders_check)
        
        extensions_check = QCheckBox("Включать расширения файлов")
        extensions_check.setChecked(True)
        layout.addWidget(extensions_check)
        
        fullpath_check = QCheckBox("Включать полные пути")
        fullpath_check.setChecked(False)
        layout.addWidget(fullpath_check)
        
        size_check = QCheckBox("Включать размер файлов")
        size_check.setChecked(True)
        layout.addWidget(size_check)
        
        date_check = QCheckBox("Включать дату изменения")
        date_check.setChecked(True)
        layout.addWidget(date_check)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        if dialog.exec() != QDialog.DialogCode.Accepted: return
        
        include_subfolders = subfolders_check.isChecked()
        include_extensions = extensions_check.isChecked()
        include_fullpath = fullpath_check.isChecked()
        include_size = size_check.isChecked()
        include_date = date_check.isChecked()

        try:
            file_data = []
            base_path = Path(folder_path)
            pattern = "**/*" if include_subfolders else "*"
            
            for item in base_path.glob(pattern):
                if item.is_file():
                    row = {}
                    if include_extensions:
                        row['Имя файла'] = item.name
                    else:
                        row['Имя файла'] = item.stem
                    
                    if include_fullpath:
                        row['Полный путь'] = str(item.absolute())
                    else:
                        try:
                            row['Относительный путь'] = str(item.relative_to(base_path.parent))
                        except:
                            row['Относительный путь'] = str(item.relative_to(base_path))
                    
                    if include_size:
                        size = item.stat().st_size
                        if size < 1024: row['Размер'] = f"{size} B"
                        elif size < 1024 * 1024: row['Размер'] = f"{size / 1024:.2f} KB"
                        elif size < 1024 * 1024 * 1024: row['Размер'] = f"{size / (1024 * 1024):.2f} MB"
                        else: row['Размер'] = f"{size / (1024 * 1024 * 1024):.2f} GB"
                    
                    if include_date:
                        mtime = datetime.fromtimestamp(item.stat().st_mtime)
                        row['Дата изменения'] = mtime.strftime("%Y-%m-%d %H:%M:%S")
                    
                    file_data.append(row)

            if not file_data:
                QMessageBox.information(self, "Информация", "В выбранной папке нет файлов.")
                return

            df = pd.DataFrame(file_data)
            save_path, _ = QFileDialog.getSaveFileName(self, "Сохранить список", str(base_path / "file_list.xlsx"), "Excel files (*.xlsx);;CSV files (*.csv)")
            
            if save_path:
                if save_path.endswith('.csv'):
                    df.to_csv(save_path, index=False, encoding='utf-8-sig')
                else:
                    df.to_excel(save_path, index=False, engine='openpyxl')
                
                if QMessageBox.question(self, "Успех", f"Найдено и сохранено {len(file_data)} файлов.\n\nОткрыть файл?") == QMessageBox.StandardButton.Yes:
                    subprocess.Popen(f'explorer "{save_path}"')
                
                if QMessageBox.question(self, "Загрузить имена?", "Загрузить имена файлов в программу для создания папок?") == QMessageBox.StandardButton.Yes:
                    self.folder_names = df['Имя файла'].tolist()
                    self.update_preview()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка сканирования: {str(e)}")

    def scan_folders_to_excel(self):
        """Scan multiple folders and create Excel with file names"""
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.FileMode.Directory)
        dialog.setOption(QFileDialog.Option.DontUseNativeDialog, True)
        dialog.setWindowTitle("Выберите несколько папок (Ctrl+Click)")

        list_view = dialog.findChild(QListWidget)
        if list_view:
            list_view.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

        tree_view = dialog.findChild(QTreeView)
        if tree_view:
            tree_view.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

        if dialog.exec() != QDialog.DialogCode.Accepted: return
        
        selected_dirs = dialog.selectedFiles()
        if not selected_dirs: return
        
        opts_dialog = QDialog(self)
        opts_dialog.setWindowTitle(f"Настройки ({len(selected_dirs)} папок)")
        layout = QVBoxLayout(opts_dialog)
        
        info_label = QLabel(f"Выбрано папок: {len(selected_dirs)}")
        layout.addWidget(info_label)
        
        folders_text = QTextEdit()
        folders_text.setPlainText("\n".join(selected_dirs))
        folders_text.setMaximumHeight(80)
        folders_text.setReadOnly(True)
        layout.addWidget(folders_text)
        
        options_group = QGroupBox("Опции")
        options_layout = QVBoxLayout(options_group)
        
        subfolders_check = QCheckBox("Включить подпапки")
        subfolders_check.setChecked(False)
        options_layout.addWidget(subfolders_check)
        
        extensions_check = QCheckBox("Включить расширения")
        extensions_check.setChecked(True)
        options_layout.addWidget(extensions_check)
        
        fullpath_check = QCheckBox("Полные пути")
        fullpath_check.setChecked(True)
        options_layout.addWidget(fullpath_check)
        
        folder_col_check = QCheckBox("Колонка Исходная папка")
        folder_col_check.setChecked(True)
        options_layout.addWidget(folder_col_check)
        
        size_check = QCheckBox("Размер файлов")
        size_check.setChecked(True)
        options_layout.addWidget(size_check)
        
        date_check = QCheckBox("Дата изменения")
        date_check.setChecked(True)
        options_layout.addWidget(date_check)
        
        layout.addWidget(options_group)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(opts_dialog.accept)
        btn_box.rejected.connect(opts_dialog.reject)
        layout.addWidget(btn_box)
        
        if opts_dialog.exec() != QDialog.DialogCode.Accepted: return
        
        include_subfolders = subfolders_check.isChecked()
        include_extensions = extensions_check.isChecked()
        include_fullpath = fullpath_check.isChecked()
        include_folder_col = folder_col_check.isChecked()
        include_size = size_check.isChecked()
        include_date = date_check.isChecked()

        try:
            all_file_data = []
            for folder_path in selected_dirs:
                base_path = Path(folder_path)
                folder_name = base_path.name
                pattern = "**/*" if include_subfolders else "*"
                
                for item in base_path.glob(pattern):
                    if item.is_file():
                        row = {}
                        if include_folder_col:
                            row["Исходная папка"] = folder_name
                        if include_extensions:
                            row["Имя файла"] = item.name
                        else:
                            row["Имя файла"] = item.stem
                        if include_fullpath:
                            row["Путь"] = str(item.absolute())
                        else:
                            row["Путь"] = str(item.relative_to(base_path))
                        if include_size:
                            size = item.stat().st_size
                            if size < 1024: row["Размер"] = f"{size} B"
                            elif size < 1024*1024: row["Размер"] = f"{size/1024:.1f} KB"
                            elif size < 1024*1024*1024: row["Размер"] = f"{size/(1024*1024):.1f} MB"
                            else: row["Размер"] = f"{size/(1024*1024*1024):.1f} GB"
                        if include_date:
                            mtime = datetime.fromtimestamp(item.stat().st_mtime)
                            row["Дата"] = mtime.strftime("%Y-%m-%d %H:%M")
                        all_file_data.append(row)

            if not all_file_data:
                QMessageBox.information(self, "Информация", "В выбранных папках нет файлов.")
                return

            df = pd.DataFrame(all_file_data)
            cols = []
            if include_folder_col: cols.append("Исходная папка")
            cols.append("Имя файла")
            cols.append("Путь")
            if include_size: cols.extend(["Размер", "Размер (байт)"])
            if include_date: cols.append("Дата")
            df = df[cols]

            save_path, _ = QFileDialog.getSaveFileName(self, "Сохранить список", str(Path.home() / "folders_file_list.xlsx"), "Excel (*.xlsx);;CSV (*.csv)")

            if save_path:
                if save_path.endswith(".csv"):
                    df.to_csv(save_path, index=False, encoding="utf-8-sig")
                else:
                    df.to_excel(save_path, index=False, engine="openpyxl")

                if QMessageBox.question(self, "Успех", f"Сохранено {len(all_file_data)} файлов. Открыть?") == QMessageBox.StandardButton.Yes:
                    subprocess.Popen(f'explorer "{save_path}"')

                if QMessageBox.question(self, "Загрузить?", "Загрузить имена в программу?") == QMessageBox.StandardButton.Yes:
                    self.folder_names = df["Имя файла"].tolist()
                    self.update_preview()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка: {str(e)}")

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        if urls:
            file_path = urls[0].toLocalFile()
            if file_path:
                self.load_file(file_path)


# --- TAB 2: MERGE CARNETS (Full Logic Restored) ---

class ExcelWorker(QThread):
    """Поток для обработки Excel файлов без блокировки UI"""
    progress = pyqtSignal(int)
    file_loaded = pyqtSignal(dict)
    finished_loading = pyqtSignal()
    error = pyqtSignal(str)
    
    def __init__(self, file_paths: List[str]):
        super().__init__()
        self.file_paths = file_paths
        
    def run(self):
        total = len(self.file_paths)
        for idx, file_path in enumerate(self.file_paths):
            try:
                self.progress.emit(int((idx / total) * 100))
                file_data = self._load_excel_file(file_path)
                if file_data:
                    self.file_loaded.emit(file_data)
            except Exception as e:
                self.error.emit(f"Ошибка загрузки {os.path.basename(file_path)}: {str(e)}")
                
        self.progress.emit(100)
        self.finished_loading.emit()
    
    def _load_excel_file(self, file_path: str) -> Optional[dict]:
        """Загрузка Excel файла любого формата включая XLSB"""
        file_ext = Path(file_path).suffix.lower()
        file_name = Path(file_path).name
        
        try:
            if file_ext == '.xlsx':
                return self._load_xlsx(file_path, file_name)
            elif file_ext == '.xls':
                return self._load_xls(file_path, file_name)
            elif file_ext == '.xlsm':
                return self._load_xlsx(file_path, file_name, data_only=False)
            elif file_ext == '.xlsb':
                return self._load_xlsb(file_path, file_name)
            else:
                return None
        except Exception as e:
            raise Exception(f"Ошибка чтения файла: {e}")
    
    def _load_xlsx(self, file_path: str, file_name: str, data_only=True) -> dict:
        """Загрузка XLSX и XLSM"""
        if not openpyxl: raise ImportError("openpyxl не установлен")
        wb = load_workbook(file_path, data_only=data_only, read_only=True)
        sheets = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            is_hidden = ws.sheet_state != 'visible'
            max_row = ws.max_row if ws.max_row else 1
            max_col = ws.max_column if ws.max_column else 1
            
            headers = self._get_headers_from_worksheet(ws, 1, max_col)
            
            sheets.append({
                'name': sheet_name,
                'hidden': is_hidden,
                'selected': False,
                'header_row': 1,
                'max_row': max_row,
                'max_col': max_col,
                'headers': headers
            })
        
        wb.close()
        return {
            'name': file_name,
            'path': file_path,
            'type': 'xlsx',
            'sheets': sheets
        }
    
    def _load_xls(self, file_path: str, file_name: str) -> dict:
        """Загрузка XLS (старый формат)"""
        if not xlrd: raise ImportError("xlrd не установлен")
        wb = xlrd.open_workbook(file_path, formatting_info=True)
        sheets = []
        
        for idx, sheet_name in enumerate(wb.sheet_names()):
            ws = wb.sheet_by_index(idx)
            is_hidden = wb.sheet_hidden(idx) if hasattr(wb, 'sheet_hidden') else False
            
            headers = []
            if ws.nrows > 0:
                for col in range(ws.ncols):
                    cell_value = ws.cell_value(0, col)
                    headers.append(str(cell_value) if cell_value else f"Column {col+1}")
            
            sheets.append({
                'name': sheet_name,
                'hidden': is_hidden,
                'selected': False,
                'header_row': 1,
                'max_row': ws.nrows,
                'max_col': ws.ncols,
                'headers': headers
            })
        
        wb.release_resources()
        return {
            'name': file_name,
            'path': file_path,
            'type': 'xls',
            'sheets': sheets
        }
    
    def _load_xlsb(self, file_path: str, file_name: str) -> dict:
        """Загрузка XLSB (бинарный формат Excel)"""
        if not pyxlsb: raise ImportError("pyxlsb не установлен")
        try:
            with open_xlsb(file_path) as wb:
                sheets = []
                sheet_names = list(wb.sheets)
                
                for sheet_name in sheet_names:
                    try:
                        with wb.get_sheet(sheet_name) as sheet:
                            max_row = 0
                            max_col = 0
                            headers = []
                            
                            first_row_data = {}
                            for row_idx, row in enumerate(sheet.rows()):
                                if row_idx == 0:
                                    for cell in row:
                                        if cell:
                                            col_idx = cell.col_idx if hasattr(cell, 'col_idx') else 0
                                            max_col = max(max_col, col_idx + 1)
                                            first_row_data[col_idx] = str(cell.v) if cell.v else f"Column {col_idx+1}"
                                
                                max_row += 1
                                if row_idx >= 100000:
                                    break
                            
                            if first_row_data:
                                for i in range(max_col):
                                    headers.append(first_row_data.get(i, f"Column {i+1}"))
                            
                            if not headers and max_col > 0:
                                headers = [f"Column {i+1}" for i in range(max_col)]
                            
                            sheets.append({
                                'name': sheet_name,
                                'hidden': False,
                                'selected': False,
                                'header_row': 1,
                                'max_row': max_row,
                                'max_col': max_col if max_col > 0 else 1,
                                'headers': headers
                            })
                    except Exception as sheet_error:
                        sheets.append({
                            'name': sheet_name,
                            'hidden': False,
                            'selected': False,
                            'header_row': 1,
                            'max_row': 0,
                            'max_col': 0,
                            'headers': [],
                            'error': str(sheet_error)
                        })
                
                return {
                    'name': file_name,
                    'path': file_path,
                    'type': 'xlsb',
                    'sheets': sheets
                }
        except Exception as e:
            raise Exception(f"XLSB ошибка: {e}")
    
    def _get_headers_from_worksheet(self, ws, row_num: int, max_col: int) -> List[str]:
        """Получение заголовков из указанной строки"""
        headers = []
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            value = cell.value if cell.value else f"Column {col}"
            headers.append(str(value))
        return headers


class MergeWorker(QThread):
    """Поток для объединения файлов Karnet"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(bool, str, int)
    error = pyqtSignal(str, str)
    log_message = pyqtSignal(str)
    
    def __init__(self, mapping_file: str, source_folder: str, output_file: str):
        super().__init__()
        self.mapping_file = mapping_file
        self.source_folder = source_folder
        self.output_file = output_file
        self._is_running = True
        
        self.target_headers = [
            "Табельный номер", "Фактическая должность", "Участок по факту", "Прораб",
            "Позиция", "ДЕНЬ/НОЧЬ", "да/нет", "ОП/Проект", "Оценка", "ФИО",
            "Должность (по утвержденном списку в параметрах)",
            "Фактическое структурное подразделение (по утвержденном списку в параметрах)",
            "Гражданство", "Должность", "Разряд", "Подразделение",
            "Сотрудник официально трудоустроен на проекте (в 1с Территория): (данные на последний день в месяце)",
            "Удостоверение Серия", "Удостоверение Номер", "Компания", "ИТР",
            "Сектор", "Вид работ", "Виза/ гражданство"
        ] + [str(i) for i in range(1, 32)] + ["ПРИМЕЧАНИЕ", "Итого произв. Часов", "Итого актираных часов"]
    
    def run(self):
        try:
            self.log_message.emit("=== Начало объединения файлов ===")
            
            if not os.path.exists(self.mapping_file):
                self.finished.emit(False, f"Файл сопоставления не найден: {self.mapping_file}", 0)
                return
            
            self.log_message.emit(f"Загрузка сопоставления: {os.path.basename(self.mapping_file)}")
            
            try:
                df_map = pd.read_excel(self.mapping_file, dtype={"Столбец": int})
            except Exception as e:
                self.finished.emit(False, f"Ошибка загрузки сопоставления: {str(e)}", 0)
                return
            
            df_map = df_map[
                (df_map["Целевое поле"].notna()) &
                (df_map["Целевое поле"].astype(str).str.strip() != "")
            ].copy()
            
            self.log_message.emit(f"Записей после фильтрации: {len(df_map)}")
            
            file_config = {}
            for _, row in df_map.iterrows():
                fname = str(row["Имя файла"]).strip() if pd.notna(row["Имя файла"]) else ""
                target_header = str(row["Целевое поле"]).strip()
                col_1based = row["Столбец"]
                
                if not fname or not target_header: continue
                if target_header not in self.target_headers: continue
                
                if fname not in file_config:
                    file_config[fname] = {}
                file_config[fname][col_1based - 1] = target_header
            
            if not file_config:
                self.finished.emit(False, "Нет данных для объединения (пустой файл сопоставления)", 0)
                return
            
            self.log_message.emit(f"Файлов для обработки: {len(file_config)}")
            
            all_rows = []
            processed_files = 0
            total_files = len(file_config)
            
            for i, (filename, col_map) in enumerate(file_config.items()):
                if not self._is_running: break
                
                progress = int((i / total_files) * 100)
                self.progress.emit(progress, f"Обработка {filename}")
                
                filepath = os.path.join(self.source_folder, filename)
                if not os.path.exists(filepath):
                    self.log_message.emit(f"⚠️ Файл не найден: {filename}")
                    continue
                
                try:
                    df = pd.read_excel(filepath, sheet_name='Karnet', header=1, dtype=str)
                    
                    max_col = max(col_map.keys()) if col_map else -1
                    while len(df.columns) <= max_col:
                        df[len(df.columns)] = None
                    
                    file_rows = 0
                    for _, row in df.iterrows():
                        fio_val = None
                        for col_idx, hdr in col_map.items():
                            if hdr == "ФИО" and col_idx < len(row):
                                fio_val = row.iloc[col_idx]
                                break
                        
                        if pd.isna(fio_val) or str(fio_val).strip() == "":
                            continue
                        
                        new_row = {h: None for h in self.target_headers}
                        for col_idx, hdr in col_map.items():
                            if col_idx < len(row):
                                value = row.iloc[col_idx]
                                if pd.notna(value) and str(value).strip().lower() != 'nan':
                                    new_row[hdr] = str(value).strip()
                        
                        all_rows.append(new_row)
                        file_rows += 1
                    
                    processed_files += 1
                    self.log_message.emit(f"✅ {filename}: {file_rows} строк")
                    
                except Exception as e:
                    self.error.emit(filename, str(e))
                    self.log_message.emit(f"❌ Ошибка {filename}: {str(e)}")
            
            if self._is_running and all_rows:
                result_df = pd.DataFrame(all_rows, columns=self.target_headers)
                result_df.to_excel(self.output_file, index=False)
                self.log_message.emit(f"✅ Сохранено: {self.output_file}")
                self.log_message.emit(f"📊 Всего строк: {len(all_rows)}")
                self.finished.emit(True, f"Обработано {processed_files} файлов, {len(all_rows)} строк", len(all_rows))
            elif not self._is_running:
                self.finished.emit(False, "Обработка остановлена", 0)
            else:
                self.finished.emit(False, "Нет данных для сохранения", 0)
                
        except Exception as e:
            self.finished.emit(False, f"Критическая ошибка: {str(e)}", 0)
    
    def stop(self):
        self._is_running = False


class MergeCarnetTab(BaseTabWidget):
    """Вкладка 2: Объединение Карнетов"""
    REQUIRED_FIELDS = [
        "Табельный номер", "Фактическая должность", "Участок по факту", "Прораб",
        "Позиция", "ДЕНЬ/НОЧЬ", "да/нет", "ОП/Проект", "Оценка", "ФИО",
        "Должность (по утвержденном списку в параметрах)",
        "Фактическое структурное подразделение (по утвержденном списку в параметрах)",
        "Гражданство", "Должность", "Разряд", "Подразделение",
        "Сотрудник официально трудоустроен на проекте (в 1с Территория): (данные на последний день в месяце)",
        "Удостоверение Серия", "Удостоверение Номер", "Компания", "ИТР",
        "Сектор", "Вид работ", "Виза/ гражданство",
        "1", "2", "3", "4", "5", "6", "7", "8", "9", "10",
        "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
        "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31",
        "ПРИМЕЧАНИЕ", "Итого произв. Часов", "Итого актираных часов"
    ]
    
    ALL_TARGET_FIELDS = REQUIRED_FIELDS.copy()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.files_data: List[dict] = []
        self.current_worker: Optional[ExcelWorker] = None
        self.merge_worker: Optional[MergeWorker] = None

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        header = QLabel("🔗 Объединение Карнетов")
        header.setProperty("class", "title")
        layout.addWidget(header)

        # Toolbar
        toolbar_layout = QHBoxLayout()
        open_btn = QPushButton("📂 Файлы")
        open_btn.clicked.connect(self._open_files_dialog)
        toolbar_layout.addWidget(open_btn)
        
        folder_btn = QPushButton("📁 Папка")
        folder_btn.clicked.connect(self._open_folder_dialog)
        toolbar_layout.addWidget(folder_btn)
        
        toolbar_layout.addStretch()
        
        self.generate_btn = QPushButton("✨ Заголовки")
        self.generate_btn.clicked.connect(self._generate_headers)
        self.generate_btn.setEnabled(False)
        toolbar_layout.addWidget(self.generate_btn)
        
        self.target_btn = QPushButton("🎯 Целевое поле")
        self.target_btn.clicked.connect(self._fill_target_fields)
        self.target_btn.setEnabled(False)
        toolbar_layout.addWidget(self.target_btn)
        
        self.check_btn = QPushButton("✅ Проверить")
        self.check_btn.clicked.connect(self._check_target_fields)
        self.check_btn.setEnabled(False)
        toolbar_layout.addWidget(self.check_btn)
        
        toolbar_layout.addStretch()
        
        merge_btn = QPushButton("⚡ Объединить")
        merge_btn.setProperty("class", "success")
        merge_btn.clicked.connect(self._show_merge_dialog)
        toolbar_layout.addWidget(merge_btn)
        
        layout.addLayout(toolbar_layout)

        # Splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left: File Tree
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0,0,0,0)
        
        header_layout = QHBoxLayout()
        title_label = QLabel("📁 Структура файлов")
        title_label.setProperty("class", "heading")
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        self.file_count_label = QLabel("0")
        self.file_count_label.setProperty("class", "caption")
        header_layout.addWidget(self.file_count_label)
        left_layout.addLayout(header_layout)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        left_layout.addWidget(self.progress_bar)
        
        self.file_tree = QTreeWidget()
        self.file_tree.setHeaderHidden(True)
        self.file_tree.setIndentation(20)
        self.file_tree.setAnimated(True)
        self.file_tree.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        left_layout.addWidget(self.file_tree)
        
        btn_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("☑️ Все")
        self.select_all_btn.setEnabled(False)
        self.select_all_btn.clicked.connect(self._select_all_sheets)
        btn_layout.addWidget(self.select_all_btn)
        
        btn_layout.addStretch()
        clear_list_btn = QPushButton("🗑️ Очистить")
        clear_list_btn.clicked.connect(self._clear_file_list)
        btn_layout.addWidget(clear_list_btn)
        left_layout.addLayout(btn_layout)
        
        splitter.addWidget(left_panel)

        # Right: Data Table
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0,0,0,0)
        
        toolbar_table = QHBoxLayout()
        self.row_count_label = QLabel("Строк: 0")
        self.row_count_label.setProperty("class", "caption")
        toolbar_table.addWidget(self.row_count_label)
        
        toolbar_table.addStretch()
        clear_btn = QPushButton("🗑️ Очистить")
        clear_btn.clicked.connect(self._clear_workspace)
        toolbar_table.addWidget(clear_btn)
        
        export_btn = QPushButton("💾 Экспорт")
        export_btn.clicked.connect(self._export_table)
        toolbar_table.addWidget(export_btn)
        
        right_layout.addLayout(toolbar_table)
        
        self.data_table = QTableWidget()
        self.data_table.setColumnCount(6)
        self.data_table.setHorizontalHeaderLabels([
            "Имя файла", "Заголовок таблицы", "Строка заголовка", 
            "Столбец", "Статус столбца", "Целевое поле"
        ])
        self.data_table.horizontalHeader().setStretchLastSection(True)
        self.data_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.data_table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.data_table.setAlternatingRowColors(True)
        right_layout.addWidget(self.data_table)
        
        splitter.addWidget(right_panel)
        splitter.setSizes([400, 1000])
        layout.addWidget(splitter)

    def _open_files_dialog(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Выберите Excel файлы", "", "Excel Files (*.xlsx *.xls *.xlsm *.xlsb);;All Files (*)"
        )
        if files:
            self._load_files(files)
    
    def _open_folder_dialog(self):
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку с Excel файлами")
        if folder:
            files = []
            for ext in ['*.xlsx', '*.xls', '*.xlsm', '*.xlsb']:
                files.extend([str(p) for p in Path(folder).glob(ext)])
            if files:
                self._load_files(files)
            else:
                QMessageBox.information(self, "Информация", "В выбранной папке не найдено Excel файлов")
    
    def _load_files(self, file_paths: List[str]):
        self.files_data = []
        self.file_tree.clear()
        self.data_table.setRowCount(0)
        
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.generate_btn.setEnabled(False)
        self.target_btn.setEnabled(False)
        self.check_btn.setEnabled(False)
        self.select_all_btn.setEnabled(False)
        
        self.current_worker = ExcelWorker(file_paths)
        self.current_worker.progress.connect(self.progress_bar.setValue)
        self.current_worker.file_loaded.connect(self._add_file_to_tree)
        self.current_worker.finished_loading.connect(self._loading_finished)
        self.current_worker.error.connect(self._show_error)
        self.current_worker.start()
    
    def _add_file_to_tree(self, file_data: dict):
        self.files_data.append(file_data)
        
        file_item = QTreeWidgetItem(self.file_tree)
        type_icons = {'xlsx': '📊', 'xls': '📄', 'xlsm': '⚙️', 'xlsb': '⚡'}
        icon = type_icons.get(file_data['type'], '📊')
        
        file_item.setText(0, f"{icon} {file_data['name']}")
        file_item.setData(0, Qt.ItemDataRole.UserRole, file_data)
        file_item.setExpanded(True)
        
        for idx, sheet in enumerate(file_data['sheets']):
            sheet_item = QTreeWidgetItem(file_item)
            
            container = QWidget()
            l = QHBoxLayout(container)
            l.setContentsMargins(20, 0, 4, 0)
            l.setSpacing(8)
            
            checkbox = QCheckBox()
            checkbox.setChecked(sheet['selected'])
            checkbox.setFixedSize(14, 14)
            # Важно: используем замыкание для захвата правильных индексов
            checkbox.stateChanged.connect(
                lambda state, f=len(self.files_data)-1, s=idx: self._toggle_sheet(f, s, state)
            )
            l.addWidget(checkbox)
            
            sheet_icon = "🚫" if sheet.get('hidden') else "📄"
            name_label = QLabel(f"{sheet_icon} {sheet['name']}")
            l.addWidget(name_label)
            
            l.addStretch()
            
            row_label = QLabel("Строка:")
            l.addWidget(row_label)
            
            combo = QComboBox()
            combo.addItems([str(i) for i in range(1, 16)])
            combo.setCurrentText(str(sheet['header_row']))
            combo.setFixedWidth(45)
            combo.setFixedHeight(20)
            # Важно: используем замыкание для захвата правильных индексов
            combo.currentTextChanged.connect(
                lambda text, f=len(self.files_data)-1, s=idx: self._update_header_row(f, s, text)
            )
            l.addWidget(combo)
            
            size_text = f"({len(sheet.get('headers', []))}×{sheet['max_row']})"
            size_label = QLabel(size_text)
            l.addWidget(size_label)
            
            self.file_tree.setItemWidget(sheet_item, 0, container)
        
        self.file_count_label.setText(str(len(self.files_data)))
    
    def _toggle_sheet(self, file_idx: int, sheet_idx: int, state: int):
        if 0 <= file_idx < len(self.files_data):
            self.files_data[file_idx]['sheets'][sheet_idx]['selected'] = (state == 2)
            self._update_buttons_state()
    
    def _update_header_row(self, file_idx: int, sheet_idx: int, value: str):
        if 0 <= file_idx < len(self.files_data):
            try:
                new_row = int(value)
                self.files_data[file_idx]['sheets'][sheet_idx]['header_row'] = new_row
                self._reload_headers_for_sheet(file_idx, sheet_idx, new_row)
            except ValueError:
                pass
    
    def _reload_headers_for_sheet(self, file_idx: int, sheet_idx: int, row_num: int):
        file_data = self.files_data[file_idx]
        sheet_data = file_data['sheets'][sheet_idx]
        file_path = file_data['path']
        sheet_name = sheet_data['name']
        
        try:
            new_headers = []
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext in ['.xlsx', '.xlsm']:
                wb = load_workbook(file_path, data_only=True, read_only=True)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    max_col = ws.max_column if ws.max_column else 1
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=row_num, column=col)
                        val = cell.value if cell.value else f"Column {col}"
                        new_headers.append(str(val))
                wb.close()
                
            elif file_ext == '.xls':
                wb = xlrd.open_workbook(file_path)
                sheet_idx_num = wb.sheet_names().index(sheet_name)
                ws = wb.sheet_by_index(sheet_idx_num)
                if row_num <= ws.nrows:
                    for col in range(ws.ncols):
                        cell_value = ws.cell_value(row_num - 1, col)
                        new_headers.append(str(cell_value) if cell_value else f"Column {col+1}")
                wb.release_resources()
                
            elif file_ext == '.xlsb':
                with open_xlsb(file_path) as wb:
                    with wb.get_sheet(sheet_name) as sheet:
                        current_row = 0
                        for row in sheet.rows():
                            current_row += 1
                            if current_row == row_num:
                                row_data = {}
                                for cell in row:
                                    if cell:
                                        col_idx = cell.col_idx if hasattr(cell, 'col_idx') else 0
                                        row_data[col_idx] = str(cell.v) if cell.v else f"Column {col_idx+1}"
                                if row_data:
                                    max_col = max(row_data.keys()) + 1
                                    new_headers = [row_data.get(i, f"Column {i+1}") for i in range(max_col)]
                                break
            
            if new_headers:
                sheet_data['headers'] = new_headers
                sheet_data['max_col'] = len(new_headers)
                
        except Exception as e:
            print(f"Ошибка перезагрузки заголовков: {e}")

    def _update_buttons_state(self):
        has_selection = any(sheet['selected'] for file in self.files_data for sheet in file['sheets'])
        self.generate_btn.setEnabled(has_selection)
        self.target_btn.setEnabled(has_selection)
        self.check_btn.setEnabled(has_selection)
    
    def _loading_finished(self):
        self.progress_bar.setVisible(False)
        self.select_all_btn.setEnabled(True)
    
    def _select_all_sheets(self):
        def process_item(item):
            widget = self.file_tree.itemWidget(item, 0)
            if widget:
                checkbox = widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.setChecked(True)
            for i in range(item.childCount()):
                process_item(item.child(i))
        
        for i in range(self.file_tree.topLevelItemCount()):
            process_item(self.file_tree.topLevelItem(i))
    
    def _clear_file_list(self):
        self.files_data = []
        self.file_tree.clear()
        self.generate_btn.setEnabled(False)
        self.target_btn.setEnabled(False)
        self.check_btn.setEnabled(False)
        self.select_all_btn.setEnabled(False)
        self.file_count_label.setText("0")
    
    def _show_error(self, message: str):
        QMessageBox.warning(self, "Ошибка загрузки", message)
    
    def _generate_headers(self):
        self.data_table.setRowCount(0)
        row = 0
        for file_idx, file in enumerate(self.files_data):
            for sheet_idx, sheet in enumerate(file['sheets']):
                if not sheet['selected']: continue
                
                headers = sheet.get('headers', [])
                header_row = sheet['header_row']
                
                if not headers:
                    headers = [f"Column {i+1}" for i in range(sheet['max_col'])]
                
                for col_idx, header_name in enumerate(headers):
                    self.data_table.insertRow(row)
                    
                    self.data_table.setItem(row, 0, QTableWidgetItem(file['name']))
                    
                    header_item = QTableWidgetItem(str(header_name))
                    header_item.setToolTip(f"Лист: {sheet['name']}, Строка: {header_row}, Столбец: {col_idx + 1}")
                    self.data_table.setItem(row, 1, header_item)
                    
                    row_item = QTableWidgetItem(str(header_row))
                    row_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.data_table.setItem(row, 2, row_item)
                    
                    col_num_item = QTableWidgetItem(str(col_idx + 1))
                    col_num_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.data_table.setItem(row, 3, col_num_item)
                    
                    status = "Скрытый" if sheet['hidden'] else "Видимый"
                    status_item = QTableWidgetItem(status)
                    if sheet['hidden']:
                        status_item.setBackground(QColor("#fed7d7"))
                    else:
                        status_item.setBackground(QColor("#c6f6d5"))
                    status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.data_table.setItem(row, 4, status_item)
                    
                    target_item = QTableWidgetItem("")
                    target_item.setFlags(target_item.flags() | Qt.ItemFlag.ItemIsEditable)
                    self.data_table.setItem(row, 5, target_item)
                    
                    row += 1
        
        self.row_count_label.setText(f"Строк: {row}")
    
    def _fill_target_fields(self):
        if self.data_table.rowCount() == 0: return
        
        filled_count = 0
        for row in range(self.data_table.rowCount()):
            header_item = self.data_table.item(row, 1)
            target_item = self.data_table.item(row, 5)
            
            if header_item and target_item:
                header_text = header_item.text().strip()
                
                if header_text in self.ALL_TARGET_FIELDS:
                    target_item.setText(header_text)
                    filled_count += 1
                else:
                    target_item.setText("")
    
    def _check_target_fields(self):
        if self.data_table.rowCount() == 0: return
        
        file_fields: Dict[str, Set[str]] = {}
        
        for row in range(self.data_table.rowCount()):
            file_item = self.data_table.item(row, 0)
            target_item = self.data_table.item(row, 5)
            
            if file_item and target_item:
                file_name = file_item.text()
                target_value = target_item.text().strip()
                
                if file_name not in file_fields:
                    file_fields[file_name] = set()
                
                if target_value:
                    file_fields[file_name].add(target_value)
        
        check_results = []
        all_passed = True
        
        for file_name, fields in file_fields.items():
            missing = []
            for required in self.REQUIRED_FIELDS:
                if required not in fields:
                    missing.append(required)
            
            if missing:
                all_passed = False
                check_results.append({
                    'file': file_name,
                    'status': '❌ Не пройдена',
                    'missing': ', '.join(missing) if len(missing) <= 3 else f"{len(missing)} полей"
                })
            else:
                check_results.append({
                    'file': file_name,
                    'status': '✅ Пройдена',
                    'missing': ''
                })
        
        self._show_check_results(check_results, all_passed)
    
    def _show_check_results(self, results: List[dict], all_passed: bool):
        dialog = QDialog(self)
        dialog.setWindowTitle("Результаты проверки")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout(dialog)
        text = QTextEdit()
        text.setReadOnly(True)
        
        html = "<h3>Проверка обязательных полей</h3>"
        html += "<table border='1' cellpadding='5' style='border-collapse: collapse; width: 100%;'>"
        html += "<tr style='background-color: #f0f0f0;'><th>Название файла</th><th>Статус</th><th>Отсутствуют значения</th></tr>"
        
        for result in results:
            color = "#c6f6d5" if result['status'] == '✅ Пройдена' else "#fed7d7"
            html += f"<tr style='background-color: {color};'>"
            html += f"<td>{result['file']}</td>"
            html += f"<td>{result['status']}</td>"
            html += f"<td>{result['missing']}</td>"
            html += "</tr>"
        
        html += "</table>"
        text.setHtml(html)
        layout.addWidget(text)
        
        export_btn = QPushButton("💾 Сохранить отчет")
        export_btn.clicked.connect(lambda: self._export_check_results(results))
        layout.addWidget(export_btn)
        
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn)
        
        dialog.exec()
    
    def _export_check_results(self, results: List[dict]):
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить отчет", "check_results.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            try:
                df = pd.DataFrame(results)
                df.to_excel(file_path, index=False)
                QMessageBox.information(self, "Успех", "Отчет сохранен")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def _clear_workspace(self):
        self.data_table.setRowCount(0)
        self.row_count_label.setText("Строк: 0")
    
    def _export_table(self):
        if self.data_table.rowCount() == 0: return
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить", "headers_export.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            try:
                data = []
                for row in range(self.data_table.rowCount()):
                    row_data = []
                    for col in range(self.data_table.columnCount()):
                        item = self.data_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    data.append(row_data)
                
                df = pd.DataFrame(data, columns=[
                    "Имя файла", "Заголовок таблицы", "Строка заголовка",
                    "Столбец", "Статус столбца", "Целевое поле"
                ])
                df.to_excel(file_path, index=False)
                QMessageBox.information(self, "Успех", "Данные сохранены")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    # --- MERGE LOGIC ---
    def _show_merge_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("⚡ Объединение файлов Karnet")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout(dialog)
        
        mapping_group = QGroupBox("📋 Файл сопоставления")
        mapping_layout = QHBoxLayout()
        self.merge_mapping_path = QLineEdit()
        self.merge_mapping_path.setPlaceholderText("Выберите файл сопоставления...")
        mapping_layout.addWidget(self.merge_mapping_path)
        
        mapping_btn = QPushButton("📂 Обзор")
        mapping_btn.clicked.connect(lambda: self._browse_merge_mapping())
        mapping_layout.addWidget(mapping_btn)
        
        mapping_group.setLayout(mapping_layout)
        layout.addWidget(mapping_group)
        
        source_group = QGroupBox("📁 Папка с файлами Karnet")
        source_layout = QHBoxLayout()
        self.merge_source_path = QLineEdit()
        self.merge_source_path.setPlaceholderText("Выберите папку...")
        source_layout.addWidget(self.merge_source_path)
        
        source_btn = QPushButton("📂 Обзор")
        source_btn.clicked.connect(lambda: self._browse_merge_source())
        source_layout.addWidget(source_btn)
        
        source_group.setLayout(source_layout)
        layout.addWidget(source_group)
        
        output_group = QGroupBox("💾 Файл результата")
        output_layout = QHBoxLayout()
        self.merge_output_path = QLineEdit()
        self.merge_output_path.setText("Итоговый_карнет.xlsx")
        output_layout.addWidget(self.merge_output_path)
        
        output_btn = QPushButton("📂 Обзор")
        output_btn.clicked.connect(lambda: self._browse_merge_output())
        output_layout.addWidget(output_btn)
        
        output_group.setLayout(output_layout)
        layout.addWidget(output_group)
        
        self.merge_progress = QProgressBar()
        self.merge_progress.setVisible(False)
        layout.addWidget(self.merge_progress)
        
        self.merge_log = QTextEdit()
        self.merge_log.setMaximumHeight(150)
        layout.addWidget(self.merge_log)
        
        btn_layout = QHBoxLayout()
        self.merge_start_btn = QPushButton("⚡ Начать объединение")
        self.merge_start_btn.setProperty("class", "success")
        self.merge_start_btn.clicked.connect(lambda: self._start_merge(dialog))
        btn_layout.addWidget(self.merge_start_btn)
        
        self.merge_stop_btn = QPushButton("⏹ Остановить")
        self.merge_stop_btn.setEnabled(False)
        self.merge_stop_btn.clicked.connect(self._stop_merge)
        btn_layout.addWidget(self.merge_stop_btn)
        
        btn_layout.addStretch()
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(dialog.close)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        dialog.exec()
    
    def _browse_merge_mapping(self):
        file, _ = QFileDialog.getOpenFileName(self, "Выберите файл", "", "Excel Files (*.xlsx *.xls);;All Files (*)")
        if file:
            self.merge_mapping_path.setText(file)

    def _browse_merge_source(self):
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку")
        if folder:
            self.merge_source_path.setText(folder)

    def _browse_merge_output(self):
        file, _ = QFileDialog.getSaveFileName(self, "Сохранить", "Итоговый_карнет.xlsx", "Excel Files (*.xlsx)")
        if file:
            self.merge_output_path.setText(file)

    def _start_merge(self, dialog):
        mapping_file = self.merge_mapping_path.text()
        source_folder = self.merge_source_path.text()
        output_file = self.merge_output_path.text()
        
        if not mapping_file or not os.path.exists(mapping_file):
            QMessageBox.warning(self, "Ошибка", "Укажите файл сопоставления")
            return
        if not source_folder or not os.path.isdir(source_folder):
            QMessageBox.warning(self, "Ошибка", "Укажите папку с файлами")
            return
        
        self.merge_log.clear()
        self.merge_log.append("🚀 Начало объединения...")
        self.merge_progress.setVisible(True)
        self.merge_progress.setValue(0)
        self.merge_start_btn.setEnabled(False)
        self.merge_stop_btn.setEnabled(True)
        
        self.merge_worker = MergeWorker(mapping_file, source_folder, output_file)
        self.merge_worker.progress.connect(self._update_merge_progress)
        self.merge_worker.finished.connect(self._merge_finished)
        self.merge_worker.error.connect(self._merge_error)
        self.merge_worker.log_message.connect(self._merge_log)
        self.merge_worker.start()

    def _update_merge_progress(self, value: int, message: str):
        self.merge_progress.setValue(value)
        self.merge_log.append(message)

    def _merge_log(self, message: str):
        self.merge_log.append(message)

    def _merge_error(self, file_name: str, error: str):
        self.merge_log.append(f"❌ Ошибка в {file_name}: {error}")

    def _merge_finished(self, success: bool, message: str, row_count: int):
        self.merge_progress.setVisible(False)
        self.merge_start_btn.setEnabled(True)
        self.merge_stop_btn.setEnabled(False)
        
        if success:
            self.merge_log.append(f"✅ {message}")
            QMessageBox.information(self, "Готово", message)
        else:
            self.merge_log.append(f"❌ {message}")
            QMessageBox.critical(self, "Ошибка", message)

    def _stop_merge(self):
        if self.merge_worker and self.merge_worker.isRunning():
            self.merge_worker.stop()
            self.merge_worker.wait()
            self.merge_log.append("⏹ Обработка остановлена")
            self.merge_start_btn.setEnabled(True)
            self.merge_stop_btn.setEnabled(False)


# --- TAB 3: TICKET PARSER (Full Logic Restored) ---

@dataclass
class TicketData:
    passenger: str = 'Не указано'
    ticket_number: str = 'Не указано'
    order_number: str = 'Не указано'
    issue_date: str = 'Не указано'
    carrier: str = 'Не указано'
    flight_number: str = 'Не указано'
    departure_time: str = 'Не указано'
    departure_date: str = 'Не указано'
    route: str = 'Не указано'
    arrival_airport: str = 'Не указано'
    arrival_time: str = 'Не указано'
    arrival_date: str = 'Не указано'
    total_price: str = 'Не указано'
    currency: str = 'RUB'
    source_file: str = ''
    
    def to_dict(self) -> Dict[str, str]:
        return {
            'Пассажир': self.passenger,
            'Номер билета': self.ticket_number,
            'Номер заказа': self.order_number,
            'Дата выдачи': self.issue_date,
            'Перевозчик': self.carrier,
            'Номер рейса': self.flight_number,
            'Время отправления': self.departure_time,
            'Дата отправления': self.departure_date,
            'Маршрут': self.route,
            'Аэропорт прибытия': self.arrival_airport,
            'Время прибытия': self.arrival_time,
            'Дата прибытия': self.arrival_date,
            'Стоимость': self.total_price,
            'Валюта': self.currency,
            'Источник': self.source_file
        }


class TicketParser:
    """Парсер авиабилетов из PDF и изображений"""
    
    IATA_CODES = {
        'SVO': 'МОСКВА', 'DME': 'МОСКВА', 'VKO': 'МОСКВА', 'ZIA': 'МОСКВА',
        'LED': 'САНКТ-ПЕТЕРБУРГ', 'AER': 'СОЧИ', 'KRR': 'КРАСНОДАР',
        'ROV': 'РОСТОВ-НА-ДОНУ', 'KZN': 'КАЗАНЬ', 'UFA': 'УФА',
        'KUF': 'САМАРА', 'GOJ': 'НИЖНИЙ НОВГОРОД', 'SVX': 'ЕКАТЕРИНБУРГ',
        'CEK': 'ЧЕЛЯБИНСК', 'TJM': 'ТЮМЕНЬ', 'OVB': 'НОВОСИБИРСК',
        'OMS': 'ОМСК', 'KJA': 'КРАСНОЯРСК', 'IKT': 'ИРКУТСК',
        'TAS': 'ТАШКЕНТ', 'TMJ': 'ТЕРМЕЗ', 'BHK': 'БУХАРА',
        'SKD': 'САМАРКАНД', 'NVI': 'НАВОИ', 'NCU': 'НУКУС',
        'UGC': 'УРГЕНЧ', 'FEG': 'ФЕРГАНА', 'KSQ': 'КАРШИ',
        'AZN': 'АНДИЖАН', 'NMA': 'НАМАНГАН', 'ALA': 'АЛМАТЫ',
        'NQZ': 'АСТАНА', 'CIT': 'ШЫМКЕНТ', 'MSQ': 'МИНСК',
        'EVN': 'ЕРЕВАН', 'GYD': 'БАКУ', 'TBS': 'ТБИЛИСИ',
        'FRU': 'БИШКЕК', 'DYU': 'ДУШАНБЕ', 'LBD': 'ХУДЖАНД',
        'IST': 'СТАМБУЛ', 'SAW': 'СТАМБУЛ', 'AYT': 'АНТАЛЬЯ',
        'DXB': 'ДУБАЙ', 'AUH': 'АБУ-ДАБИ'
    }
    
    MONTH_MAP = {
        'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04', 'MAY': '05', 'JUN': '06',
        'JUL': '07', 'AUG': '08', 'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12',
        'ЯНВ': '01', 'ФЕВ': '02', 'МАР': '03', 'АПР': '04', 'МАЙ': '05', 'ИЮН': '06',
        'ИЮЛ': '07', 'АВГ': '08', 'СЕН': '09', 'ОКТ': '10', 'НОЯ': '11', 'ДЕК': '12',
        'MAP': '03'
    }
    
    def parse_file(self, file_path: str) -> TicketData:
        path = Path(file_path)
        filename = path.stem
        
        if path.suffix.lower() == '.pdf':
            text = self._extract_text_from_pdf(file_path)
        elif path.suffix.lower() in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
            text = self._extract_text_from_image(file_path)
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {path.suffix}")
        
        return self._parse_text(text, filename)
    
    def _extract_text_from_pdf(self, pdf_path: str) -> str:
        if not fitz: raise ImportError("PyMuPDF не установлен")
        try:
            doc = fitz.open(pdf_path)
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            return text
        except Exception as e:
            logger.error(f"Ошибка чтения PDF {pdf_path}: {e}")
            return ""
    
    def _extract_text_from_image(self, image_path: str) -> str:
        # Placeholder for OCR logic (requires pytesseract)
        return ""
    
    def _parse_text(self, text: str, filename: str) -> TicketData:
        data = TicketData()
        data.source_file = filename
        
        # 1. ПАССАЖИР
        data.passenger = self._extract_passenger(filename, text)
        
        # 2. НОМЕР БИЛЕТА
        data.ticket_number = self._extract_ticket_number(text)
        
        # 3. НОМЕР ЗАКАЗА
        data.order_number = self._extract_order_number(text, filename)
        
        # 4. ДАТА ВЫДАЧИ
        data.issue_date = self._extract_issue_date(text)
        
        # 5. ПЕРЕВОЗЧИК
        data.carrier = self._extract_carrier(text)
        
        # 6. НОМЕР РЕЙСА
        data.flight_number = self._extract_flight_number(text)
        
        # 7. ДАТЫ И ВРЕМЯ
        flight_info = self._extract_flight_times(text, data.issue_date)
        data.departure_time = flight_info.get('dep_time', 'Не указано')
        data.departure_date = flight_info.get('dep_date', 'Не указано')
        data.arrival_time = flight_info.get('arr_time', 'Не указано')
        data.arrival_date = flight_info.get('arr_date', 'Не указано')
        
        # 8. МАРШРУТ
        route_info = self._extract_route(text)
        data.route = route_info.get('route', 'Не указано')
        data.arrival_airport = route_info.get('arrival', 'Не указано')
        
        # 9. СТОИМОСТЬ
        price_info = self._extract_price(text)
        data.total_price = price_info.get('price', 'Не указано')
        data.currency = price_info.get('currency', 'RUB')
        
        return data
    
    def _extract_passenger(self, filename: str, text: str) -> str:
        name_match = re.match(r'^([A-Za-zА-Яа-я]+ [A-Za-zА-Яа-я]+)', filename)
        if name_match:
            return name_match.group(1).strip()
        
        name_match = re.search(r'ФАМИЛИЯ:\s*([A-Z]+/[A-Z]+)', text)
        if name_match:
            return name_match.group(1).replace('/', ' ')
        
        return 'Не указано'
    
    def _extract_ticket_number(self, text: str) -> str:
        ticket_match = re.search(r'НОМЕР БИЛЕТА\s*:\s*(\d+\s+\d+)', text)
        if ticket_match:
            return ticket_match.group(1).replace(' ', '')
        return 'Не указано'
    
    def _extract_order_number(self, text: str, filename: str) -> str:
        order_match = re.search(r'-\s*([A-Z0-9]+)', filename)
        if order_match:
            return order_match.group(1)
        return 'Не указано'
    
    def _extract_issue_date(self, text: str) -> str:
        date_issue = re.search(r'ДАТА:\s*(\d{2})([А-ЯA-Z]{3})(\d{2})', text)
        if date_issue:
            day, month_str, year = date_issue.groups()
            month_num = self.MONTH_MAP.get(month_str[:3].upper(), '01')
            return f"{day}.{month_num}.20{year}"
        return 'Не указано'
    
    def _extract_carrier(self, text: str) -> str:
        text_upper = text.upper()
        lines = text.split('\n')
        
        carrier_patterns = [
            r'ВЫДАН\s+ОТ\s*[:;]?\s*([A-ZА-Я][A-ZА-Я\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'ISSUED\s+BY\s*[:;]?\s*([A-Z][A-Z\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'CARRIER\s*[:;]?\s*([A-Z][A-Z\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'АВИАКОМПАНИЯ\s*[:;]?\s*([A-ZА-Я][A-ZА-Я\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'OPERATED\s+BY\s*[:;]?\s*([A-Z][A-Z\s&\.\-]+?)(?:\s{2,}|\n|$)',
            r'VALIDATING\s+CARRIER\s*[:;]?\s*([A-Z]{2})',
            r'MARKETING\s+CARRIER\s*[:;]?\s*([A-Z]{2})',
        ]
        
        for pattern in carrier_patterns:
            match = re.search(pattern, text_upper)
            if match:
                carrier = match.group(1).strip()
                carrier = re.sub(r'\s+', ' ', carrier)
                if len(carrier) > 2 and len(carrier) < 50:
                    return self._clean_carrier_name(carrier)
        
        flight_patterns = [
            r'\b([A-Z]{2})\s*[-]?\s*\d{3,4}\b',
            r'FLIGHT\s*[:;]?\s*([A-Z]{2})\s*\d{3,4}',
            r'РЕЙС\s*[:;]?\s*([A-Z]{2})\s*\d{3,4}',
        ]
        
        for pattern in flight_patterns:
            match = re.search(pattern, text_upper)
            if match:
                code = match.group(1)
                if code not in ['TK', 'BG', 'NP', 'PC', 'NDC', 'NUC', 'ROE', 'RUB', 'USD', 'EUR']:
                    full_name = self._find_carrier_name_by_code(code, text)
                    return full_name if full_name else code
        
        known_carriers = [
            'LUFTHANSA', 'AIR FRANCE', 'KLM', 'BRITISH AIRWAYS', 'VIRGIN ATLANTIC',
            'TURKISH AIRLINES', 'PEGASUS AIRLINES', 'SUNEXPRESS', 'CORENDON AIRLINES',
            'AEROFLOT', 'S7 AIRLINES', 'ROSSIYA', 'URAL AIRLINES', 'POBEDA', 'AZIMUT',
            'NORDWIND', 'SMARTAVIA', 'AZUR AIR', 'ROYAL FLIGHT', 'IKAR', 'RED WINGS',
            'BELAVIA', 'AIR MOLDOVA', 'LOT POLISH AIRLINES', 'CZECH AIRLINES',
            'AUSTRIAN AIRLINES', 'SWISS', 'SAS', 'NORWEGIAN', 'FINNAIR',
            'IBERIA', 'TAP PORTUGAL', 'ALITALIA', 'AEGEAN AIRLINES', 'EL AL',
            'EMIRATES', 'ETIHAD AIRWAYS', 'QATAR AIRWAYS', 'FLY DUBAI', 'AIR ARABIA',
            'SAUDI ARABIAN AIRLINES', 'KUWAIT AIRWAYS', 'GULF AIR', 'OMAN AIR',
            'SINGAPORE AIRLINES', 'CATHAY PACIFIC', 'THAI AIRWAYS', 'MALAYSIA AIRLINES',
            'GARUDA INDONESIA', 'PHILIPPINE AIRLINES', 'VIETNAM AIRLINES',
            'JAPAN AIRLINES', 'ALL NIPPON AIRWAYS', 'KOREAN AIR', 'ASIANA AIRLINES',
            'CHINA SOUTHERN', 'CHINA EASTERN', 'AIR CHINA', 'HAINAN AIRLINES',
            'INDIGO', 'SPICEJET', 'AIR INDIA', 'GO FIRST',
            'UZBEKISTAN AIRWAYS', 'AIR ASTANA', 'SCAT AIRLINES', 'QANOT SHARQ',
            'KYRGYZ AIRLINES', 'TAJIK AIR', 'TURKMENISTAN AIRLINES',
            'AMERICAN AIRLINES', 'DELTA AIR LINES', 'UNITED AIRLINES', 'SOUTHWEST',
            'AIR CANADA', 'WESTJET', 'JETBLUE', 'ALASKA AIRLINES', 'SPIRIT AIRLINES',
            'LATAM', 'GOL', 'AZUL', 'AEROMEXICO', 'COPA AIRLINES', 'AVIANCA',
            'EGYPTAIR', 'ROYAL AIR MAROC', 'ETHIOPIAN AIRLINES', 'SOUTH AFRICAN AIRWAYS',
            'KENYA AIRWAYS', 'QANTAS', 'VIRGIN AUSTRALIA', 'AIR NEW ZEALAND',
            'RYANAIR', 'EASYJET', 'WIZZ AIR', 'VUELING', 'TRANSAVIA', 'EUROWINGS',
            'JET2', 'TUI AIRWAYS', 'NORWEGIAN AIR SHUTTLE',
            'AIR BALTIC', 'UKRAINE INTERNATIONAL AIRLINES', 'GEORGIAN AIRWAYS',
            'ARMENIA AIRWAYS', 'BUTA AIRWAYS', 'IRAN AIR', 'MAHAN AIR',
        ]
        
        known_carriers.sort(key=len, reverse=True)
        
        for carrier in known_carriers:
            if carrier.upper() in text_upper:
                return carrier
        
        generic_match = re.search(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\s+(?:AIRWAYS|AIRLINES|AIR|AVIA))\b', text)
        if generic_match:
            return generic_match.group(1)
        
        return 'Не указано'

    def _clean_carrier_name(self, name: str) -> str:
        remove_words = ['TICKET', 'E-TICKET', 'ETICKET', 'BOOKING', 'PNR', 'FARE', 'TAX', 'TOTAL']
        name_upper = name.upper()
        for word in remove_words:
            name_upper = name_upper.replace(word, '')
        
        name = name_upper.strip()
        name = re.sub(r'\s+', ' ', name)
        name = re.sub(r'\s+[A-Z0-9]$', '', name)
        
        return name if len(name) > 1 else 'Не указано'

    def _find_carrier_name_by_code(self, code: str, text: str) -> Optional[str]:
        iata_to_name = {
            'HY': 'UZBEKISTAN AIRWAYS', 'KC': 'AIR ASTANA', 'DV': 'SCAT AIRLINES',
            'SU': 'AEROFLOT', 'S7': 'S7 AIRLINES', 'U6': 'URAL AIRLINES',
            'DP': 'POBEDA', 'A4': 'AZIMUT', 'B2': 'BELAVIA', 'TK': 'TURKISH AIRLINES',
            'PC': 'PEGASUS AIRLINES', 'XQ': 'SUNEXPRESS', 'LH': 'LUFTHANSA',
            'AF': 'AIR FRANCE', 'BA': 'BRITISH AIRWAYS', 'EK': 'EMIRATES',
            'EY': 'ETIHAD AIRWAYS', 'QR': 'QATAR AIRWAYS', 'FZ': 'FLY DUBAI',
            'AA': 'AMERICAN AIRLINES', 'DL': 'DELTA AIR LINES', 'UA': 'UNITED AIRLINES',
            'FR': 'RYANAIR', 'U2': 'EASYJET', 'W6': 'WIZZ AIR', 'LX': 'SWISS',
            'OS': 'AUSTRIAN AIRLINES', 'KL': 'KLM', 'VS': 'VIRGIN ATLANTIC',
            'SQ': 'SINGAPORE AIRLINES', 'CX': 'CATHAY PACIFIC', 'JL': 'JAPAN AIRLINES',
            'NH': 'ALL NIPPON AIRWAYS', 'QF': 'QANTAS', 'NZ': 'AIR NEW ZEALAND',
            'SA': 'SOUTH AFRICAN AIRWAYS', 'MS': 'EGYPTAIR', 'ET': 'ETHIOPIAN AIRLINES',
            'AI': 'AIR INDIA', '6E': 'INDIGO', 'SG': 'SPICEJET', 'CZ': 'CHINA SOUTHERN',
            'MU': 'CHINA EASTERN', 'CA': 'AIR CHINA', 'HU': 'HAINAN AIRLINES',
            'TG': 'THAI AIRWAYS', 'MH': 'MALAYSIA AIRLINES', 'GA': 'GARUDA INDONESIA',
            'PR': 'PHILIPPINE AIRLINES', 'VN': 'VIETNAM AIRLINES', 'KE': 'KOREAN AIR',
            'OZ': 'ASIANA AIRLINES', 'AC': 'AIR CANADA', 'WS': 'WESTJET',
            'B6': 'JETBLUE', 'AS': 'ALASKA AIRLINES', 'LA': 'LATAM', 'G3': 'GOL',
            'AD': 'AZUL', 'AM': 'AEROMEXICO', 'CM': 'COPA AIRLINES', 'AV': 'AVIANCA',
            'ME': 'MIDDLE EAST AIRLINES', 'RJ': 'ROYAL JORDANIAN', 'KU': 'KUWAIT AIRWAYS',
            'GF': 'GULF AIR', 'WY': 'OMAN AIR', 'SV': 'SAUDI ARABIAN AIRLINES',
            'AT': 'ROYAL AIR MAROC', 'KM': 'AIR MALTA', 'CY': 'CYPRUS AIRWAYS',
            'A3': 'AEGEAN AIRLINES', 'RO': 'TAROM', 'LO': 'LOT POLISH AIRLINES',
            'OK': 'CZECH AIRLINES', 'JU': 'AIR SERBIA', 'FB': 'BULGARIA AIR',
            'OU': 'CROATIA AIRLINES', 'BT': 'AIR BALTIC', 'PS': 'UKRAINE INTERNATIONAL AIRLINES',
            'A9': 'GEORGIAN AIRWAYS', 'SZ': 'SOMON AIR', '5W': 'WIZZ AIR ABU DHABI',
            '9U': 'AIR MOLDOVA', '7B': 'BEES AIRLINE', 'IU': 'SUPER AIR JET',
        }
        
        if code in iata_to_name:
            return iata_to_name[code]
        
        pattern = rf'\b{code}\b\s*[-:]\s*["\']?([A-Z][A-Za-z\s&]+?)["\']?(?:\n|$)'
        match = re.search(pattern, text)
        if match:
            name = match.group(1).strip()
            if len(name) > 2 and len(name) < 40:
                return name
        
        return None
    
    def _extract_flight_number(self, text: str) -> str:
        flight_match = re.search(r'\b([A-Z]{2})\s*(\d{3,4})\b', text)
        if flight_match:
            return f"{flight_match.group(1)}{flight_match.group(2)}"
        return 'Не указано'
    
    def _extract_flight_times(self, text: str, issue_date: str) -> Dict[str, str]:
        result = {
            'dep_time': 'Не указано',
            'dep_date': 'Не указано',
            'arr_time': 'Не указано',
            'arr_date': 'Не указано'
        }
        
        time_match = re.search(r'(\d{2})([А-ЯA-Z]{3})\s+(\d{2})(\d{2})\s+(\d{2})(\d{2})', text)
        
        if time_match:
            day, month_str, dep_h, dep_m, arr_h, arr_m = time_match.groups()
            month_num = self.MONTH_MAP.get(month_str[:3].upper(), '01')
            
            year = '2026'
            if issue_date != 'Не указано':
                year = issue_date.split('.')[-1]
            
            result['dep_time'] = f"{dep_h}:{dep_m}"
            result['arr_time'] = f"{arr_h}:{arr_m}"
            result['dep_date'] = f"{day}.{month_num}.{year}"
            
            try:
                dep_dt = datetime(int(year), int(month_num), int(day), int(dep_h), int(dep_m))
                arr_dt = datetime(int(year), int(month_num), int(day), int(arr_h), int(arr_m))
                
                if arr_dt < dep_dt:
                    arr_dt += timedelta(days=1)
                
                result['arr_date'] = arr_dt.strftime('%d.%m.%Y')
            except Exception as e:
                logger.warning(f"Ошибка расчета даты прилета: {e}")
                result['arr_date'] = result['dep_date']
        
        return result
    
    def _extract_route(self, text: str) -> Dict[str, str]:
        result = {'route': 'Не указано', 'arrival': 'Не указано'}
        
        codes = re.findall(r'\b([A-Z]{3})\b', text)
        exclude = {'NUC', 'ROE', 'RUB', 'EUR', 'PC', 'NDC', 'USD', 'END', 'MOW'}
        valid = [c for c in codes if c not in exclude]
        
        dep_code = arr_code = None
        for c in valid:
            if c == 'TMJ':
                dep_code = c
            elif c in ['VKO', 'DME', 'SVO', 'ZIA']:
                arr_code = c
        
        if dep_code and arr_code:
            dep_city = self.IATA_CODES.get(dep_code, dep_code)
            arr_city = self.IATA_CODES.get(arr_code, arr_code)
            result['route'] = f"{dep_city} - {arr_city}"
            result['arrival'] = f"{arr_city} ({arr_code})"
        
        return result
    
    def _extract_price(self, text: str) -> Dict[str, str]:
        result = {'price': 'Не указано', 'currency': 'RUB'}
        
        total_match = re.search(r'ИТОГО\s*:\s*(\d+)\s*РУБ', text, re.I)
        if total_match:
            result['price'] = total_match.group(1)
        else:
            total_match = re.search(r'ИТОГО К ОПЛАТЕ\s*:\s*RUB(\d+)', text)
            if total_match:
                result['price'] = total_match.group(1)
        
        return result


class ProcessingThread(QThread):
    """Поток для обработки файлов без блокировки UI"""
    progress = pyqtSignal(int)
    finished_signal = pyqtSignal(list)
    error_signal = pyqtSignal(str)
    
    def __init__(self, files: List[str], parser: TicketParser):
        super().__init__()
        self.files = files
        self.parser = parser
        self._is_running = True
        
    def run(self):
        results = []
        for i, file_path in enumerate(self.files):
            if not self._is_running:
                break
            
            try:
                data = self.parser.parse_file(file_path)
                results.append(data)
                self.progress.emit(i + 1)
            except Exception as e:
                logger.error(f"Ошибка обработки {file_path}: {e}")
                self.error_signal.emit(f"Ошибка в файле {Path(file_path).name}: {str(e)}")
        
        self.finished_signal.emit(results)
    
    def stop(self):
        self._is_running = False


class TicketDataTab(BaseTabWidget):
    """Вкладка 3: Данные из Билетов"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.tickets: List[TicketData] = []
        self.parser = TicketParser()
        self.processing_thread: Optional[ProcessingThread] = None
        
        # Enable Drag and Drop
        self.setAcceptDrops(True)

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        header = QLabel("🎫 Данные из Билетов")
        header.setProperty("class", "title")
        layout.addWidget(header)

        # Controls
        controls = QHBoxLayout()
        self.folder_input = QLineEdit()
        self.folder_input.setPlaceholderText("Выберите папку с PDF билетами...")
        controls.addWidget(self.folder_input)
        
        browse_btn = QPushButton("📂 Обзор")
        browse_btn.clicked.connect(self.select_folder)
        controls.addWidget(browse_btn)
        
        scan_btn = QPushButton("🔍 Сканировать")
        scan_btn.setProperty("class", "primary")
        scan_btn.clicked.connect(self.scan_folder)
        controls.addWidget(scan_btn)
        
        layout.addLayout(controls)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(13)
        self.table.setHorizontalHeaderLabels([
            'Пассажир', 'Номер билета', 'Номер заказа', 'Дата выдачи',
            'Перевозчик', 'Номер рейса', 'Время вылета', 'Дата вылета',
            'Маршрут', 'Аэропорт прибытия', 'Время прилета', 'Дата прилета', 'Стоимость'
        ])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_table_context_menu)
        layout.addWidget(self.table)
        
        # Export
        export_layout = QHBoxLayout()
        self.stats_label = QLabel("Билетов: 0")
        self.stats_label.setProperty("class", "caption")
        export_layout.addWidget(self.stats_label)
        
        export_layout.addStretch()
        
        export_btn = QPushButton("📊 Экспорт в Excel")
        export_btn.clicked.connect(self.export_to_excel)
        export_btn.setEnabled(False)
        self.export_btn = export_btn
        export_layout.addWidget(export_btn)
        
        layout.addLayout(export_layout)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        if urls:
            folder_path = urls[0].toLocalFile()
            if os.path.isdir(folder_path):
                self.folder_input.setText(folder_path)
                self.scan_folder()
            elif os.path.isfile(folder_path):
                # Если перетащили один файл, обрабатываем его
                self.process_files([folder_path])

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку с билетами", self.parser.settings.value('last_folder', ''))
        if folder:
            self.folder_input.setText(folder)
            self.scan_folder()

    def scan_folder(self):
        folder = self.folder_input.text()
        if not folder or not Path(folder).exists():
            QMessageBox.warning(self, "Ошибка", "Выберите корректную папку!")
            return
        
        path = Path(folder)
        files = []
        
        if fitz:
            files.extend(path.glob('*.pdf'))
            files.extend(path.glob('*.PDF'))
        
        for ext in ['*.jpg', '*.jpeg', '*.png', '*.bmp', '*.tiff']:
            files.extend(path.glob(ext))
            files.extend(path.glob(ext.upper()))
        
        files = sorted(list(set(files)))
        
        if not files:
            QMessageBox.information(self, "Нет файлов", "В папке не найдено PDF файлов или изображений.")
            return
        
        reply = QMessageBox.question(
            self, 'Найдены файлы',
            f'Найдено файлов: {len(files)}\n\nНачать обработку?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.process_files([str(f) for f in files])

    def process_files(self, files: List[str]):
        self.tickets = []
        self.update_table()
        
        self.processing_thread = ProcessingThread(files, self.parser)
        self.processing_thread.progress.connect(self.on_progress)
        self.processing_thread.finished_signal.connect(self.on_processing_finished)
        self.processing_thread.error_signal.connect(self.on_processing_error)
        self.processing_thread.start()
        
        self.statusbar_message("Обработка...")

    def on_progress(self, value: int):
        # Здесь можно добавить прогресс бар, если нужно
        pass

    def on_processing_finished(self, results: List[TicketData]):
        self.tickets = results
        self.update_table()
        self.update_stats()
        
        success_count = len([t for t in results if t.ticket_number != 'Не указано'])
        self.statusbar_message(f"Обработано: {len(results)} файлов, распознано: {success_count}")
        
        if success_count > 0:
            self.export_btn.setEnabled(True)
        
        QMessageBox.information(
            self, 'Готово',
            f'Обработано файлов: {len(results)}\nУспешно распознано: {success_count}'
        )

    def on_processing_error(self, error_msg: str):
        logger.error(error_msg)
        self.statusbar_message(f"Ошибка: {error_msg}", 5000)

    def update_table(self):
        self.table.setRowCount(len(self.tickets))
        
        for i, ticket in enumerate(self.tickets):
            data = ticket.to_dict()
            values = [
                data['Пассажир'],
                data['Номер билета'],
                data['Номер заказа'],
                data['Дата выдачи'],
                data['Перевозчик'],
                data['Номер рейса'],
                data['Время отправления'],
                data['Дата отправления'],
                data['Маршрут'],
                data['Аэропорт прибытия'],
                data['Время прибытия'],
                data['Дата прибытия'],
                f"{data['Стоимость']} {data['Валюта']}" if data['Стоимость'] != 'Не указано' else 'Не указано'
            ]
            
            for j, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if value == 'Не указано' or value == '':
                    item.setBackground(QColor("#fff3cd")) # Желтый фон для ошибок
                self.table.setItem(i, j, item)
        
        self.table.resizeColumnsToContents()

    def update_stats(self):
        total = len(self.tickets)
        with_tickets = len([t for t in self.tickets if t.ticket_number != 'Не указано'])
        with_flights = len([t for t in self.tickets if t.flight_number != 'Не указано'])
        with_times = len([t for t in self.tickets if t.departure_time != 'Не указано'])
        
        stats_text = (
            f'<b>Всего билетов:</b> {total}<br>'
            f'<b>С номером билета:</b> {with_tickets}<br>'
            f'<b>С номером рейса:</b> {with_flights}<br>'
            f'<b>С временем вылета:</b> {with_times}'
        ) if total > 0 else 'Билетов: 0'
        
        self.stats_label.setText(stats_text)

    def export_to_excel(self):
        if not self.tickets:
            QMessageBox.warning(self, 'Нет данных', 'Нет данных для экспорта!')
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            'Сохранить как',
            f'tickets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            'Excel Files (*.xlsx)'
        )
        
        if not file_path:
            return
        
        try:
            data = [t.to_dict() for t in self.tickets]
            df = pd.DataFrame(data)
            
            column_order = [
                'Пассажир', 'Номер билета', 'Номер заказа', 'Дата выдачи',
                'Перевозчик', 'Номер рейса', 'Маршрут', 'Аэропорт прибытия',
                'Дата отправления', 'Время отправления',
                'Дата прибытия', 'Время прибытия',
                'Стоимость', 'Валюта', 'Источник'
            ]
            df = df[column_order]
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Билеты', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Билеты']
                
                # Автоматическая ширина колонок
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Стилизация заголовков
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            self.statusbar_message(f'Сохранено: {file_path}')
            QMessageBox.information(self, 'Успех', f'Данные сохранены в:\n{file_path}')
            
        except Exception as e:
            logger.error(f"Ошибка экспорта в Excel: {e}")
            QMessageBox.critical(self, 'Ошибка', f'Не удалось сохранить:\n{str(e)}')

    def show_table_context_menu(self, position):
        menu = QMenu()
        
        action_copy = QAction('Копировать', self)
        action_copy.triggered.connect(self.copy_selected_cells)
        menu.addAction(action_copy)
        
        action_copy_row = QAction('Копировать строку', self)
        action_copy_row.triggered.connect(self.copy_selected_row)
        menu.addAction(action_copy_row)
        
        menu.exec_(self.table.viewport().mapToGlobal(position))

    def copy_selected_cells(self):
        selected = self.table.selectedItems()
        if selected:
            text = '\t'.join([item.text() for item in selected])
            QApplication.clipboard().setText(text)

    def copy_selected_row(self):
        row = self.table.currentRow()
        if row >= 0:
            values = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                values.append(item.text() if item else '')
            text = '\t'.join(values)
            QApplication.clipboard().setText(text)
    
    def statusbar_message(self, msg: str, timeout=0):
        main_win = self.window()
        if isinstance(main_win, MainWindow):
            if timeout > 0:
                main_win.statusbar.showMessage(msg, timeout)
            else:
                main_win.statusbar.showMessage(msg)


# --- TAB 4 & 5: PLACEHOLDERS ---

class PassportDataTab(BaseTabWidget):
    def setup_ui(self):
        layout = QVBoxLayout(self)
        header = QLabel("🛂 Данные из Паспортов")
        header.setProperty("class", "title")
        layout.addWidget(header)
        layout.addWidget(QLabel("В разработке..."))
        layout.addStretch()


class ReportTab(BaseTabWidget):
    def setup_ui(self):
        layout = QVBoxLayout(self)
        header = QLabel("📊 Отчет по Уволенному персоналу")
        header.setProperty("class", "title")
        layout.addWidget(header)
        layout.addWidget(QLabel("Функционал будет добавлен сегодня."))
        layout.addStretch()


# --- MAIN WINDOW ---

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("FolderForge Pro - Ultimate Toolset")
        self.setMinimumSize(1200, 800)
        self.resize(1400, 900)

        self.scale_factor = 1.0
        self.apply_theme()
        self.setup_ui()
        self.setup_menu()
        self.setup_statusbar()

    def _detect_scale_factor(self) -> float:
        screen = QApplication.primaryScreen()
        if screen:
            dpi = screen.logicalDotsPerInch()
            return max(1.0, min(2.0, dpi / 96.0))
        return 1.0

    def apply_theme(self):
        base_font_size = int(10 * self.scale_factor)
        stylesheet = ModernStyle.get_stylesheet(base_font_size)
        self.setStyleSheet(stylesheet)
        font = QFont("Segoe UI", base_font_size)
        font.setStyleHint(QFont.StyleHint.SansSerif)
        QApplication.setFont(font)

    def setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)

        self.tabs = QTabWidget()
        
        self.tab_folders = FolderCreationTab()
        self.tab_carnet = MergeCarnetTab()
        self.tab_tickets = TicketDataTab()
        self.tab_passports = PassportDataTab()
        self.tab_reports = ReportTab()

        self.tabs.addTab(self.tab_folders, "📂 Создание папок")
        self.tabs.addTab(self.tab_carnet, "🔗 Объединение Карнетов")
        self.tabs.addTab(self.tab_tickets, "🎫 Данные из Билетов")
        self.tabs.addTab(self.tab_passports, "🛂 Данные из Паспортов")
        self.tabs.addTab(self.tab_reports, "📊 Отчет по Уволенному персоналу")

        main_layout.addWidget(self.tabs)

    def setup_menu(self):
        menubar = self.menuBar()
        
        file_menu = menubar.addMenu("📁 Файл")
        exit_action = QAction("Выход", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        nav_menu = menubar.addMenu("🧭 Навигация")
        nav_menu.addAction("Создание папок", lambda: self.tabs.setCurrentIndex(0))
        nav_menu.addAction("Объединение Карнетов", lambda: self.tabs.setCurrentIndex(1))
        nav_menu.addAction("Данные из Билетов", lambda: self.tabs.setCurrentIndex(2))
        nav_menu.addAction("Данные из Паспортов", lambda: self.tabs.setCurrentIndex(3))
        nav_menu.addAction("Отчет по Уволенному персоналу", lambda: self.tabs.setCurrentIndex(4))

        help_menu = menubar.addMenu("❓ Помощь")
        help_menu.addAction("О программе", self.show_about)

    def setup_statusbar(self):
        self.statusbar = QStatusBar()
        self.setStatusBar(self.statusbar)
        self.statusbar.showMessage("Готов")

    def show_about(self):
        QMessageBox.about(self, "О программе", "<h2>FolderForge Pro</h2><p>Версия 3.0</p>")

    def closeEvent(self, event):
        if hasattr(self.tab_folders, 'worker') and self.tab_folders.worker.isRunning():
            self.tab_folders.worker.stop()
            self.tab_folders.worker.wait(2000)
        if hasattr(self.tab_carnet, 'merge_worker') and self.tab_carnet.merge_worker.isRunning():
            self.tab_carnet.merge_worker.stop()
            self.tab_carnet.merge_worker.wait()
        if hasattr(self.tab_tickets, 'processing_thread') and self.tab_tickets.processing_thread.isRunning():
            self.tab_tickets.processing_thread.stop()
            self.tab_tickets.processing_thread.wait()
        event.accept()


def main():
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.ApplicationAttribute.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.ApplicationAttribute.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    app.setApplicationName("FolderForge Pro - Ultimate")
    
    font = QFont("Segoe UI", 10)
    app.setFont(font)

    window = MainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
