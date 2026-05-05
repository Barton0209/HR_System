# passport_ocr.py
"""
Точка входа для обработки паспортов.
Сбор файлов → ocr_pipeline (Passport режим) → Excel.

Запуск: python passport_ocr.py
"""

import sys
import logging
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ocr_pipeline.runner import run_batch, Mode

sys.stdout.reconfigure(encoding='utf-8')
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

# ── Пути ──────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
PASS_DIR    = BASE_DIR / "Download" / "Pass"
FSB_DIR     = PASS_DIR / "ФСБ"
FOREIGN_DIR = PASS_DIR / "Иностранные_паспорта"
OCR_DIR     = PASS_DIR / "OCR"
EXCEL_OUT   = BASE_DIR / "Upload" / "Data_Pass_OCR.xlsx"
SHEET_NAME  = "Данные_OCR"

# ── Сбор файлов ───────────────────────────────────────────────────────────────

def collect_files() -> list:
    files = []

    # ФСБ: папки с именами → все OCR-страницы или оригинал
    if FSB_DIR.exists():
        for person_dir in sorted(FSB_DIR.iterdir()):
            if not person_dir.is_dir():
                continue
            fio = person_dir.name
            # Берём все OCR файлы этого человека (все страницы паспорта)
            ocr_files = sorted(OCR_DIR.glob(f"{fio}_*-ocr.pdf")) if OCR_DIR.exists() else []
            if ocr_files:
                # Первый файл — главная страница с фото и данными
                files.append({"path": ocr_files[0], "fio": fio,
                               "source": "ФСБ", "lang": "rus"})
            else:
                # OCR ещё не создан — берём оригинальный PDF
                pdfs = sorted(person_dir.glob("*.pdf"))
                if pdfs:
                    files.append({"path": pdfs[0], "fio": fio,
                                   "source": "ФСБ", "lang": "rus"})

    # Иностранные: отдельные файлы
    if FOREIGN_DIR.exists():
        for f in sorted(FOREIGN_DIR.iterdir()):
            if f.suffix.lower() in (".pdf", ".jpg", ".jpeg", ".png"):
                files.append({"path": f, "fio": f.stem.strip(),
                               "source": "Иностранные", "lang": "rus+eng"})

    return files


# ── Запись в Excel ────────────────────────────────────────────────────────────

HEADERS = {
    "A": "П/П",
    "B": "Фамилия, имя, отчество",
    "C": "Фамилия, имя, отчество (латиница)",
    "D": "Дата рождения",
    "E": "Гражданство",
    "F": "Серия и номер паспорта",
    "G": "Кем выдан",
    "H": "Дата выдачи",
    "I": "Дата окончания",
    "J": "Код подразделения",
    "K": "Прописка, Регистрация",
    "L": "Методы распознавания",
}

COL_WIDTHS = {
    "A": 6,  "B": 35, "C": 35, "D": 14, "E": 16,
    "F": 18, "G": 45, "H": 14, "I": 14, "J": 14,
    "K": 50, "L": 40,
}


def write_excel(records: list):
    EXCEL_OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(EXCEL_OUT) if EXCEL_OUT.exists() else Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    hdr_fill  = PatternFill("solid", fgColor="217346")
    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font  = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="center", wrap_text=False)
    thin   = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_letter, header in HEADERS.items():
        cell = ws[f"{col_letter}1"]
        cell.value, cell.font, cell.fill = header, hdr_font, hdr_fill
        cell.alignment, cell.border = hdr_align, border
    ws.row_dimensions[1].height = 30

    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    even_fill = PatternFill("solid", fgColor="EEF4EE")

    for i, rec in enumerate(records, 1):
        row  = i + 1
        fill = odd_fill if i % 2 == 1 else even_fill
        methods = ", ".join(rec.get("_methods", []))
        values = [
            i,
            rec.get("fio", ""),
            rec.get("fio_lat", ""),
            rec.get("birth_date", ""),
            rec.get("citizenship", ""),
            rec.get("doc_series_num", ""),
            rec.get("issuer", ""),
            rec.get("issue_date", ""),
            rec.get("expiry_date", ""),
            rec.get("dept_code", ""),
            rec.get("address", ""),
            methods,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font, cell.alignment = cell_font, cell_align
            cell.fill, cell.border = fill, border
        ws.row_dimensions[row].height = 18

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(records) + 1}"
    wb.save(EXCEL_OUT)
    logger.info("Сохранено: %s  (%d записей)", EXCEL_OUT, len(records))


# ── main ──────────────────────────────────────────────────────────────────────

BATCH_SIZE = 20  # обрабатываем порциями, чтобы не переполнять память


def main():
    import gc
    logger.info("Сбор файлов...")
    files = collect_files()
    logger.info("Найдено файлов: %d", len(files))

    all_records = []
    for start in range(0, len(files), BATCH_SIZE):
        chunk = files[start:start + BATCH_SIZE]
        logger.info("Порция %d-%d...", start + 1, start + len(chunk))
        records = run_batch(chunk, Mode.PASSPORT)
        all_records.extend(records)
        gc.collect()  # явная очистка памяти после каждой порции

    if all_records:
        write_excel(all_records)
        print(f"\n✅ Готово! Записей: {len(all_records)}")
        print(f"   Файл: {EXCEL_OUT}")
    else:
        print("\n⚠ Нет данных для записи")


if __name__ == "__main__":
    main()
