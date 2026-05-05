"""
Система обработки заявок на билеты v3.0
+ Паспорта 23 стран
+ Универсальный OCR

Запуск: python -m ticket_app.main
"""

import sys
import os

# Добавляем корень проекта в путь
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

import logging
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.font import Font

import pandas as pd

from ticket_app.auth import AuthManager, LoginWindow
from ticket_app.config import ALL_COLUMNS, ROUTES, REASONS
from ticket_app.database import (
    load_employees_base, get_employees_db, set_employees_db,
    find_employee_by_fio, get_all_employees,
)
from ticket_app.excel_handler import save_as_excel, create_application_row, create_empty_row
from ticket_app.pdf_processor import process_pdf_file
from ticket_app.storage import save_database_to_file, load_database_from_file
from ticket_app.dialogs.pdf_viewer import PDFViewer
from ticket_app.dialogs.catalog import EmployeeCatalogDialog
from ticket_app.dialogs.wizard import FillFromBaseWizard
from ticket_app.dialogs.ticket_parser_tab import TicketParserTab

# Новые модули OCR
try:
    from tabs import PassportTab, UniversalOCRTab
    OCR_TABS_AVAILABLE = True
except ImportError as e:
    OCR_TABS_AVAILABLE = False
    print(f"OCR tabs not available: {e}")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('ticket_app.log', encoding='utf-8'),
    ]
)
logger = logging.getLogger(__name__)


class ExcelTable(tk.Frame):
    """Treeview с поведением Excel"""

    CLR_HEADER_BG  = '#217346'
    CLR_HEADER_FG  = '#ffffff'
    CLR_ROW_ODD    = '#ffffff'
    CLR_ROW_EVEN   = '#EEF4EE'
    CLR_ROW_SEL    = '#C6EFCE'
    CLR_ROW_SEL_FG = '#000000'
    CLR_NOT_FOUND  = '#FFE0E0'
    FONT_HEADER    = ('Calibri', 10, 'bold')
    FONT_CELL      = ('Calibri', 10)
    ROW_HEIGHT     = 22

    def __init__(self, parent, columns: list, col_widths: dict,
                 on_cell_edited=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.columns = columns
        self.col_widths = col_widths
        self.on_cell_edited = on_cell_edited
        self._sort_col = None
        self._sort_asc = True
        self._edit_widget = None
        self._df = None

        self._build()

    def _build(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Excel.Treeview',
                         font=self.FONT_CELL,
                         rowheight=self.ROW_HEIGHT,
                         fieldbackground=self.CLR_ROW_ODD,
                         background=self.CLR_ROW_ODD)
        style.configure('Excel.Treeview.Heading',
                         font=self.FONT_HEADER,
                         background=self.CLR_HEADER_BG,
                         foreground=self.CLR_HEADER_FG)
        style.map('Excel.Treeview',
                  background=[('selected', self.CLR_ROW_SEL)])

        vsb = ttk.Scrollbar(self, orient='vertical')
        hsb = ttk.Scrollbar(self, orient='horizontal')
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')

        self.tree = ttk.Treeview(
            self,
            columns=self.columns,
            show='headings',
            style='Excel.Treeview',
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
        )
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        self.tree.pack(fill='both', expand=True)

        self.tree.tag_configure('odd', background=self.CLR_ROW_ODD)
        self.tree.tag_configure('even', background=self.CLR_ROW_EVEN)
        self.tree.tag_configure('not_found', background=self.CLR_NOT_FOUND)

        for col in self.columns:
            w = self.col_widths.get(col, 80)
            self.tree.heading(col, text=col, command=lambda c=col: self._sort_by(c))
            self.tree.column(col, width=w, minwidth=30, anchor='w')

        self._footer_frame = tk.Frame(self, bg='#D8E4BC', height=self.ROW_HEIGHT)
        self._footer_frame.pack(fill='x', side='bottom')
        self._footer_label = tk.Label(
            self._footer_frame, text='', font=('Calibri', 9, 'italic'),
            bg='#D8E4BC', fg='#333', anchor='w')
        self._footer_label.pack(fill='x', padx=6)

        self.tree.bind('<Double-1>', self._on_double_click)
        self.tree.bind('<F2>', self._start_edit_selected)
        self.tree.bind('<Delete>', self._delete_selected)

    def load_dataframe(self, df):
        self._df = df.copy() if df is not None else pd.DataFrame(columns=self.columns)
        self._refresh()

    def _refresh(self):
        self.tree.delete(*self.tree.get_children())
        if self._df is None or self._df.empty:
            self._footer_label.config(text='Нет данных')
            return
        for i, (_, row) in enumerate(self._df.iterrows()):
            vals = [str(v) if pd.notna(v) else '' for v in row]
            note = str(row.get('Примечание', ''))
            tag = 'not_found' if 'НЕ НАЙДЕН' in note.upper() else ('odd' if i % 2 == 0 else 'even')
            self.tree.insert('', 'end', iid=str(i), values=vals, tags=(tag,))
        self._update_footer()

    def _update_footer(self):
        if self._df is None or self._df.empty:
            self._footer_label.config(text='Нет данных')
            return
        total = len(self._df)
        not_found = self._df['Примечание'].astype(str).str.contains('НЕ НАЙДЕН', case=False, na=False).sum()
        self._footer_label.config(text=f'  Всего: {total} | Не найдено: {not_found}')

    def _sort_by(self, col):
        if self._df is None or self._df.empty:
            return
        self._sort_asc = not self._sort_asc if self._sort_col == col else True
        self._sort_col = col
        try:
            self._df = self._df.sort_values(by=col, ascending=self._sort_asc,
                                            key=lambda s: s.astype(str).str.lower()).reset_index(drop=True)
        except Exception:
            pass
        self._refresh()

    def _on_double_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != 'cell':
            return
        iid = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not iid or not col_id:
            return
        col_idx = int(col_id.replace('#', '')) - 1
        self._start_edit(iid, col_idx)

    def _start_edit_selected(self, event=None):
        sel = self.tree.selection()
        if sel:
            self._start_edit(sel[0], 0)

    def _start_edit(self, iid, col_idx):
        if self._edit_widget:
            self._edit_widget.destroy()
        col_name = self.columns[col_idx]
        bbox = self.tree.bbox(iid, column=f'#{col_idx + 1}')
        if not bbox:
            return
        x, y, w, h = bbox
        current_val = self.tree.item(iid, 'values')[col_idx]

        entry = tk.Entry(self.tree, font=self.FONT_CELL, relief='solid', bd=1)
        entry.insert(0, current_val)
        entry.select_range(0, 'end')
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()
        entry.bind('<Return>', lambda e: self._commit_edit(entry, iid, col_idx, col_name))
        entry.bind('<Escape>', lambda e: entry.destroy())
        self._edit_widget = entry

    def _commit_edit(self, entry, iid, col_idx, col_name):
        new_val = entry.get()
        entry.destroy()
        self._edit_widget = None
        vals = list(self.tree.item(iid, 'values'))
        vals[col_idx] = new_val
        self.tree.item(iid, values=vals)
        if self._df is not None:
            row_idx = int(iid)
            if row_idx < len(self._df):
                self._df.at[self._df.index[row_idx], col_name] = new_val
                if self.on_cell_edited:
                    self.on_cell_edited(row_idx, col_name, new_val)

    def _delete_selected(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        if messagebox.askyesno('Удаление', f'Удалить {len(sel)} строк?'):
            indices = sorted([int(iid) for iid in sel], reverse=True)
            for idx in indices:
                if self._df is not None and idx < len(self._df):
                    self._df = self._df.drop(self._df.index[idx]).reset_index(drop=True)
            self._refresh()


class MainApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Система обработки заявок на билеты v3.0")
        self.root.geometry("1800x1000")
        self.root.state('zoomed')

        self.auth = AuthManager()
        self.current_department = None
        self.is_admin = False

        self.pdf_results = []
        self.pdf_files = {}
        self.applications_df = None
        self.existing_tab_nums = set()

        self._build_ui()
        self._auto_load_database()
        self._show_login()
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _auto_load_database(self):
        df = load_database_from_file()
        if df is not None and not df.empty:
            set_employees_db(df)
            self.status_label.config(text=f"Автозагрузка: {len(df)} сотрудников")
            self.btn_catalog.config(state='normal')

    def _on_closing(self):
        self._save_current_database()
        self.root.quit()
        self.root.destroy()

    def _save_current_database(self):
        db = get_employees_db()
        if db is not None and not db.empty:
            save_database_to_file(db)

    def _build_ui(self):
        # Шапка
        top = tk.Frame(self.root, bg='#1a1a2e', height=60)
        top.pack(fill='x')
        tk.Label(top, text="СИСТЕМА ЗАЯВОК НА БИЛЕТЫ v3.0",
                 font=('Arial', 18, 'bold'), bg='#1a1a2e', fg='white').pack(side='left', padx=20, pady=15)
        self.user_label = tk.Label(top, text="Не авторизован",
                                   font=('Arial', 11), bg='#1a1a2e', fg='#e94560')
        self.user_label.pack(side='right', padx=20)

        # Тулбар
        toolbar = tk.Frame(self.root, bg='#16213e', height=50)
        toolbar.pack(fill='x')
        toolbar.pack_propagate(False)

        btn_cfg = dict(fg='white', font=('Arial', 10), padx=15, pady=8)

        self.btn_load_base = tk.Button(
            toolbar, text="Загрузить Базу", command=self._load_base,
            bg='#0f3460', state='disabled', **btn_cfg)
        self.btn_load_base.pack(side='left', padx=5, pady=10)

        self.btn_catalog = tk.Button(
            toolbar, text="Каталог сотрудников", command=self._open_catalog,
            bg='#17a2b8', state='disabled', **btn_cfg)
        self.btn_catalog.pack(side='left', padx=5, pady=10)

        tk.Button(toolbar, text="Выбрать файл", command=self._select_single_file,
                  bg='#0f3460', **btn_cfg).pack(side='left', padx=5, pady=10)

        tk.Button(toolbar, text="Выбрать папку", command=self._select_folder,
                  bg='#0f3460', **btn_cfg).pack(side='left', padx=5, pady=10)
     

        def _download_eju_from_outlook(self):
            import simpledialog, messagebox
            email = simpledialog.askstring("Email", "Ваш Outlook email:")
            if not email: return
            import getpass
            password = getpass.getpass("Пароль (в терминале):")
            try:
                client = OutlookClient(email, password)
                # Спросить дату у пользователя
                date_str = simpledialog.askstring("Дата", "Дата в формате ДД.ММ.ГГГГ (например, 02.05.2026):")
                if not date_str: return
                files = client.download_eju_attachments(date_str)
                if files:
                    messagebox.showinfo("Успех", f"Скачано {len(files)} файлов ЕЖУ в {os.path.abspath('EJU/Download/' + date_str)}")
                    # ВАЖНО: Сразу запустить обработку (как в вашем описании)
                    self._process_eju_folder(os.path.abspath("EJU/Download/" + date_str))
                else:
                    messagebox.showwarning("Нет файлов", "Файлы за эту дату не найдены")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка IMAP: {e}")

        self.btn_fill = tk.Button(
            toolbar, text="Заполнить из Базы", command=self._fill_from_database,
            bg='#0f3460', state='disabled', **btn_cfg)
        self.btn_fill.pack(side='left', padx=5, pady=10)

        self.btn_export = tk.Button(
            toolbar, text="Экспорт в Excel", command=self._export_to_excel,
            bg='#e94560', state='disabled', **btn_cfg)
        self.btn_export.pack(side='left', padx=5, pady=10)

        self.btn_clear = tk.Button(
            toolbar, text="Очистить заявку", command=self._clear_all,
            bg='#dc3545', state='disabled', **btn_cfg)
        self.btn_clear.pack(side='left', padx=5, pady=10)

        tk.Button(toolbar, text="🔄 Сменить пользователя", command=self._switch_user,
                  bg='#6c757d', **btn_cfg).pack(side='right', padx=5, pady=10)

        tk.Button(toolbar, text="Выход", command=self._logout,
                  bg='#333', **btn_cfg).pack(side='right', padx=10, pady=10)

        # Вкладки
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True)

        # Вкладка 1: Заявки на билеты
        tab_orders = ttk.Frame(self.notebook)
        self.notebook.add(tab_orders, text="  📋 Заявки на билеты  ")
        self._build_tickets_tab(tab_orders)

        # Вкладка 2: Парсер авиабилетов
        tab_parser = ttk.Frame(self.notebook)
        self.notebook.add(tab_parser, text="  ✈ Парсер билетов  ")
        self.ticket_parser_tab = TicketParserTab(tab_parser)
        self.ticket_parser_tab.pack(fill='both', expand=True)

        # Вкладка 3: Паспорта
        if OCR_TABS_AVAILABLE:
            tab_passports = PassportTab(self.notebook)
            self.notebook.add(tab_passports, text="  🛂 Паспорта 23 стран  ")

            # Вкладка 4: Universal OCR
            tab_ocr = UniversalOCRTab(self.notebook)
            self.notebook.add(tab_ocr, text="  🔍 Universal OCR  ")
        else:
            logger.warning("OCR вкладки недоступны")

        # Статус бар
        self.status_label = tk.Label(
            self.root, text="Готов к работе", bd=1, relief='sunken',
            anchor='w', font=('Arial', 9), bg='#f0f0f0')
        self.status_label.pack(side='bottom', fill='x')

        # Прогресс бар
        self.progress_frame = tk.Frame(self.root)
        self.progress_label = tk.Label(self.progress_frame, text="", font=('Arial', 10))
        self.progress_label.pack(side='left', padx=10)
        self.progress_bar = ttk.Progressbar(self.progress_frame, mode='determinate', length=300)
        self.progress_bar.pack(side='left', padx=10)

    def _build_tickets_tab(self, parent):
        """Построение вкладки заявок на билеты."""
        main_paned = ttk.PanedWindow(parent, orient='horizontal')
        main_paned.pack(fill='both', expand=True)

        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=2)

        # Список PDF
        pdf_list_frame = ttk.LabelFrame(left_frame, text="Обработанные PDF", padding=5)
        pdf_list_frame.pack(fill='x', padx=5, pady=5)
        scrollbar = ttk.Scrollbar(pdf_list_frame)
        self.pdf_listbox = tk.Listbox(pdf_list_frame, height=4, font=('Arial', 10),
                                      yscrollcommand=scrollbar.set)
        self.pdf_listbox.pack(side='left', fill='x', expand=True)
        scrollbar.config(command=self.pdf_listbox.yview)
        scrollbar.pack(side='right', fill='y')
        self.pdf_listbox.bind('<<ListboxSelect>>', self._on_pdf_select)

        # Таблица
        col_widths = {
            "№": 40, "Подразделение": 110, "Отдел": 70, "Операция": 70,
            "Классификация": 85, "Дата заказа": 90, "Организация": 80,
            "Ф.И.О.": 160, "Ф.И.О лат": 160, "Табельный номер": 75,
            "Гражданство": 90, "Дата рождения": 90, "Вид документа": 140,
            "Серия": 55, "Номер": 75, "Дата выдачи": 90, "Дата окончания": 90,
            "Кем выдан": 130, "Адрес": 130, "Маршрут": 160, "Обоснование": 130,
            "ПС": 40, "АВИА/ЖД": 65, "Дата вылета": 90, "Примечание": 110,
            "Ответственный": 110, "Дата выписки": 85, "Билет": 75, "Сумма": 65,
            "Оплата": 75, "Причина возврата": 90, "Последний перелет": 90,
            "Телефон": 110, "Трансфер": 65,
        }
        self.excel_table = ExcelTable(
            left_frame, columns=ALL_COLUMNS, col_widths=col_widths,
            on_cell_edited=self._on_cell_edited)
        self.excel_table.pack(fill='both', expand=True, padx=5, pady=5)

        # Правая панель — просмотр PDF
        right_frame = ttk.Frame(main_paned)
        main_paned.add(right_frame, weight=3)
        self.pdf_viewer = PDFViewer(right_frame)

    def _show_login(self):
        login = LoginWindow(self.root, self.auth)
        self.root.wait_window(login.window)
        if self.auth.current_user:
            self.current_department = self.auth.current_user
            self.is_admin = self.auth.is_admin
            self._update_ui()
            self.status_label.config(text=f"Вход: {self.current_department}")
        else:
            self.root.quit()

    def _update_ui(self):
        if not self.current_department:
            return
        label = f"{self.current_department} (Admin)" if self.is_admin else self.current_department
        self.user_label.config(text=label)
        self.btn_load_base.config(state='normal' if self.is_admin else 'disabled')
        if get_employees_db() is not None:
            self.btn_catalog.config(state='normal')

    def _switch_user(self):
        self.auth.logout()
        self.current_department = None
        self.is_admin = False
        self.user_label.config(text="Не авторизован")
        self.btn_load_base.config(state='disabled')
        self._show_login()

    def _logout(self):
        if messagebox.askyesno("Выход", "Выйти из системы?"):
            self._on_closing()

    def _load_base(self):
        fp = filedialog.askopenfilename(
            title="Выберите файл базы", filetypes=[("Excel", "*.xlsx *.xls")])
        if not fp:
            return
        self.status_label.config(text="Загрузка базы...")
        self.root.update_idletasks()
        success, msg, count = load_employees_base(fp)
        if success:
            messagebox.showinfo("Успех", msg)
            self.status_label.config(text=f"База: {count} сотрудников")
            self.btn_catalog.config(state='normal')
            self._save_current_database()
        else:
            messagebox.showerror("Ошибка", msg)
        self._update_ui()

    def _open_catalog(self):
        def on_sel(emps):
            if not emps:
                return
            rows = []
            start_num = (len(self.applications_df) + 1
                         if self.applications_df is not None and not self.applications_df.empty
                         else 1)
            skipped = 0
            for i, emp in enumerate(emps):
                tn = emp.get('tab_num', '')
                if tn and tn in self.existing_tab_nums:
                    skipped += 1
                    continue
                if tn:
                    self.existing_tab_nums.add(tn)
                pdf_data = {
                    'source_file': 'Каталог', 'fio': emp.get('fio', ''),
                    'route': '', 'date': '', 'reason': '', 'phone': emp.get('phone', ''),
                }
                rows.append(create_application_row(start_num + i - skipped,
                                                   self.current_department, pdf_data, emp))
            if rows:
                df = pd.DataFrame(rows, columns=ALL_COLUMNS)
                self._append_to_table(df)

        EmployeeCatalogDialog(self.root, on_sel)

    def _select_single_file(self):
        if get_employees_db() is None:
            messagebox.showwarning("Внимание", "Сначала загрузите базу сотрудников!")
            return
        fp = filedialog.askopenfilename(title="Выберите PDF", filetypes=[("PDF", "*.pdf")])
        if fp:
            self._process_single_pdf(fp)

    def _select_folder(self):
        if get_employees_db() is None:
            messagebox.showwarning("Внимание", "Сначала загрузите базу сотрудников!")
            return
        fd = filedialog.askdirectory(title="Выберите папку с PDF")
        if fd:
            self._process_pdfs_folder(fd)

    def _process_single_pdf(self, fp):
        import os
        fn = os.path.basename(fp)
        self.pdf_files[fn] = fp
        self._show_progress("Обработка...")
        try:
            results = process_pdf_file(fp)
            self.pdf_results.extend(results)
            self._update_pdf_list()
            self.btn_fill.config(state='normal')
            if results:
                self.pdf_viewer.load_pdf(fp)
                self._fill_from_database()
            else:
                messagebox.showwarning("Внимание", "Не удалось извлечь данные из PDF")
        except Exception as e:
            logger.exception("Ошибка обработки PDF")
            messagebox.showerror("Ошибка", str(e))
        finally:
            self._hide_progress()

    def _process_pdfs_folder(self, folder):
        from pathlib import Path
        pdf_files = list(Path(folder).glob("*.pdf"))
        if not pdf_files:
            messagebox.showwarning("Внимание", "PDF-файлы не найдены")
            return

        self.progress_frame.pack(side='bottom', fill='x', padx=10, pady=5)
        self.progress_bar['maximum'] = len(pdf_files)
        try:
            for i, pf in enumerate(pdf_files):
                self.progress_bar['value'] = i
                self.progress_label.config(text=f"Обработка: {pf.name}")
                self.root.update_idletasks()
                results = process_pdf_file(str(pf))
                self.pdf_results.extend(results)
                self.pdf_files[pf.name] = str(pf)
            self._update_pdf_list()
            self.btn_fill.config(state='normal')
        except Exception as e:
            logger.exception("Ошибка обработки папки")
            messagebox.showerror("Ошибка", str(e))
        finally:
            self._hide_progress()

    def _update_pdf_list(self):
        self.pdf_listbox.delete(0, 'end')
        for r in self.pdf_results:
            self.pdf_listbox.insert('end', f"{r.get('source_file', '')} — {r.get('fio', '')}")

    def _on_pdf_select(self, event):
        sel = self.pdf_listbox.curselection()
        if sel and sel[0] < len(self.pdf_results):
            fn = self.pdf_results[sel[0]].get('source_file', '')
            if fn in self.pdf_files:
                self.pdf_viewer.load_pdf(self.pdf_files[fn])

    def _fill_from_database(self):
        if not self.pdf_results:
            return

        def on_complete(final_results):
            start_num = (len(self.applications_df) + 1
                         if self.applications_df is not None and not self.applications_df.empty
                         else 1)
            rows = []
            row_num = start_num
            for res in final_results:
                emp = res.get('employee')
                legs = res.get('legs')
                responsible = res.get('responsible', '')

                if legs:
                    for leg in legs:
                        pdf_data = {
                            'fio': res.get('fio', ''),
                            'phone': emp.get('phone', '') if emp else '',
                            'responsible': responsible,
                        }
                        if res.get('status') == 'skipped':
                            row = create_empty_row(row_num, self.current_department, {**pdf_data, **leg})
                        else:
                            row = create_application_row(row_num, self.current_department,
                                                         pdf_data, emp, leg)
                        rows.append(row)
                        tn = emp.get('tab_num', '') if emp else ''
                        if tn:
                            self.existing_tab_nums.add(tn)
                        row_num += 1
                else:
                    if res.get('status') == 'skipped':
                        row = create_empty_row(row_num, self.current_department, res)
                    else:
                        row = create_application_row(row_num, self.current_department, res, emp)
                    rows.append(row)
                    tn = emp.get('tab_num', '') if emp else ''
                    if tn:
                        self.existing_tab_nums.add(tn)
                    row_num += 1

            if rows:
                df = pd.DataFrame(rows, columns=ALL_COLUMNS)
                self._append_to_table(df)

        FillFromBaseWizard(
            self.root, self.pdf_results, self.pdf_files,
            self.current_department, on_complete
        )

    def _append_to_table(self, df):
        if self.applications_df is not None and not self.applications_df.empty:
            self.applications_df = pd.concat([self.applications_df, df], ignore_index=True)
        else:
            self.applications_df = df
        self.applications_df['№'] = range(1, len(self.applications_df) + 1)
        self.excel_table.load_dataframe(self.applications_df)
        self.btn_export.config(state='normal')
        self.btn_clear.config(state='normal')

    def _on_cell_edited(self, row_idx, col_name, new_value):
        if self.applications_df is not None and row_idx < len(self.applications_df):
            self.applications_df.at[self.applications_df.index[row_idx], col_name] = new_value

    def _clear_all(self):
        if not messagebox.askyesno("Подтверждение", "Очистить все заявки?"):
            return
        self.applications_df = None
        self.pdf_results = []
        self.pdf_files = {}
        self.existing_tab_nums.clear()
        self.excel_table.load_dataframe(None)
        self.pdf_listbox.delete(0, 'end')
        self.pdf_viewer.close()
        self.btn_fill.config(state='disabled')
        self.btn_export.config(state='disabled')
        self.btn_clear.config(state='disabled')

    def _export_to_excel(self):
        if self.applications_df is None or self.applications_df.empty:
            messagebox.showwarning("Внимание", "Нет данных для экспорта!")
            return

        default_name = f"НОВАЯ-{self.current_department}. Заказ билетов.xlsx"
        fp = filedialog.asksaveasfilename(
            title="Сохранить как", defaultextension=".xlsx",
            initialfile=default_name, filetypes=[("Excel", "*.xlsx")])
        if not fp:
            return

        try:
            from openpyxl import Workbook, load_workbook
            wb = load_workbook(fp) if os.path.exists(fp) else Workbook()

            for sheet_name in ("Заявка", "Sheet"):
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

            ws = wb.create_sheet("Заявка")
            for col_idx, col_name in enumerate(self.applications_df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(
                    self.applications_df.itertuples(index=False, name=None), 2):
                for col_idx, value in enumerate(row_data, 1):
                    val = value if pd.notna(value) else ""
                    if hasattr(val, 'strftime'):
                        val = val.strftime("%d.%m.%Y")
                    ws.cell(row=row_idx, column=col_idx, value=val)

            wb.save(fp)
            messagebox.showinfo("Успех", f"Данные сохранены:\n{fp}")
            logger.info("Экспорт в Excel: %s", fp)
        except Exception as e:
            logger.exception("Ошибка экспорта")
            messagebox.showerror("Ошибка", f"Ошибка сохранения:\n{e}")

    def _show_progress(self, text=""):
        self.progress_frame.pack(side='bottom', fill='x', padx=10, pady=5)
        self.progress_bar['value'] = 30
        self.progress_label.config(text=text)
        self.root.update()

    def _hide_progress(self):
        self.progress_frame.pack_forget()

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = MainApp()
    app.run()