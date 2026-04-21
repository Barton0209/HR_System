# ticket_app/excel_handler.py
import re
import logging
from datetime import datetime, timedelta
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import ALL_COLUMNS, COUNTRY_SCHEMA_MAP, PASSPORT_VALIDITY, ITR_KEYWORDS

logger = logging.getLogger(__name__)

IULIIA_AVAILABLE = False
try:
    import iuliia
    IULIIA_AVAILABLE = True
except ImportError:
    pass


def format_date_ddmmyyyy(date_val) -> str:
    if not date_val or str(date_val).strip() in ("", "nan", "None"):
        return ""
    date_str = str(date_val).strip()
    if re.match(r'^\d{2}\.\d{2}\.\d{4}$', date_str):
        return date_str
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d",
                "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    m = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{4})', date_str)
    if m:
        d, mo, y = m.groups()
        return f"{d.zfill(2)}.{mo.zfill(2)}.{y}"
    return date_str


def transliterate_name(name: str, citizenship: str = None) -> str:
    if not name:
        return ""
    schema = "gost_52535"
    if citizenship:
        for country, s in COUNTRY_SCHEMA_MAP.items():
            if country in citizenship.upper():
                schema = s
                break
    if IULIIA_AVAILABLE:
        try:
            return iuliia.translate(name, schema=iuliia.Schemas.get(schema)).upper()
        except Exception:
            pass
    table = {
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'E',
        'Ж': 'ZH', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
        'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
        'Ф': 'F', 'Х': 'KH', 'Ц': 'TS', 'Ч': 'CH', 'Ш': 'SH', 'Щ': 'SHCH',
        'Ы': 'Y', 'Э': 'E', 'Ю': 'YU', 'Я': 'YA', 'Ь': '', 'Ъ': '',
    }
    return ''.join(table.get(c, c) for c in str(name).upper())


def calculate_passport_expiry(doc_date: str, citizenship: str) -> str:
    if not doc_date or str(doc_date) in ("", "nan"):
        return ""
    cu = str(citizenship).upper().strip() if citizenship else ""
    if cu in ("РОССИЯ", "РФ"):
        return ""
    if cu == "БЕЛАРУСЬ":
        return "проверить паспорт"
    validity = PASSPORT_VALIDITY.get(cu)
    if not validity or not validity.get("months"):
        return ""
    try:
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(str(doc_date), fmt)
                break
            except ValueError:
                continue
        else:
            m = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{4})', str(doc_date))
            if not m:
                return ""
            dt = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        expiry = dt + timedelta(days=validity["months"] * 30 - 1)
        return expiry.strftime("%d.%m.%Y")
    except Exception:
        return ""


def get_classification(position: str) -> str:
    if not position:
        return "Рабочие"
    pos_lower = position.lower()
    return "ИТР" if any(kw in pos_lower for kw in ITR_KEYWORDS) else "Рабочие"


def get_document_type(citizenship: str, doc_series: str) -> str:
    cu = str(citizenship).upper().strip() if citizenship else ""
    if "РОССИЯ" in cu or cu == "РФ":
        return "Паспорт гражданина России"
    if str(doc_series).strip() in ("82", "83"):
        return "Вид на жительство"
    return "Паспорт иностранного гражданина"


def create_application_row(row_num: int, department: str, pdf_data: Dict,
                           employee: Dict = None) -> Dict:
    emp = employee or {}
    citizenship = emp.get('citizenship', '')
    doc_series = emp.get('doc_series', '')
    doc_date = emp.get('doc_date', '')
    doc_expiry = emp.get('doc_expiry', '')

    expiry = format_date_ddmmyyyy(doc_expiry) if doc_expiry else \
        calculate_passport_expiry(doc_date, citizenship)

    phone = pdf_data.get('phone', '') or emp.get('phone', '')

    return {
        "№": row_num,
        "Подразделение": department,
        "Отдел": str(emp.get('department_category', '')),
        "Операция": "Заказ",
        "Классификация": get_classification(emp.get('position', '')),
        "Дата заказа": datetime.now().strftime("%d.%m.%Y"),
        "Организация": "Монтаж",
        "Ф.И.О.": pdf_data.get('fio', ''),
        "Ф.И.О лат": transliterate_name(pdf_data.get('fio', ''), citizenship),
        "Табельный номер": emp.get('tab_num', ''),
        "Гражданство": citizenship,
        "Дата рождения": format_date_ddmmyyyy(emp.get('birth_date', '')),
        "Вид документа": get_document_type(citizenship, doc_series),
        "Серия": doc_series,
        "Номер": emp.get('doc_num', ''),
        "Дата выдачи": format_date_ddmmyyyy(doc_date),
        "Дата окончания": expiry,
        "Кем выдан": emp.get('doc_issuer', ''),
        "Адрес": str(emp.get('address', '')),
        "Маршрут": pdf_data.get('route', ''),
        "Обоснование": pdf_data.get('reason', ''),
        "ПС": "",
        "АВИА/ЖД": "АВИА",
        "Дата вылета": format_date_ddmmyyyy(pdf_data.get('date', '')),
        "Примечание": "",
        "Ответственный": "",
        "Дата выписки": "",
        "Билет": "",
        "Сумма": "",
        "Оплата": "Монтаж",
        "Причина возврата": "",
        "Последний перелет": "",
        "Телефон": phone,
        "Трансфер": "",
    }


def create_empty_row(row_num: int, department: str, pdf_data: Dict) -> Dict:
    row = create_application_row(row_num, department, pdf_data, None)
    row["Примечание"] = "НЕ НАЙДЕН В БАЗЕ"
    return row


def save_as_excel(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            val = value if pd.notna(value) else ""
            if hasattr(val, 'strftime'):
                val = val.strftime("%d.%m.%Y")
            ws.cell(row=row_idx, column=col_idx, value=val)

    for column in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in column if cell.value), default=0
        )
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_len + 2, 40)

    wb.save(output_path)
