import io
import logging
from datetime import datetime
from fastapi import APIRouter, Query, Response
from fastapi.responses import StreamingResponse
from hr_app.backend.database import (
    report_headcount_by_org, report_citizenship, report_hire_fire_dynamics,
    report_age_structure, report_expiring_docs, report_no_phone,
    report_duplicates, report_foreign_workers, query_employees,
    get_distinct_values
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/headcount")
def headcount(
    status_op: str = Query("ALL"),
    platform: str = Query("ALL"),
):
    rows = report_headcount_by_org(status_op, platform)
    return {"rows": rows}


@router.get("/citizenship")
def citizenship(
    status_op: str = Query("ALL"),
    platform: str = Query("ALL"),
):
    rows = report_citizenship(status_op, platform)
    return {"rows": rows}


@router.get("/dynamics")
def dynamics(
    status_op: str = Query("ALL"),
    platform: str = Query("ALL"),
):
    return report_hire_fire_dynamics(status_op, platform)


@router.get("/age-structure")
def age_structure(
    status_op: str = Query("ALL"),
    platform: str = Query("ALL"),
):
    rows = report_age_structure(status_op, platform)
    return {"rows": rows}


@router.get("/expiring-docs")
def expiring_docs(days: int = Query(30)):
    rows = report_expiring_docs(days)
    return {"rows": rows, "total": len(rows)}


@router.get("/no-phone")
def no_phone():
    rows = report_no_phone()
    return {"rows": rows, "total": len(rows)}


@router.get("/duplicates")
def duplicates():
    rows = report_duplicates()
    return {"rows": rows, "total": len(rows)}


@router.get("/foreign-workers")
def foreign_workers(
    status_op: str = Query("ALL"),
    platform: str = Query("ALL"),
):
    rows = report_foreign_workers(status_op, platform)
    return {"rows": rows, "total": len(rows)}


@router.get("/by-schedule")
def by_schedule(
    status_op: str = Query("ALL"),
    platform: str = Query("ALL"),
):
    from hr_app.backend.database import _build_filter, get_conn
    conds, params = _build_filter(status_op, platform)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT work_schedule, COUNT(*) as cnt FROM employees {where} "
            f"GROUP BY work_schedule ORDER BY cnt DESC", params
        ).fetchall()
    return {"rows": [dict(r) for r in rows]}


@router.get("/by-position")
def by_position(
    status_op: str = Query("ALL"),
    platform: str = Query("ALL"),
    limit: int = Query(20),
):
    from hr_app.backend.database import _build_filter, get_conn
    conds, params = _build_filter(status_op, platform)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT position, COUNT(*) as cnt FROM employees {where} "
            f"GROUP BY position ORDER BY cnt DESC LIMIT ?", params + [limit]
        ).fetchall()
    return {"rows": [dict(r) for r in rows]}


@router.get("/by-department")
def by_department(
    status_op: str = Query("ALL"),
    platform: str = Query("ALL"),
):
    from hr_app.backend.database import _build_filter, get_conn
    conds, params = _build_filter(status_op, platform)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT department, org, classification, COUNT(*) as cnt FROM employees {where} "
            f"GROUP BY department, org, classification ORDER BY cnt DESC LIMIT 300", params
        ).fetchall()
    return {"rows": [dict(r) for r in rows]}


@router.get("/org-tree")
def org_tree(
    status_op: str = Query("ALL"),
    platform: str = Query("ALL"),
):
    """Иерархия: Организация → Подразделение → Должность → Кол-во"""
    from hr_app.backend.database import _build_filter, get_conn
    conds, params = _build_filter(status_op, platform)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT org, department, position, COUNT(*) as cnt FROM employees {where} "
            f"GROUP BY org, department, position ORDER BY org, department, cnt DESC", params
        ).fetchall()

    tree = {}
    for r in rows:
        org = r["org"] or "—"
        dept = r["department"] or "—"
        pos = r["position"] or "—"
        cnt = r["cnt"]
        if org not in tree:
            tree[org] = {"name": org, "count": 0, "children": {}}
        tree[org]["count"] += cnt
        if dept not in tree[org]["children"]:
            tree[org]["children"][dept] = {"name": dept, "count": 0, "children": {}}
        tree[org]["children"][dept]["count"] += cnt
        tree[org]["children"][dept]["children"][pos] = cnt

    return {"tree": tree}


@router.get("/export-excel/{report_type}")
async def export_excel(report_type: str, status_op: str = "ALL", platform: str = "ALL"):
    """Экспорт любого отчёта в Excel с графиками."""
    import pandas as pd
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList

    data_map = {
        "headcount": lambda: report_headcount_by_org(status_op, platform),
        "citizenship": lambda: report_citizenship(status_op, platform),
        "age": lambda: report_age_structure(status_op, platform),
        "expiring": lambda: report_expiring_docs(30),
        "no-phone": lambda: report_no_phone(),
        "duplicates": lambda: report_duplicates(),
        "foreign": lambda: report_foreign_workers(status_op, platform),
        "dynamics": lambda: report_hire_fire_dynamics(status_op, platform).get("data", []),
    }

    if report_type not in data_map:
        return Response(status_code=404)

    rows = data_map[report_type]()
    if not rows:
        return Response(content="Нет данных для экспорта", status_code=400)

    df = pd.DataFrame(rows)
    
    # Создаем workbook с стилями
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    
    # Заголовок
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Добавляем заголовок отчёта
    report_titles = {
        "headcount": "Численность по организациям",
        "citizenship": "Гражданство сотрудников",
        "age": "Возрастная структура",
        "expiring": "Истекающие документы (30 дней)",
        "no-phone": "Сотрудники без телефона",
        "duplicates": "Дубликаты записей",
        "foreign": "Иностранные работники",
        "dynamics": "Динамика приёма/увольнения",
    }
    
    ws.merge_cells('A1:{}1'.format(chr(ord('A') + len(df.columns) - 1)))
    cell = ws['A1']
    cell.value = f"{report_titles.get(report_type, 'Отчёт')} · {datetime.now().strftime('%d.%m.%Y')}"
    cell.font = title_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    
    # Данные
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left" if isinstance(val, str) else "center")
    
    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Добавляем график если данных достаточно
    if len(df) >= 2 and len(df.columns) >= 2:
        chart_ws = wb.create_sheet(title="График")
        
        # Копируем данные для графика
        for r_idx, row in enumerate(dataframe_to_rows(df.iloc[:10], index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                chart_ws.cell(row=r_idx, column=c_idx, value=val)
        
        # Создаём столбчатую диаграмму
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.y_axis.title = "Количество"
        chart.x_axis.title = df.columns[0] if len(df.columns) > 0 else ""
        
        # Данные для графика
        data_ref = Reference(chart_ws, min_col=2, min_row=1, max_row=min(len(df), 10) + 1, max_col=2)
        cats_ref = Reference(chart_ws, min_col=1, min_row=2, max_row=min(len(df), 10) + 1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        
        chart_ws.add_chart(chart, "D2")
    
    # Сохраняем
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/export-all")
async def export_all_employees():
    """Экспорт всей базы сотрудников в Excel с фильтрами и группировкой."""
    import pandas as pd
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    employees = query_employees()
    if not employees:
        return Response(content="Нет данных", status_code=400)
    
    df = pd.DataFrame(employees)
    
    wb = Workbook()
    
    # Лист 1: Все данные
    ws1 = wb.active
    ws1.title = "Все сотрудники"
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
    
    # Заморозка шапки
    ws1.freeze_panes = "A2"
    
    # Автофильтр
    ws1.auto_filter.ref = ws1.dimensions
    
    # Лист 2: Сводная по подразделениям
    ws2 = wb.create_sheet("По подразделениям")
    if 'department' in df.columns and 'position' in df.columns:
        pivot = df.groupby(['department', 'position']).size().reset_index(name='count')
        for r_idx, row in enumerate(dataframe_to_rows(pivot, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=val)
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"employees_full_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
